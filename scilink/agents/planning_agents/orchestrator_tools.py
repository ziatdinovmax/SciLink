"""
Tool definitions and schemas for the PlanningOrchestratorAgent.
Supports both Google Gemini (function objects) and OpenAI (JSON schemas).
"""

from datetime import datetime
import json
from .planning_rag import (author_technical_document,
                           document_to_markdown)
import logging
import os
import re
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Callable, List, Optional
import hashlib


def _natural_sort_key(s):
    """Sort key that handles embedded numbers naturally (e.g., run_2 before run_10)."""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(s))]

from .parser_utils import write_experiments_to_disk
from .instruct import (
    BO_OBJECTIVE_DISTILL_PROMPT,
    KNOWLEDGE_QUERY_CODEGEN_PROMPT,
    KNOWLEDGE_QUERY_DIRECTORY_CODEGEN_PROMPT,
    SCREEN_DATABASE_CODEGEN_PROMPT,
)
from ..lit_agents.optimize_query import optimize_search_query, is_molecule_design_objective
from ...skills.loader import list_skills, load_skill


# Progress-heartbeat cadence for long-running literature searches (seconds).
_LIT_HEARTBEAT_SECONDS = 180
# Wall-clock ceiling for ONE batch of concurrent literature searches. Deep
# searches are advertised as 10-15 min; past this a straggler is abandoned
# and whatever finished is returned, rather than holding the call open.
# Kept ABOVE the agent's own per-task budget (1500s) so a stalled task
# normally times itself out first and its worker exits cleanly — this
# deadline is the backstop for when that does not happen.
_LIT_BATCH_DEADLINE = 1800

# Character budget for AUTO-LOADED multi-file literature context (issue
# #425). A campaign accumulates literature across searches; unioning
# without a cap turns the old silent-omission bug into a silent-truncation
# bug (a refine prompt was measured at ~83k tokens with literature already
# 85% of it). Applies ONLY when several files are unioned — a campaign's
# single literature file is always loaded whole, and explicitly passed
# literature_context is never capped. Over-budget WHOLE sections are
# dropped (never mid-file truncation: a half-severed review paragraph
# reads as complete) and the omission is logged.
_LIT_AUTOLOAD_MAX_CHARS = int(os.environ.get(
    "SCILINK_LIT_AUTOLOAD_MAX_CHARS", 400_000))

# The question-heading template is the structural contract between the
# literature write site and every reader that splits a saved corpus into
# per-question sections (auto-load budgeting, the list_literature_searches
# index, #q<N> section selection). Third-party search backends supply only
# the prose INSIDE a section; these boundaries are SciLink-authored, so
# writer and splitter must share one definition.
_LIT_QUESTION_RE = re.compile(r"^# Question (\d+): (.*)$", re.MULTILINE)

# '<path>#qN' — one question section of a saved literature file, the
# selection unit surfaced by list_literature_searches (issue #425).
_LIT_SECTION_REF_RE = re.compile(r"^(?P<base>.+?)#q(?P<n>\d+)$")


def _format_lit_question_heading(n: int, objective: str) -> str:
    """The one authoring point for '# Question N: <objective>' headings —
    must stay in lockstep with _LIT_QUESTION_RE."""
    return f"# Question {n}: {objective}"


def _build_planning_skill_description(custom_skills: dict = None) -> str:
    """Build the ``skill`` parameter description for ``generate_initial_plan``.

    Auto-discovers built-in planning skill bundles (and any custom or
    graduated skills) so the orchestrator LLM can see which skills are
    available by name instead of having to be told. Mirrors the analyze-mode
    ``_build_skill_description`` helper, scoped to ``domain="planning"``.
    """
    parts = [
        "Optional domain skill: a built-in planning skill name or a path to "
        "a custom .md skill file. When set, the skill's validated domain "
        "rules are injected as mandatory constraints on the generated plan."
    ]

    try:
        names = list_skills(domain="planning")
    except Exception:
        names = []

    skill_descs = []
    for name in names:
        try:
            parsed = load_skill(name, domain="planning")
            desc = (parsed.get("meta") or {}).get("description")
            if not desc:
                desc = parsed.get("overview", "").split("\n")[0].strip()
            # Trim trailing punctuation so the join below stays clean.
            desc = desc.rstrip(".;,") if desc else desc
            skill_descs.append(f"'{name}' — {desc}" if desc else f"'{name}'")
        except Exception:
            skill_descs.append(f"'{name}'")
    if skill_descs:
        parts.append(f"Built-in planning skills: {'; '.join(skill_descs)}.")

    if custom_skills:
        parts.append(f"Custom skills: {sorted(custom_skills.keys())}.")

    return " ".join(parts)


def _build_optimization_skill_description() -> str:
    """Build the ``skill`` parameter description for ``run_optimization``.

    Lists the available optimization skill bundles (``domain="optimization"``)
    so the orchestrator LLM can activate one by name. Mirrors the planning
    helper above, scoped to the optimization domain.
    """
    parts = [
        "Optional optimization skill: a built-in optimization skill name. When "
        "set, the skill's guidance is injected into the BO strategy and "
        "inspection stages, and any surrogate/acquisition the skill contributes "
        "becomes selectable. Pass when the user's problem matches a skill below."
    ]
    try:
        names = list_skills(domain="optimization")
    except Exception:
        names = []

    skill_descs = []
    for name in names:
        try:
            parsed = load_skill(name, domain="optimization")
            desc = (parsed.get("meta") or {}).get("description")
            if not desc:
                desc = parsed.get("overview", "").split("\n")[0].strip()
            desc = desc.rstrip(".;,") if desc else desc
            skill_descs.append(f"'{name}' — {desc}" if desc else f"'{name}'")
        except Exception:
            skill_descs.append(f"'{name}'")
    if skill_descs:
        parts.append(f"Available optimization skills: {'; '.join(skill_descs)}.")

    return " ".join(parts)


def resolve_n_candidates(requested, planner_state, new_campaign: bool = False) -> int:
    """
    Best-of-N default policy for ``generate_initial_plan`` (issue #377).

    An explicit request always wins (clamped 1-4; pass 1 for a single plan).
    When the caller omits it, a campaign's FIRST plan defaults to best-of-3 —
    two rounds of live meta testing showed the opt-in default never fired
    because upstream routing narrows the objective before delegating, so
    "raise N when open-ended" prompt guidance never triggered. Any later
    plan in the campaign defaults to 1: follow-ups iterate on a committed
    strategy, and the runner-up fallback already covers plan replacement.

    ``new_campaign`` marks a call that will START a new campaign in an
    existing session (issue #396) — its plan is a campaign-first plan even
    though ``current_plan`` still holds the previous campaign's plan.
    """
    if requested is not None:
        try:
            return max(1, min(int(requested), 4))
        except (TypeError, ValueError):
            return 1
    is_first_plan = new_campaign or not (planner_state or {}).get("current_plan")
    return 3 if is_first_plan else 1


class OrchestratorTools:
    """
    Manages tool definitions, schemas, and execution for the OrchestratorAgent.
    """
    
    def __init__(self, orchestrator_instance):
        """
        Args:
            orchestrator_instance: Reference to the parent OrchestratorAgent
        """
        self.orch = orchestrator_instance

        # Build function map and schemas
        self.functions_map: Dict[str, Callable] = {}
        self.openai_schemas: list = []
        self.gemini_functions: list = []

        self._register_all_tools()

    def _get_human_feedback_enabled(self) -> bool:
        """
        Get current human feedback setting from orchestrator.
        Returns True if not set (backwards compatible default).
        """
        return getattr(self.orch, '_enable_human_feedback', True)

    def _decode_categorical_recs(self, recs: Any, level_maps: Dict[str, List[str]]) -> Any:
        """Map integer-encoded categorical values back to their level names.

        Recommendations may be either a dict (single experiment) or a list of
        dicts (batch). Continuous values are passed through unchanged. The
        decoded value lookup uses the nearest integer index, which the
        MixedSingleTaskGP path already constrains to valid levels but other
        surrogates may return as a float.
        """
        if isinstance(recs, list):
            return [self._decode_categorical_recs(r, level_maps) for r in recs]
        if not isinstance(recs, dict):
            return recs
        out = dict(recs)
        for col, levels in level_maps.items():
            if col not in out:
                continue
            try:
                idx = int(round(float(out[col])))
            except (TypeError, ValueError):
                continue
            idx = max(0, min(idx, len(levels) - 1))
            out[col] = levels[idx]
        return out

    def _capture_input_types(self, column_roles: Dict, input_columns: List[str]) -> None:
        """Persist scalarizer input_types onto the orchestrator state.

        Filters to declared input columns only; missing entries default to
        "continuous" downstream. No-op when column_roles has no input_types
        field (backward-compat with older scalarizer outputs). Also captures the
        optional fidelity role (multi-fidelity) declared in the same column_roles.
        """
        # Capture the optional fidelity role first — it is independent of
        # input_types and must run even when input_types is absent.
        self._capture_fidelity_spec(column_roles, input_columns)
        if not input_columns:
            return
        types_in = (column_roles or {}).get("input_types") or {}
        if not types_in:
            return
        filtered = {c: types_in[c] for c in input_columns if c in types_in}
        if filtered:
            self.orch.expected_input_types = filtered

    def _capture_fidelity_spec(self, column_roles: Dict, input_columns: List[str]) -> None:
        """Persist an optional scalarizer-declared fidelity column (multi-fidelity).

        A fidelity column is a normal input that also indexes evaluation
        cost/accuracy. Sets ``orch.fidelity_spec`` to a validated
        {column, target_fidelity?, costs?} dict when the scalarizer declares one
        whose column is among the inputs; otherwise resets it to None (standard
        single-fidelity BO). Backward-compatible: older scalarizer outputs that
        omit the field reset to None.
        """
        spec = (column_roles or {}).get("fidelity")
        col = spec.get("column") if isinstance(spec, dict) else None
        if col and input_columns and col in input_columns:
            clean = {"column": col}
            if spec.get("target_fidelity") is not None:
                clean["target_fidelity"] = spec["target_fidelity"]
            if isinstance(spec.get("costs"), dict) and spec["costs"]:
                clean["costs"] = spec["costs"]
            self.orch.fidelity_spec = clean
            print(f"    📶  Fidelity axis declared: '{col}'"
                  + (f" (target={clean['target_fidelity']})" if "target_fidelity" in clean else ""))
        else:
            self.orch.fidelity_spec = None

    def _compute_file_hash(self, file_path: str) -> str:
        """Compute MD5 hash of file content for deduplication."""
        hasher = hashlib.md5()
        try:
            with open(file_path, 'rb') as f:
                for chunk in iter(lambda: f.read(8192), b''):
                    hasher.update(chunk)
            return hasher.hexdigest()
        except Exception as e:
            logging.warning(f"Could not compute hash for {file_path}: {e}")
            return ""


    def _distill_objective_for_bo(self, target_cols: list) -> str:
        """
        Distill a verbose user objective into a concise BO-relevant objective.
        Uses the orchestrator's LLM to extract only optimization targets and
        directions. Result is cached on self.orch._distilled_objective.
        """
        raw = self.orch.objective
        # Skip distillation for short/default objectives
        if (not raw
                or raw == "Undefined Research Goal"
                or len(raw) <= 200):
            return raw

        cached = getattr(self.orch, '_distilled_objective', None)
        if cached is not None:
            return cached

        try:
            prompt = BO_OBJECTIVE_DISTILL_PROMPT.format(
                objective=raw,
                target_cols=", ".join(target_cols),
            )
            resp = self.orch.bo.model.generate_content(
                [prompt], generation_config=self.orch.bo.generation_config
            )
            distilled = resp.text.strip()
            if distilled:
                print(f"    🎯 Distilled objective: {distilled}")
                self.orch._distilled_objective = distilled
                return distilled
        except Exception as e:
            logging.warning(f"Objective distillation failed, using original: {e}")

        return raw

    def _parse_result_input(self, result_data: str):
        """
        Helper to parse result_data into appropriate format.

        Returns:
            - String (text input)
            - String (single file path)
            - List of strings (multiple file paths)
        """
        if len(result_data) < 500:  # Reasonable path length
            try:
                # Check if it's comma-separated file paths
                if ',' in result_data:
                    paths = [p.strip() for p in result_data.split(',')]
                    valid_paths = []
                    for p in paths:
                        resolved, error = self._resolve_data_path(p)
                        if not error:
                            valid_paths.append(resolved)

                    if valid_paths:
                        print(f"    (Detected {len(valid_paths)} file paths)")
                        return valid_paths

                # Check if it's a single file path (try session-aware resolution)
                resolved, error = self._resolve_data_path(result_data.strip())
                if not error and Path(resolved).is_file():
                    print(f"    (Detected file path: {Path(resolved).name})")
                    return str(resolved)

                # Not a valid path - treat as text
                text_preview = result_data[:100] + "..." if len(result_data) > 100 else result_data
                print(f"    (Processing text input: '{text_preview}')")
                return result_data

            except (OSError, ValueError, RuntimeError):
                # Not a valid path - treat as text
                text_preview = result_data[:100] + "..." if len(result_data) > 100 else result_data
                print(f"    (Processing text input: '{text_preview}')")
                return result_data
        else:
            # Too long to be a path - treat as text
            text_preview = result_data[:100] + "..." if len(result_data) > 100 else result_data
            print(f"    (Processing text input: '{text_preview}')")
            return result_data
        
    def _collect_scalarizer_context(self, payload) -> list:
        """Collect scalarizer metrics and plot for files that were already analyzed.

        Checks whether any file path in *payload* has a corresponding entry in
        ``self.orch.analyzed_files``.  If so, appends:
        - The computed metrics from ``optimization_data.csv`` as a text summary
        - The scalarizer debug plot image path (if it exists)

        Returns a list of extra items to append to the refinement payload
        (may be empty).
        """
        paths = payload if isinstance(payload, list) else [payload]
        extras = []
        seen_bo = False

        for item in paths:
            if not isinstance(item, str):
                continue

            abs_path = str(Path(item).resolve())
            if abs_path not in self.orch.analyzed_files:
                continue

            # Append computed metrics from optimization_data.csv (once)
            if not seen_bo and self.orch.bo_data_path.exists():
                try:
                    df = pd.read_csv(self.orch.bo_data_path)
                    summary = (
                        f"SCALARIZER ANALYSIS RESULTS (computed metrics):\n"
                        f"{df.to_string(index=False)}"
                    )
                    extras.append(summary)
                    seen_bo = True
                    print(f"    📊 Attached scalarizer metrics ({len(df)} rows)")
                except Exception:
                    pass

            # Append debug plot if it exists
            stem = Path(item).stem
            plot_path = self.orch.base_dir / "scalarizer_outputs" / f"debug_{stem}.png"
            if plot_path.exists():
                extras.append(str(plot_path))
                print(f"    📈 Attached scalarizer plot: {plot_path.name}")

        return extras

    def _resolve_knowledge_paths(self, knowledge_paths: str | None) -> list[str] | None:
        """Resolve knowledge paths with fallback to orchestrator's knowledge_dir.

        If the LLM provides explicit paths, use those.  Otherwise fall back to
        ``self.orch.knowledge_dir`` so the KB can match sources from previous
        sessions (stable path) instead of rebuilding from a session-specific dir.
        """
        if knowledge_paths:
            paths = [p.strip() for p in knowledge_paths.split(",") if p.strip()]
            if paths:
                return paths
        # Fallback: use orchestrator's configured knowledge directory
        kd = self.orch.knowledge_dir
        if kd and kd.exists():
            # A store KB (marked by its manifest) is PREBUILT: its documents
            # live under sources/ and are already embedded — handing the KB
            # root to ingestion would re-embed them and swallow the index
            # files themselves. Point the source-difference check at the
            # sources/ dir (or nothing, for index-only imported KBs); plain
            # knowledge dirs keep the legacy incremental-ingest behavior.
            if (kd / "manifest.json").is_file():
                src = kd / "sources"
                return [str(src)] if src.is_dir() else None
            return [str(kd)]
        return None

    @staticmethod
    def _resolve_section_ref(s: str) -> Optional[str]:
        """Resolve a '<path>#qN' section reference (issue #425) to that
        question section's text. Returns None when ``s`` is not a section
        reference over an existing file; returns '' (resolved but empty)
        when the file exists but holds no question N, so the caller can
        warn instead of letting the ref string masquerade as raw text.
        """
        m = _LIT_SECTION_REF_RE.match(s)
        if not m:
            return None
        base = Path(m.group("base"))
        if not base.is_file():
            return None
        n = int(m.group("n"))
        text = base.read_text()
        sections = OrchestratorTools._split_literature_sections(text)
        for _q, chunk in sections:
            hm = _LIT_QUESTION_RE.match(chunk)
            if hm and int(hm.group(1)) == n:
                return chunk
        if n == 1 and len(sections) == 1:
            # A single-question corpus is written WITHOUT question
            # headings; its whole body is section 1 (live: the model
            # selected '<file>#q1' for exactly such a file, and a strict
            # resolver skipped the campaign's most relevant corpus).
            return text
        return ""

    @staticmethod
    def _resolve_context_text(value) -> Optional[str]:
        """Resolve a literature/molecule context argument to TEXT.

        Accepts a file path, a '<path>#qN' section reference, a
        comma-separated string of these, a list of them, or raw text.
        File/section contents are read and concatenated; anything that
        isn't resolvable as existing files is treated as raw text (the
        historical behavior). This closes a live failure where a
        comma-joined pair of paths fell through the single-path check and
        the PATH STRING itself became the 'literature', leaving downstream
        consumers (plan grounding, white-paper citations) with filenames
        instead of content.
        """
        if value is None:
            return None

        def _read_one(s: str) -> Optional[str]:
            """Text for one file path or section ref; None if neither."""
            sec = OrchestratorTools._resolve_section_ref(s)
            if sec is not None:
                if not sec:
                    print(f"    ⚠️  Section reference '{s}' names a "
                          f"question that file does not contain — skipped")
                return sec
            p = Path(s)
            return p.read_text() if p.is_file() else None

        items = value if isinstance(value, list) else [value]
        pieces = []
        for item in items:
            s = str(item).strip()
            if not s:
                continue
            one = _read_one(s)
            if one is not None:
                if one:
                    pieces.append(one)
                continue
            tokens = [t.strip() for t in s.split(",") if t.strip()]
            if len(tokens) > 1:
                texts = [_read_one(t) for t in tokens]
                if all(t is not None for t in texts):
                    pieces.extend(t for t in texts if t)
                    continue
            pieces.append(s)  # raw text
        return "\n\n".join(pieces) if pieces else None

    def _write_ideation_report(self) -> str:
        """Render ALL best-of-N candidates into a detailed markdown dossier.

        Deterministic (no LLM): rendered straight from the same campaign
        state the white paper is generated from, so the two artifacts cannot
        drift apart. In ideation, runner-up candidates are deliverables, not
        rejects — the judge's pick designates the flagship, it does not
        discard the rest.

        The flagship is rendered from ``current_plan``, not from the stored
        candidate: conformance correction and critic passes rewrite the
        selected plan *after* the candidate set is frozen, and rendering the
        stored copy shipped a dossier describing a plan the user never got
        (live: the dossier carried none of the corrected portfolio's
        directions while the white paper carried them all). Runner-ups are
        still rendered as authored — nothing rewrites those.
        """
        state = self.orch.planner.state or {}
        pc = state.get("plan_candidates") or {}
        candidates = pc.get("candidates") or []
        if not candidates:
            raise ValueError("No candidate set in state — nothing to report.")
        sel = pc.get("selected_index", 1)
        judge = pc.get("judge") or {}
        scores = {s.get("candidate"): s for s in judge.get("scores", [])}
        findings = (state.get("current_plan") or {}).get("critic_findings")
        override = pc.get("human_override")

        lines = ["# Ideation Report", "",
                 f"**Objective:** {state.get('objective', '')}", ""]
        if override:
            lines += [f"**Selection:** Candidate {sel} chosen by the PI, "
                      f"overriding the judge's pick of Candidate "
                      f"{judge.get('selected_candidate', '?')}.", ""]
        if judge.get("reasoning"):
            lines += ["## Comparative Assessment (judge)",
                      judge["reasoning"], ""]
        current = state.get("current_plan") or {}
        for ci, cand in enumerate(candidates, 1):
            exp = (cand.get("proposed_experiments") or [{}])[0]
            revised = False
            if ci == sel:
                flag = (" — SELECTED (flagship, PI override)" if override
                        else " — SELECTED (flagship)")
                cur_exp = (current.get("proposed_experiments") or [{}])[0]
                if cur_exp and cur_exp != exp:
                    exp = cur_exp          # as-shipped, not as-authored
                    revised = True
            else:
                flag = ""
            lines += [f"## Candidate {ci}{flag}: "
                      f"{exp.get('experiment_name', 'Untitled')}", ""]
            if revised:
                lines += ["*Shown as shipped: this flagship was revised after "
                          "selection (conformance / reviewer passes). The "
                          "judge's scores below refer to the version it "
                          "compared.*", ""]
            sc = scores.get(ci)
            if sc:
                crit = ", ".join(f"{k}: {v}" for k, v in sc.items()
                                 if k not in ("candidate", "comment"))
                lines += [f"*Judge scores — {crit}*",
                          f"*Judge comment: {sc.get('comment', '')}*", ""]
            concepts = exp.get("concepts")
            if isinstance(concepts, list) and concepts:
                lines.append(f"### Research directions ({len(concepts)})")
                for n, c in enumerate(concepts, 1):
                    if not isinstance(c, dict):
                        lines += [f"**{n}.** {c}", ""]
                        continue
                    from .user_interface import concept_title, humanize_key
                    tier = f" *(tier {c['tier']})*" if c.get("tier") else ""
                    lines.append(f"**{c.get('id') or n}. "
                                 f"{concept_title(c, n)}**{tier}")
                    for key in ("hypothesis", "rationale", "novelty"):
                        if c.get(key):
                            lines.append(f"- *{key.capitalize()}:* {c[key]}")
                    det = c.get("details")
                    for d in (det if isinstance(det, list) else [det] if det
                              else []):
                        lines.append(f"- {d}")
                    for k, v in c.items():
                        if k not in ("id", "tier", "title", "hypothesis",
                                     "rationale", "novelty", "details") and v:
                            if isinstance(v, list):
                                v = "; ".join(str(x) for x in v)
                            lines.append(f"- *{humanize_key(k)}:* {v}")
                    lines.append("")

            for key, title in (("hypothesis", "Hypothesis"),
                               ("experimental_steps",
                                "Shared protocol" if concepts
                                else "Proposed program"),
                               ("required_equipment", "Key capabilities"),
                               ("optimization_params",
                                "Suggested exploration variables"),
                               ("expected_outcome", "Expected outcomes"),
                               ("justification", "Rationale"),
                               ("source_documents", "Sources")):
                val = exp.get(key)
                if not val:
                    continue
                lines.append(f"### {title}")
                if isinstance(val, list):
                    for item in val:
                        if isinstance(item, dict):
                            lines.append("- " + ", ".join(
                                f"{k}: {v}" for k, v in item.items()))
                        else:
                            lines.append(f"- {item}")
                else:
                    lines.append(str(val))
                lines.append("")
            if ci == sel and findings:
                lines.append("### Reviewer caveats (on the flagship)")
                for f in findings:
                    lines.append(f"- [{f.get('severity', 'note')}] "
                                 + str(f.get('note') or f.get('finding')
                                       or f))
                lines.append("")

        path = self._output_dir() / "ideation_report.md"
        path.write_text("\n".join(lines))
        from .user_interface import format_path, record_deliverable
        record_deliverable(self.orch.base_dir, path,
                           "Ideation report — all candidate directions",
                           deliverable=True)
        print(f"    📄 Ideation report saved: {format_path(path)}")
        return str(path)

    def _planner_state(self):
        """Planner session state, or None when the planner has none (yet)."""
        return getattr(getattr(self.orch, "planner", None), "state", None)

    @property
    def _pending_lit(self) -> list:
        """Campaign-literature entries recorded before the planner has any
        session state (e.g. search_literature before the first plan).
        Folded into planner state on first access afterwards — see
        _lit_registry(). Lazily created so partially-constructed tools
        instances (tests build them via __new__) still work."""
        if getattr(self, "_prestate_lit", None) is None:
            self._prestate_lit = []
        return self._prestate_lit

    def _campaign_id(self) -> int:
        """Current campaign id from planner state (1 when unset/legacy)."""
        return int((self._planner_state() or {}).get("campaign_id") or 1)

    def _lit_registry(self) -> list:
        """The campaign-literature registry: [{'path', 'campaign_id'}, ...].

        Lives in planner state (so it rides session checkpoints and meta
        restores) once state exists; entries recorded before that are held
        on the tools instance and folded in on first access afterwards.
        ``campaign_id`` None marks a pending entry — literature saved
        before any campaign was active, claimed by the next plan call.
        """
        st = self._planner_state()
        if isinstance(st, dict) and st:
            reg = st.setdefault("campaign_literature", [])
            pending = self._pending_lit
            if pending:
                known = {(e.get("path"), e.get("campaign_id")) for e in reg}
                reg.extend(e for e in pending
                           if (e.get("path"), e.get("campaign_id")) not in known)
                self._prestate_lit = []
            return reg
        return self._pending_lit

    def _emit_plan_report(self, name: str = "plan.html", ideation=None):
        """Render the campaign's protocol report — unless this is a dossier.

        An ideation campaign gets no plan.html at all: the template renders
        an ordered experimental-steps protocol, which misrepresents a
        research portfolio. That suppression used to live on the
        initial-plan path only, while the four refinement paths kept
        regenerating the report — so a long ideation session (live: nine
        delegations of cdoc use-case ideation) accumulated protocol reports
        for a portfolio it had never planned. One helper, one rule.

        `ideation` may be passed when the caller has already decided; it
        defaults to asking the campaign. Returns the path, or None when
        suppressed.
        """
        if ideation is None:
            try:
                ideation = self.orch.planner._is_ideation_campaign()
            except Exception:  # noqa: BLE001
                ideation = False
        if ideation:
            # Now that the generator has a portfolio template, an ideation
            # campaign gets a real browser-readable report — under its own
            # name, because "plan.html" is what this whole change is getting
            # ideation out of. Not starred: the deliverables stay the dossier
            # and the white paper (_record_plan_report is lab-only).
            name = "portfolio.html" if name == "plan.html" else name
        from .html_generator import HTMLReportGenerator
        html_path = self._output_dir() / name
        HTMLReportGenerator(self.orch.planner.state).generate(str(html_path))
        self._record_plan_report(html_path)
        return html_path

    def _record_plan_report(self, html_path) -> None:
        """Mark a generated plan report as the campaign's deliverable —
        but only in LAB mode.

        Lab produces no white paper or dossier, so plan.html is what the
        user asked for. An IDEATION campaign is the opposite: its
        deliverables are the dossier and the white paper, and plan.html is
        suppressed at generation precisely because the protocol view
        misrepresents a portfolio. Refinement regenerates it anyway, so
        without this check a live ideation session starred three
        "Experimental plan (report)" files beside its real artifacts.
        """
        try:
            ideation = self.orch.planner._is_ideation_campaign()
        except Exception:  # noqa: BLE001
            ideation = False
        if ideation:
            return
        from .user_interface import record_deliverable
        record_deliverable(self.orch.base_dir, html_path,
                           "Experimental plan (report)", deliverable=True)

    def _record_literature_file(self, path, label: str = None,
                                questions: list = None) -> None:
        """Register a freshly saved literature file (issue #396).

        Tagged to the current campaign when one is active (a plan exists);
        otherwise left pending for the next plan call to claim.

        ``label`` (search-type coverage, e.g. 'hypothesis_context
        +cross_domain') and ``questions`` (the objectives the file
        answers) record CONTENT coverage on the registry entry (issue
        #425) so selection and indexing need not be made on filesystem
        accident (mtime) or re-parsing — older entries simply lack the
        fields and readers fall back to parsing the file.
        """
        st = self._planner_state() or {}
        cid = (int(st.get("campaign_id") or 1)
               if st.get("current_plan") else None)
        entry = {"path": str(Path(path).resolve()), "campaign_id": cid}
        if label:
            entry["label"] = label
        if questions:
            entry["questions"] = [str(q) for q in questions]
        self._lit_registry().append(entry)

    def _adopt_literature(self, explicit_context=None) -> None:
        """Claim literature for the CURRENT campaign: pending entries plus
        any file paths explicitly passed as literature_context. Called
        after a successful plan/refine so each campaign's corpus is exactly
        what it searched for or was explicitly given (issue #396)."""
        cid = self._campaign_id()
        reg = self._lit_registry()
        for e in reg:
            if e.get("campaign_id") is None:
                e["campaign_id"] = cid
        for p in self._context_file_paths(explicit_context):
            if not any(e.get("path") == p and e.get("campaign_id") == cid
                       for e in reg):
                reg.append({"path": p, "campaign_id": cid})
        # Persist immediately. The planner's own state dump is written from
        # inside generate_plan (after each _log_action), i.e. BEFORE this
        # runs, so without a re-save the on-disk mirror shows a literature
        # file still tagged to the previous campaign — a live session's
        # planning_state.json and checkpoint.json disagreed exactly here.
        saver = getattr(getattr(self.orch, "planner", None), "_save_state", None)
        if callable(saver):
            try:
                saver()
            except Exception as e:  # noqa: BLE001 - mirror only
                logging.debug(f"Planner state re-save after adoption failed: {e}")

    def _prior_campaign_literature(self, explicit_context) -> list:
        """File paths in ``explicit_context`` that the registry attributes
        ONLY to previous campaigns — candidate cross-topic contamination
        when the current call starts a new campaign."""
        cid = self._campaign_id()
        owners: Dict[str, set] = {}
        for e in self._lit_registry():
            owners.setdefault(str(e.get("path")), set()).add(e.get("campaign_id"))
        return [p for p in self._context_file_paths(explicit_context)
                if p in owners and cid not in owners[p]
                and None not in owners[p]]

    @staticmethod
    def _context_file_paths(value) -> list:
        """Existing-file paths named by a literature_context argument —
        mirrors _resolve_context_text's path resolution (single path,
        '<path>#qN' section refs resolving to their base file,
        comma-separated paths, or a list; raw text yields nothing)."""
        if value is None:
            return []

        def _one_path(s: str) -> Optional[str]:
            m = _LIT_SECTION_REF_RE.match(s)
            if m and Path(m.group("base")).is_file():
                return str(Path(m.group("base")).resolve())
            p = Path(s)
            return str(p.resolve()) if p.is_file() else None

        items = value if isinstance(value, list) else [value]
        out = []
        for item in items:
            s = str(item).strip()
            if not s:
                continue
            one = _one_path(s)
            if one is not None:
                out.append(one)
                continue
            tokens = [t.strip() for t in s.split(",") if t.strip()]
            if len(tokens) > 1:
                token_paths = [_one_path(t) for t in tokens]
                if all(tp is not None for tp in token_paths):
                    out.extend(token_paths)
        return out

    def _campaign_literature_files(self) -> List[Path]:
        """ALL literature files belonging to the CURRENT campaign, oldest
        first (issue #425).

        Campaign-scoped via the literature registry (issue #396): a session
        can hold several unrelated campaigns, and the old session-wide
        newest-file glob handed one campaign's corpus to another campaign's
        refine / white paper. Only same-campaign entries are eligible here;
        a campaign that supplied no literature gets an empty list — an
        honest miss, never another topic's corpus.

        Oldest-first because literature accumulates as a broad foundational
        search followed by narrow top-ups: foundation leads, top-ups
        append — the right reading order and the right truncation order.

        Legacy fallback: a session restored from before the registry
        existed has literature files on disk but no entries — for those,
        and only while no campaign transition has ever happened, fall back
        to the old session-wide glob (matched by GLOB, not exact name —
        multi-type searches save under labels like
        'literature_search_hypothesis_context+cross_domain.md'; recursive
        from base_dir because under the meta each delegation writes into
        its own delegations/<NN>_<slug>/).
        """
        cid = self._campaign_id()
        reg = self._lit_registry()
        seen: set = set()
        files = []
        for e in reg:
            if e.get("campaign_id") != cid or not e.get("path"):
                continue
            p = Path(e["path"])
            if str(p) not in seen and p.is_file():
                seen.add(str(p))
                files.append(p)
        if files:
            files.sort(key=lambda p: p.stat().st_mtime)
            return files
        if reg or cid > 1:
            return []
        roots = []
        base = getattr(self.orch, "base_dir", None)
        if base:
            roots.append((Path(base), "rglob"))
        roots.append((self._output_dir(), "glob"))
        legacy = []
        for root, mode in roots:
            it = (root.rglob("literature_search_*.md") if mode == "rglob"
                  else root.glob("literature_search_*.md"))
            for p in it:
                if str(p) not in seen:
                    seen.add(str(p))
                    legacy.append(p)
        legacy.sort(key=lambda p: p.stat().st_mtime)
        return legacy

    def _latest_literature_file(self) -> Optional[Path]:
        """Newest literature file of the CURRENT campaign (compat wrapper
        over _campaign_literature_files, which is oldest-first)."""
        files = self._campaign_literature_files()
        return files[-1] if files else None

    @staticmethod
    def _split_literature_sections(text: str) -> list:
        """Split a saved literature corpus into [(question, chunk), ...].

        Boundaries are the SciLink-authored '# Question N: <objective>'
        headings (_LIT_QUESTION_RE) — third-party search content only ever
        appears INSIDE a section, so a format change upstream cannot move
        them. A chunk includes its heading. Content before the first
        heading (the file title, or the entire body of a single-question
        file, which is written without question headings) is one chunk
        with question=None. Rejoining all chunks reproduces the input
        byte-for-byte.
        """
        matches = list(_LIT_QUESTION_RE.finditer(text))
        if not matches:
            return [(None, text)]
        sections = []
        if matches[0].start() > 0:
            sections.append((None, text[:matches[0].start()]))
        for i, m in enumerate(matches):
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
            sections.append((m.group(2).strip(), text[m.start():end]))
        return sections

    def _load_campaign_literature(self) -> Optional[Dict[str, Any]]:
        """Auto-load the CURRENT campaign's literature corpus (issue #425).

        Returns {'text', 'files', 'n_files', 'dropped'} or None when the
        campaign has no literature. Selection policy:

        - ONE file: loaded whole, verbatim, uncapped — there is no choice
          to make, and the single-search case is self-limiting. This path
          is byte-identical to the historical newest-file behavior.
        - SEVERAL files: oldest-first union. Verbatim-duplicate sections
          (identical content — e.g. a re-recorded file or a search re-run
          that returned the same text) are dropped first-occurrence-wins;
          different prose about the same question is NEVER fused. Then a
          character budget (_LIT_AUTOLOAD_MAX_CHARS) drops whole
          over-budget sections (oldest-first fill, never mid-section
          truncation), and every drop is reported — no silent caps.
        """
        files = self._campaign_literature_files()
        if not files:
            return None
        if len(files) == 1:
            text = files[0].read_text()
            return {"text": text, "files": [files[0].name],
                    "n_files": 1, "dropped": []}
        chunks = []  # (file_name, question, chunk_text)
        seen_content: set = set()
        dup_dropped = []
        for p in files:
            for question, chunk in self._split_literature_sections(
                    p.read_text()):
                key = chunk.strip()
                if key in seen_content:
                    dup_dropped.append((p.name, question))
                    continue
                seen_content.add(key)
                chunks.append((p.name, question, chunk))
        kept, dropped, total = [], [], 0
        for name, question, chunk in chunks:
            if kept and total + len(chunk) > _LIT_AUTOLOAD_MAX_CHARS:
                dropped.append((name, question))
                continue
            kept.append(chunk)
            total += len(chunk)
        if dup_dropped:
            print(f"    📚 Literature union: skipped "
                  f"{len(dup_dropped)} duplicate section(s) repeated "
                  f"verbatim across files")
        if dropped:
            what = "; ".join(
                f"{name}: {q or 'untitled section'}" for name, q in dropped)
            print(f"    ⚠️  Literature budget "
                  f"({_LIT_AUTOLOAD_MAX_CHARS:,} chars): dropped "
                  f"{len(dropped)} whole section(s) — {what}")
        return {"text": "\n\n".join(c.strip("\n") for c in kept),
                "files": [p.name for p in files],
                "n_files": len(files),
                "dropped": dropped}

    def _get_diagram_agent(self):
        """Lazy DiagramAgent sharing the BO agent's model client (vision-
        capable; no new credential plumbing)."""
        if getattr(self, "_diagram_agent", None) is None:
            from .diagram_agent import DiagramAgent
            self._diagram_agent = DiagramAgent(
                model=self.orch.bo.model, output_dir=str(self.orch.base_dir))
        return self._diagram_agent

    def _maybe_embed_workflow_diagram(self, text: str, out_dir,
                                      stem: str = "campaign_workflow") -> str:
        """Append a compact workflow diagram section to a document when the
        renderer is available and the QC'd diagram succeeds. Any failure
        returns the text unchanged — a document must never be lost to a
        figure."""
        try:
            from ...utils.mermaid_render import mermaid_available
            if not mermaid_available():
                return text
            plan = (self.orch.planner.state or {}).get("current_plan") or {}
            if not plan:
                # A roadmap or memo is often authored with no campaign
                # state at all (that is the point of the document tool),
                # so fall back to diagramming the document's own flow.
                if len(text or "") < 400:
                    return text
                plan = {"objective": "Document workflow",
                        "scientific_context": text[:4000]}
            res = self._get_diagram_agent().generate_workflow_diagram(
                plan=plan, out_dir=out_dir, stem=stem, detail="simple")
            if res.get("status") == "success":
                rel = Path(res["png_path"]).name
                print(f"    🗺️  Workflow diagram embedded "
                      f"({res['attempts']} attempt(s), "
                      f"{res['qc_rounds']} QC round(s))")
                # Recorded like the PDF twin — an ordinary produced file,
                # not a second starred deliverable: the document is what
                # the session points at. Without this the figure is
                # invisible to the files listing and to a UI resume,
                # which re-embeds from these manifests.
                try:
                    from .user_interface import record_deliverable
                    record_deliverable(self.orch.base_dir, res["png_path"],
                                       "Campaign workflow diagram")
                except Exception:  # noqa: BLE001
                    pass
                return (f"{text}\n\n## Campaign Workflow\n\n"
                        f"![Campaign workflow diagram]({rel})\n")
        except Exception as exc:  # noqa: BLE001
            print(f"    ⚠️  Workflow diagram skipped: {exc}")
        return text

    def _refresh_pdf_twin(self, md_path: Path) -> bool:
        """Re-export a markdown document's PDF twin, if one exists.

        Shared core in utils/file_edit (live: a diagram swap updated the
        white paper's markdown while its forwarded PDF kept the old
        image). Returns True when a twin was refreshed; failure must
        never block the edit that triggered it — this wrapper only owns
        the transcript prints.
        """
        from ...utils.file_edit import refresh_pdf_twin
        refreshed, err = refresh_pdf_twin(md_path)
        if refreshed:
            from .user_interface import format_path
            print(f"    📄 PDF twin refreshed: "
                  f"{format_path(md_path.with_suffix('.pdf'))}")
        elif err:
            print(f"    ⚠️  PDF twin not refreshed: {err}")
        return refreshed

    def _write_white_paper(self, audience_context: str = None) -> str:
        """Generate the sponsor-facing white paper from the current plan and
        save it beside the plan artifacts. Returns the saved path."""
        # Last-resort literature continuity: if neither the current plan nor
        # the CURRENT CAMPAIGN's history carries literature, seed the state
        # from the newest same-campaign saved search so citations survive
        # plan restructuring. Other campaigns' literature is invisible here
        # (issue #396): _latest_literature_file is campaign-scoped, and an
        # earlier campaign's corpus in history must not mask this one's
        # missing literature.
        state = self.orch.planner.state or {}
        cid = self._campaign_id()
        has_lit = any(p.get("literature_search")
                      for p in ([state.get("current_plan") or {}]
                                + list(state.get("plan_history") or []))
                      if int(p.get("campaign_id") or 1) == cid)
        if not has_lit:
            lit = self._load_campaign_literature()
            if lit is not None and state.get("current_plan"):
                state["current_plan"]["literature_search"] = lit["text"]
                print(f"    📚 White paper literature restored from "
                      f"{', '.join(lit['files'])}")
        text = self.orch.planner.generate_white_paper(
            audience_context=audience_context
        )
        text = self._maybe_embed_workflow_diagram(text, self._output_dir())
        wp_path = self._output_dir() / "white_paper.md"
        wp_path.write_text(text)
        from .user_interface import format_path, record_deliverable
        record_deliverable(self.orch.base_dir, wp_path,
                           "White paper", deliverable=True)
        print(f"    📄 White paper saved: {format_path(wp_path)}")
        # A PDF twin, because this is the one document that gets forwarded.
        # Recorded as an ordinary produced file, not a second deliverable —
        # the markdown stays the single thing the session points at. Failure
        # here must not lose the white paper that was just written.
        try:
            from ...utils.md_to_pdf import markdown_to_pdf
            pdf_path = markdown_to_pdf(wp_path, title="White paper")
            record_deliverable(self.orch.base_dir, pdf_path, "White paper PDF")
            print(f"    📄 PDF version: {format_path(pdf_path)}")
        except Exception as exc:  # noqa: BLE001
            print(f"    ⚠️  PDF version unavailable: {exc}")
        return str(wp_path)

    @staticmethod
    def _build_objective_guidance(n_data: int, numeric_cols: list) -> dict:
        """Return data-aware guidance on how many targets the data can support."""
        # Estimate max feasible inputs (assume at least 2)
        n_numeric = len(numeric_cols)
        est_inputs = max(2, n_numeric // 2)

        supported = {}
        for n_t in range(1, min(n_numeric, 5)):
            needed = 5 * est_inputs * n_t
            supported[n_t] = {"min_recommended": needed, "feasible": n_data >= needed}

        max_feasible = max((k for k, v in supported.items() if v["feasible"]), default=1)

        return {
            "data_points": n_data,
            "numeric_columns": n_numeric,
            "supported_targets": supported,
            "max_feasible_targets": max_feasible,
            "recommendation": (
                f"With {n_data} data points, up to {max_feasible} target(s) can be "
                f"optimized reliably. Pick the {max_feasible} most important target(s) "
                f"aligned with the stated objective. Additional targets can be added "
                f"later when more data is collected."
            ),
        }

    def _resolve_data_path(self, path_input: str) -> tuple[str, str]:
        """
        Resolves user input to actual file path with fuzzy matching for typos.
        
        Returns:
            (resolved_path, None) on success
            (None, error_json) on failure (with suggestions if available)
        """
        from difflib import get_close_matches
        
        path = Path(path_input.strip())
        
        # Case 1: Path exists as-is
        if path.exists():
            return str(path), None
        
        # Case 2: Try common extensions if no extension provided
        if not path.suffix:
            for ext in ['.csv', '.xlsx', '.xls']:
                candidate = path.with_suffix(ext)
                if candidate.exists():
                    print(f"    🔍 Resolved: {path.name} → {candidate.name}")
                    return str(candidate), None

        # Case 2b: a bare name that exists ANYWHERE in this session. Files
        # written by save_file land in the delegation directory that was
        # active at the time, so a later turn asking for them by name must
        # be able to find them without knowing which delegation wrote them.
        if not path.is_absolute() and len(path.parts) == 1:
            try:
                hits = sorted(self.orch.base_dir.rglob(path.name),
                              key=lambda p: p.stat().st_mtime, reverse=True)
                hits = [h for h in hits if h.is_file()]
                if hits:
                    print(f"    🔍 Resolved in session: {hits[0]}")
                    return str(hits[0]), None
            except Exception:  # noqa: BLE001 - fall through to the usual search
                pass
        
        # Case 3: Try in common data folders (session dirs first, then cwd-relative)
        session = self.orch.base_dir
        search_folders = [
            str(session / "uploads"),
            str(session / "uploads" / "series"),
            str(session / "data"),
            str(session),
            './experimental_results', './data', './results', './',
        ]
        all_candidates = []  # Track all files we find for fuzzy matching
        
        if not path.is_absolute():
            stem = path.stem if path.suffix else path.name
            
            for folder in search_folders:
                folder_path = Path(folder)
                if not folder_path.exists():
                    continue
                
                # Collect all data files in this folder
                for ext in ['.csv', '.xlsx', '.xls']:
                    all_candidates.extend(folder_path.glob(f"*{ext}"))
                
                # Try exact match with provided extension
                if path.suffix:
                    candidate = folder_path / path.name
                    if candidate.exists():
                        print(f"    🔍 Found: {path.name} in {folder}/")
                        return str(candidate), None
                
                # Try common extensions
                for ext in ['.csv', '.xlsx', '.xls']:
                    candidate = folder_path / f"{stem}{ext}"
                    if candidate.exists():
                        print(f"    🔍 Found: {stem}{ext} in {folder}/")
                        return str(candidate), None
        
        # Case 4: File not found - use fuzzy matching to suggest alternatives
        if all_candidates:
            # Get filenames without path
            candidate_names = [f.name for f in all_candidates]
            
            # Try fuzzy match on the input filename
            input_name = path.name
            matches = get_close_matches(input_name, candidate_names, n=3, cutoff=0.6)
            
            if matches:
                # Find full paths for the matches
                suggested_files = []
                for match in matches:
                    for candidate in all_candidates:
                        if candidate.name == match:
                            suggested_files.append(str(candidate))
                            break
                
                return None, json.dumps({
                    "status": "error",
                    "message": f"File not found: {path_input}",
                    "did_you_mean": matches,
                    "full_paths": suggested_files,
                    "hint": f"Did you mean '{matches[0]}'? Use: primary_data_set='{suggested_files[0]}'"
                })
        
        # No matches found at all
        return None, json.dumps({
            "status": "error",
            "message": f"Could not find file: {path_input}",
            "searched_in": [str(f) for f in search_folders if Path(f).exists()],
            "hint": "Check filename spelling or use /files command to see available files"
        })
    
    def _output_dir(self):
        """Directory for plan artifacts (plan.json, tea_analysis, output_scripts,
        literature_search*, molecule_design, ...).

        During a meta-agent delegation this is a per-delegation sub-directory
        so a reused planning child does not overwrite an earlier delegation's
        artifacts; for direct `scilink plan` use it is the campaign root
        (``_active_output_subdir`` is None, so behaviour is unchanged).

        Read defensively: file-writing tools now route through here, and a
        partially-built orchestrator (tests, early construction) has no
        delegation attribute yet — falling back to the campaign root is
        always the right answer there."""
        return getattr(self.orch, "_active_output_subdir", None) or self.orch.base_dir

    def _register_all_tools(self):
        """Register all tools with both OpenAI and Gemini formats."""
        
        # 0. LIST WORKSPACE FILES
        def list_workspace_files():
            """Lists files in the campaign directory including analysis artifacts."""
            print(f"  ⚡ Tool: Listing files in {self.orch.base_dir}...")
            files = [f.name for f in self.orch.base_dir.iterdir() if f.is_file()]
            artifacts_dir = self.orch.base_dir / "analysis_artifacts"
            artifact_names = []
            if artifacts_dir.exists():
                 artifact_names = [f"analysis_artifacts/{f.name}" for f in artifacts_dir.iterdir() if f.is_file()]
            
            all_files = files + artifact_names

            # Include data point count for optimization readiness
            data_count = 0
            if self.orch.bo_data_path.exists():
                try:
                    df = pd.read_csv(self.orch.bo_data_path)
                    data_count = len(df)
                except Exception:
                    pass

            # Under a meta session every artifact lives in
            # delegations/<slug>/, which the flat listing cannot see — live,
            # an agent probed six guessed paths for a companion whose real
            # filename it could not know (folders are named by task slug,
            # files by the authoring turn) and only found it by reading
            # deliverables.json as a last resort. Surface both halves of the
            # answer here: the registry (title -> path, the semantic index)
            # and a one-level listing of each delegation folder.
            deliverables_index = []
            try:
                from .user_interface import load_deliverables
                for e in load_deliverables(self.orch.base_dir):
                    p = e.get("path")
                    if not p or not Path(p).exists():
                        continue
                    deliverables_index.append({
                        "title": e.get("title") or Path(p).name,
                        "path": p,
                        "deliverable": bool(e.get("deliverable")),
                    })
            except Exception:  # noqa: BLE001 - listing must never fail
                pass

            delegation_folders = {}
            try:
                droot = self.orch.base_dir / "delegations"
                if droot.is_dir():
                    for d in sorted(droot.iterdir()):
                        if d.is_dir():
                            delegation_folders[d.name] = sorted(
                                f.name + ("/" if f.is_dir() else "")
                                for f in d.iterdir())
            except Exception:  # noqa: BLE001
                pass

            payload = {
                "status": "success",
                "files": all_files,
                "data_points_collected": data_count,
                "optimization_ready": data_count >= 3,
                "active_analysis_script": Path(self.orch.active_scalarizer_script).name if self.orch.active_scalarizer_script else None
            }
            if deliverables_index:
                payload["deliverables_index"] = deliverables_index
                payload["hint"] = ("deliverables_index maps document titles "
                                   "to their real paths — use it instead of "
                                   "guessing filenames.")
            if delegation_folders:
                payload["delegation_folders"] = delegation_folders
            return json.dumps(payload)

        self._register_tool(
            func=list_workspace_files,
            name="list_workspace_files",
            description=(
                "Lists files in the session directory (checkpoints, analysis "
                "artifacts, etc.), plus — when present — a deliverables index "
                "(document title -> real path) and the contents of each "
                "delegations/ folder. To find a document produced earlier, "
                "use this index rather than guessing paths: delegation "
                "folders are named by task slug and filenames rarely match "
                "a document's topic. User data files may exist outside the "
                "session folder."
            ),
            parameters={}
        )
        
        # --- LITERATURE SEARCH TOOL ---
        def search_literature(objective, search_type: str = "hypothesis_context"):
            """
            Searches scientific literature using the FutureHouse Edison API.
            Call this BEFORE generate_initial_plan to enrich the plan with
            external literature context.
            """
            if not self.orch.lit_agent:
                return json.dumps({
                    "status": "error",
                    "message": "Literature search not available (no FutureHouse API key configured)"
                })

            search_methods = {
                "hypothesis_context": self.orch.lit_agent.search_for_hypothesis_context,
                "cross_domain": self.orch.lit_agent.search_for_cross_domain,
                "economic_data": self.orch.lit_agent.search_for_economic_data,
                "fitting_models": self.orch.lit_agent.search_for_fitting_models,
            }
            # Multiple types — and multiple decomposed objectives — run
            # CONCURRENTLY: each Edison call is minutes of waiting, so a
            # serial pair doubles the user's wait for no reason.
            # PER-TYPE objectives. `objective` may be a mapping
            # {search_type: question(s)} — because a paired call otherwise
            # cross-multiplies ONE text across every type, and the two legs
            # want different text: grounding wants "what are the gaps in X",
            # transfer wants the FUNCTION to transfer toward. A plain string
            # or list keeps the historical cross-product.
            per_type: Dict[str, List[str]] = {}
            if isinstance(objective, dict):
                for t, objs in objective.items():
                    key = str(t).strip()
                    vals = ([objs] if isinstance(objs, str)
                            else [str(o) for o in (objs or [])])
                    vals = [v.strip() for v in vals if v and v.strip()]
                    if vals:
                        per_type[key] = vals
                types = list(per_type)
            else:
                types = [t.strip() for t in str(search_type).split(",")
                         if t.strip()]
            bad = [t for t in types if t not in search_methods]
            if bad or not types:
                return json.dumps({
                    "status": "error",
                    "message": (f"Invalid search_type '{bad or search_type}'. "
                                f"Use one or more (comma-separated, or as the "
                                f"keys of an objective mapping) of: "
                                f"{', '.join(search_methods)}")
                })

            if per_type:
                # objective index -> text, deduped so an identical question
                # asked of two types is optimized once.
                objectives = list(dict.fromkeys(
                    o for objs in per_type.values() for o in objs))
                idx = {o: i for i, o in enumerate(objectives)}
                task_pairs = [(idx[o], t) for t in types for o in per_type[t]]
            else:
                objectives = ([objective] if isinstance(objective, str)
                              else [str(o) for o in (objective or [])])
                objectives = [o.strip() for o in objectives if o and o.strip()]
                task_pairs = None
            if not objectives:
                return json.dumps({"status": "error",
                                   "message": "No objective provided."})
            # Concurrency is bounded by BATCHING inside this single call —
            # up to MAX_CONCURRENT searches run in parallel, extras spill
            # into the next back-to-back batch. This fills the parallel
            # budget before going sequential, and keeps everything in ONE
            # tool call so the LLM never splits across turns (which would
            # serialize). MAX_TOTAL is a wall-clock guardrail (each batch is
            # ~10-15 min), not a concurrency limit.
            MAX_CONCURRENT = 6
            MAX_TOTAL = 12
            n_jobs = (len(task_pairs) if task_pairs is not None
                      else len(objectives) * len(types))
            if n_jobs > MAX_TOTAL:
                return json.dumps({
                    "status": "error",
                    "message": (
                        f"{n_jobs} searches requested, over the "
                        f"{MAX_TOTAL}-per-"
                        f"call ceiling (~{-(-n_jobs // MAX_CONCURRENT)} "
                        f"sequential batches of {MAX_CONCURRENT} = too long). "
                        f"Prioritize down to <= {MAX_TOTAL} searches in this "
                        f"one call — do NOT split into separate calls."),
                })

            label = "+".join(types)
            n_batches = -(-n_jobs // MAX_CONCURRENT)
            # The Q-list below shows DISTINCT questions; when several search
            # types are active each question runs once per type, so say that
            # multiplication out loud — '5 questions but 10 searches' read
            # as a mismatch in live sessions.
            _q = f"{len(objectives)} question{'s' if len(objectives) != 1 else ''}"
            if per_type:
                # Each type has its OWN question(s); the multiplication line
                # would be a lie here.
                _mapping = (", ".join(
                    f"{len(per_type[t])} for {t}" for t in types)
                    + f" = {n_jobs} searches")
            elif len(types) > 1:
                _mapping = (f"{_q}, each searched {len(types)} ways "
                            f"({', '.join(types)}) = {n_jobs} searches")
            else:
                _mapping = (f"{_q} x 1 search type ({label}) = {n_jobs} "
                            f"search{'es' if n_jobs != 1 else ''}")
            # Spell out the actual batch sizes — 'running 6 at a time in 2
            # batches' reads as 6 x 2 = 12 when the last batch only carries
            # the remainder.
            _batch_sizes = [min(MAX_CONCURRENT, n_jobs - s)
                            for s in range(0, n_jobs, MAX_CONCURRENT)]
            _batch_note = (f", run in {n_batches} sequential batches of "
                           + " then ".join(str(b) for b in _batch_sizes)
                           if n_batches > 1 else "")
            print(f"  ⚡ Tool: Searching literature: {_mapping}{_batch_note}")
            for qi, o in enumerate(objectives, 1):
                _for = ""
                if per_type:
                    _owners = [t for t in types if o in per_type[t]]
                    _for = f" [{'+'.join(_owners)}]"
                print(f"     Q{qi}{_for}: '{o[:90]}"
                      f"{'...' if len(o) > 90 else ''}'")
            if n_batches > 1:
                print(f"     ⏱️  Deep literature searches typically take "
                      f"10-15 minutes per batch — {n_batches} sequential "
                      f"batches, so expect ~{10 * n_batches}-"
                      f"{15 * n_batches} minutes total; progress is "
                      f"reported every few minutes.")
            else:
                print("     ⏱️  Deep literature searches typically take 10-15 "
                      "minutes — the system is working; progress is reported "
                      "every few minutes.")

            # Heartbeat so a minutes-long silent wait doesn't read as a hang
            # (Edison jobs produce no output until they complete). Tracks LIVE
            # completion — '1 of 5 still running', not a static total — via
            # per-future done-callbacks updating the shared pending set.
            import threading as _threading
            import time as _time
            _hb_stop = _threading.Event()
            _hb_lock = _threading.Lock()
            # `t0` is PER BATCH, not per call. Batches run back-to-back and
            # each gets its own deadline, so a global clock made batch 2's
            # first minute read as "17 min elapsed — longer than the usual
            # 10-15" and threatened a give-up time that had already passed.
            _hb_state = {"pending": None, "running": set(), "total": n_jobs,
                         "t0": _time.time(), "batch": 0, "n_batches": 1}

            def _hb_mark_running(label):
                with _hb_lock:
                    _hb_state["running"].add(label)

            def _hb_mark_done(label):
                with _hb_lock:
                    if _hb_state["pending"] is not None:
                        _hb_state["pending"].discard(label)
                    _hb_state["running"].discard(label)

            def _heartbeat():
                # Batches are sequential, so 'remaining' includes jobs that
                # have not been SUBMITTED yet — report running and queued
                # separately ('10 still running' overstated concurrency in
                # live multi-batch sessions).
                _eta = ("typically 10-15 min total" if n_batches == 1
                        else "~10-15 min per batch")
                while not _hb_stop.wait(_LIT_HEARTBEAT_SECONDS):
                    with _hb_lock:
                        pending = (set(_hb_state["pending"])
                                   if _hb_state["pending"] is not None
                                   else None)
                        running = len(_hb_state["running"])
                    remaining = (len(pending) if pending is not None
                                 else _hb_state["total"])
                    if remaining == 0:
                        continue
                    queued = max(0, remaining - running)
                    with _hb_lock:
                        _t0 = _hb_state["t0"]
                        _bi, _bn = _hb_state["batch"], _hb_state["n_batches"]
                    mins = int((_time.time() - _t0) / 60)
                    _of = f" [batch {_bi} of {_bn}]" if _bn > 1 else ""
                    # Past the advertised window, stop repeating it: saying
                    # "typically 10-15 min" at minute 45 reads as a hang
                    # with no end in sight. Say when we will give up.
                    if mins >= 16:
                        give_up = _LIT_BATCH_DEADLINE // 60
                        print(f"  ⏳ {running} still running after {mins} min{_of} "
                              f"— longer than the usual 10-15; abandoning "
                              f"stragglers at {give_up} min and returning "
                              f"whatever finished")
                        continue
                    if queued:
                        print(f"  ⏳ {running} running, {queued} queued of "
                              f"{_hb_state['total']} literature searches "
                              f"({mins} min elapsed{_of}; {_eta})")
                    else:
                        print(f"  ⏳ {remaining} of {_hb_state['total']} "
                              f"literature search"
                              f"{'es' if _hb_state['total'] != 1 else ''} still "
                              f"running ({mins} min elapsed{_of}; {_eta})")

            _threading.Thread(target=_heartbeat, daemon=True).start()

            try:
                clean_queries = [optimize_search_query(
                    objective=o, model=self.orch.planner.model
                ) for o in objectives]

                tasks = (task_pairs if task_pairs is not None
                         else [(oi, t) for oi in range(len(objectives))
                               for t in types])

                def _task_label(oi, t):
                    return f"q{oi + 1}:{t}" if len(objectives) > 1 else t

                with _hb_lock:
                    _hb_state["pending"] = {_task_label(oi, t)
                                            for oi, t in tasks}
                results = {}
                if len(tasks) == 1:
                    oi, t = tasks[0]
                    _hb_mark_running(_task_label(oi, t))
                    results[(oi, t)] = search_methods[t](clean_queries[oi])
                    _hb_mark_done(_task_label(oi, t))
                else:
                    from concurrent.futures import (
                        ThreadPoolExecutor, TimeoutError as FuturesTimeout)
                    # Greedy batches of <= MAX_CONCURRENT, back-to-back. A
                    # 5-search request is ONE parallel batch; an 8-search
                    # request is 6 concurrent then 2 concurrent — never 2+3
                    # sub-optimal splits, and never sequential when it fits
                    # the parallel budget.
                    _n_batches = -(-len(tasks) // MAX_CONCURRENT)
                    with _hb_lock:
                        _hb_state["n_batches"] = _n_batches
                    for start in range(0, len(tasks), MAX_CONCURRENT):
                        batch = tasks[start:start + MAX_CONCURRENT]
                        # Fresh clock: this batch's deadline starts here, so
                        # the elapsed time reported against it must too.
                        with _hb_lock:
                            _hb_state["t0"] = _time.time()
                            _hb_state["batch"] = start // MAX_CONCURRENT + 1
                        # NOT a context manager: its __exit__ joins every
                        # worker, so one wedged remote task would still hold
                        # the call open after the deadline below fires.
                        ex = ThreadPoolExecutor(max_workers=len(batch))
                        try:
                            futures = {}
                            for oi, t in batch:
                                _l = _task_label(oi, t)
                                _hb_mark_running(_l)
                                f = ex.submit(search_methods[t],
                                              clean_queries[oi])
                                f.add_done_callback(
                                    lambda _f, _l=_l: _hb_mark_done(_l))
                                futures[(oi, t)] = f
                            # Bounded wait. One wedged remote task used to
                            # pin the whole call (live: 45+ min on a
                            # 5-search request) while four finished results
                            # sat unusable behind it. Past the deadline the
                            # stragglers are abandoned and reported as
                            # failures — the merge below already handles a
                            # partial result set.
                            deadline = _time.time() + _LIT_BATCH_DEADLINE
                            for k, f in futures.items():
                                remaining = deadline - _time.time()
                                try:
                                    results[k] = f.result(
                                        timeout=max(0.0, remaining))
                                except FuturesTimeout:
                                    label = _task_label(*k)
                                    _hb_mark_done(label)
                                    print(f"  ⚠️  {label} exceeded "
                                          f"{_LIT_BATCH_DEADLINE // 60} min "
                                          f"— abandoning it and returning "
                                          f"the searches that finished.")
                                    results[k] = {
                                        "status": "timeout",
                                        "message": (
                                            f"abandoned after "
                                            f"{_LIT_BATCH_DEADLINE}s"),
                                    }
                                except Exception as e:  # noqa: BLE001
                                    results[k] = {"status": "error",
                                                  "message": str(e)}
                        finally:
                            # Do not join: an abandoned worker keeps polling
                            # until the agent's own max_wait_time ends it.
                            ex.shutdown(wait=False, cancel_futures=True)

                ok = {k: r for k, r in results.items()
                      if r.get("status") == "success"}
                if not ok:
                    first = next(iter(results.values()))
                    return json.dumps({
                        "status": first.get("status", "error"),
                        "message": first.get("message", "Literature search did not succeed")
                    })

                # Sections are LABELLED, not concatenated: candidates must be
                # able to tell established results in this field (usable as
                # constraints) from cross-domain analogies (usable as
                # mechanism inspiration, but not established here).
                SECTION = {
                    "hypothesis_context": "## ESTABLISHED IN THIS FIELD (known methods, "
                                          "parameter ranges, failure modes)",
                    "cross_domain": "## TRANSFERABLE MECHANISMS FROM OTHER DOMAINS "
                                    "(analogies — NOT established results in this field)",
                    "economic_data": "## ECONOMIC CONTEXT",
                    "fitting_models": "## MODELS AND EQUATIONS",
                }
                parts = []
                for oi in range(len(objectives)):
                    secs = [f"{SECTION.get(t, '## ' + t.upper())}\n{ok[(oi, t)]['content']}"
                            for t in types if (oi, t) in ok]
                    if not secs:
                        continue
                    if len(objectives) > 1:
                        # Multi-question runs keep per-question grouping so
                        # downstream readers see which evidence answers what.
                        parts.append(
                            _format_lit_question_heading(oi + 1,
                                                         objectives[oi])
                            + "\n\n" + "\n\n".join(secs))
                    else:
                        parts.extend(secs)
                content = "\n\n".join(parts)

                # The name is keyed only by search TYPE, so a second search
                # of the same type in one output directory used to truncate
                # the first — silently, since the registry then held two
                # entries pointing at one surviving file. Under the meta
                # each delegation has its own directory, so this bites the
                # standalone campaign root hardest. Suffix only on
                # collision: the first (overwhelmingly common) write keeps
                # the historical name exactly.
                lit_path = self._output_dir() / f"literature_search_{label}.md"
                _n = 1
                while lit_path.exists():
                    _n += 1
                    lit_path = (self._output_dir()
                                / f"literature_search_{label}_{_n}.md")
                with open(lit_path, 'w') as f:
                    f.write(f"# Literature Search Results ({label})\n\n")
                    f.write(content)
                self._record_literature_file(
                    lit_path, label=label,
                    questions=[objectives[oi]
                               for oi in range(len(objectives))
                               if any((oi, t) in ok for t in types)])

                failed = [(f"q{oi + 1}:{t}" if len(objectives) > 1 else t)
                          for oi, t in tasks if (oi, t) not in ok]
                print(f"  ✅ Literature search completed ({len(ok)}/{len(tasks)}). "
                      f"Saved to {lit_path.name}")

                out = {
                    "status": "success",
                    "searches_run": [
                        (f"q{oi + 1}:{t}" if len(objectives) > 1 else t)
                        for oi, t in tasks if (oi, t) in ok
                    ],
                    "file_path": str(lit_path),
                    "content_preview": content[:500] + "..." if len(content) > 500 else content,
                    "hint": "Pass file_path as literature_context to generate_initial_plan()",
                }
                if failed:
                    out["failed_searches"] = failed
                if ok:
                    out["caveat"] = ("Literature context of any type was measured to "
                                     "reduce constraint COVERAGE — plans stay on-topic "
                                     "but silently omit individual requirements. When "
                                     "the objective carries hard equipment/process "
                                     "constraints, pass them as additional_context so "
                                     "each is mapped to a named step; do not withhold "
                                     "the literature.")
                return json.dumps(out)

            except Exception as e:
                logging.error(f"Literature search error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})
            finally:
                _hb_stop.set()

        self._register_tool(
            func=search_literature,
            name="search_literature",
            description=(
                "Searches scientific literature via FutureHouse Edison API. "
                "A deep search takes ~10-15 minutes — tell the user before "
                "calling so the wait is expected. "
                "Call BEFORE generate_initial_plan() to enrich the plan with external context. "
                "Pass the returned file_path as literature_context to generate_initial_plan(). "
                "DECOMPOSE, don't concatenate: each objective must be ONE "
                "focused question — a query stuffed with several distinct "
                "aspects returns a shallow synthesis of all of them. Pass a "
                "LIST of objectives (and, if useful, multiple search types) "
                "in a SINGLE call — put ALL the aspects you want covered "
                "into this one call. The tool runs up to 6 searches "
                "concurrently and automatically batches any extras into "
                "back-to-back rounds, merging everything into one labelled "
                "document. So do NOT make a second search_literature call to "
                "cover more aspects (a second call is a later turn and runs "
                "sequentially anyway) — just list them all here. Total "
                "searches must be <= 12. When pairing grounding with "
                "cross_domain, give each type its OWN objective via the "
                "object form (see `objective`) rather than one shared "
                "question."
            ),
            parameters={
                "objective": {
                    "type": ["string", "array", "object"],
                    "items": {"type": "string"},
                    "description": (
                        "ONE focused research question, or a LIST of them "
                        "(run in parallel). Split a broad objective — or "
                        "one containing too many distinct aspects — into "
                        "separate entries; each entry should stand alone as "
                        "a question one review could answer well. "
                        "PER-TYPE form: pass an OBJECT keyed by search type, "
                        "e.g. {\"hypothesis_context\": [\"what is known "
                        "about X\"], \"cross_domain\": [\"capture a state "
                        "that exists only under drive and relaxes in "
                        "milliseconds\"]}. Use it whenever you pair the two: "
                        "grounding wants the question about your field, "
                        "transfer wants the FUNCTION to transfer toward, and "
                        "a single shared string cannot be both. With the "
                        "object form, search_type is taken from its keys."
                    ),
                },
                "search_type": {
                    "type": "string",
                    "description": (
                        "One type, or several comma-separated (run in parallel). "
                        "'hypothesis_context' (default): established methods, "
                        "parameter ranges and pitfalls in the problem's own field — "
                        "the grounding a runnable plan needs. "
                        "'cross_domain': mechanisms from ADJACENT/UNRELATED fields "
                        "that could transfer — use for IDEATION (pair it with "
                        "hypothesis_context: 'hypothesis_context,cross_domain'). "
                        "cross_domain needs a FUNCTION or CHALLENGE to transfer "
                        "toward, so phrase that objective as the thing to achieve "
                        "or overcome; a survey question ('what are the frontiers "
                        "in X', 'what is hard to measure in X') asks for X's own "
                        "field and belongs to hypothesis_context — sending it to "
                        "cross_domain asks for a review and forbids it in the "
                        "same breath. "
                        "'economic_data' (TEA); 'fitting_models' "
                        "(curve fitting)."
                    ),
                }
            },
            required=["objective"]
        )

        # --- LITERATURE INDEX TOOL (issue #425) ---
        def list_literature_searches():
            """
            Index of the CURRENT campaign's saved literature searches:
            every file, the question each section answers, and the
            beginning of each answer — the information needed to DECIDE
            what a refine / plan call should read, without loading
            hundreds of KB to find out.
            """
            print("  ⚡ Tool: Listing campaign literature searches...")
            try:
                files = self._campaign_literature_files()
                if not files:
                    return json.dumps({
                        "status": "success", "count": 0, "files": [],
                        "message": ("No literature saved for the current "
                                    "campaign. Use search_literature() to "
                                    "gather some.")})
                meta = {}
                for e in self._lit_registry():
                    if e.get("path"):
                        meta[str(Path(e["path"]).resolve())] = e
                out_files = []
                for p in files:
                    text = p.read_text()
                    entry = meta.get(str(p.resolve()), {})
                    reg_qs = entry.get("questions") or []
                    chunks = self._split_literature_sections(text)
                    sections = []
                    for question, chunk in chunks:
                        m = _LIT_QUESTION_RE.match(chunk)
                        body = (chunk[m.end():] if m else chunk).strip()
                        if (question is None and len(body) < 200
                                and body.startswith(
                                    "# Literature Search Results")):
                            continue  # bare title chunk, not content
                        if question is None and len(reg_qs) == 1:
                            # Single-question files are written without a
                            # question heading; the registry knows the
                            # objective.
                            question = reg_qs[0]
                        item = {"question": question,
                                "chars": len(chunk),
                                "answer_preview": body[:300]}
                        # Headingless single-question files are addressable
                        # as their one-and-only section, '#q1' — keep the
                        # selection syntax uniform across file shapes.
                        if m:
                            item["section_ref"] = f"{p}#q{m.group(1)}"
                        elif len(chunks) == 1:
                            item["section_ref"] = f"{p}#q1"
                        sections.append(item)
                    out_files.append({
                        "path": str(p),
                        "label": entry.get("label"),
                        "chars": len(text),
                        "modified": datetime.fromtimestamp(
                            p.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                        "sections": sections,
                    })
                return json.dumps({
                    "status": "success",
                    "count": len(out_files),
                    "files": out_files,
                    "hint": (
                        "Decide from the questions/previews what the next "
                        "plan or refine call should read, then pass that "
                        "selection as literature_context — whole files by "
                        "path, or individual sections by their "
                        "section_ref ('<path>#qN'), comma-separated. "
                        "Omit literature_context to auto-load ALL of the "
                        "above (oldest-first union, deduped, capped at "
                        f"{_LIT_AUTOLOAD_MAX_CHARS:,} chars).")})
            except Exception as e:
                logging.error(f"Literature listing error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=list_literature_searches,
            name="list_literature_searches",
            description=(
                "Lists the current campaign's saved literature searches as "
                "an index: each file's questions and the beginning of each "
                "answer, with per-section sizes. Costs almost nothing — "
                "it reads no corpus into context. Call it before "
                "refine_plan_with_results() or write_technical_document() "
                "when the campaign may hold MORE THAN ONE literature file, "
                "and pass the relevant selection (file paths, or "
                "'<path>#qN' section refs) as literature_context. If you "
                "pass nothing, ALL campaign literature is auto-loaded — "
                "safe, but larger than a considered selection."
            ),
            parameters={},
            required=[]
        )

        # --- MOLECULES QUERY TOOL ---
        def query_molecules(objective: str):
            """
            Queries the FutureHouse Molecules agent for molecular design,
            synthesis planning, or cheminformatics tasks.
            Call this BEFORE generate_initial_plan() when the objective
            involves molecular design or discovery.
            """
            if not self.orch.mol_agent:
                return json.dumps({
                    "status": "error",
                    "message": "Molecules agent not available (no FutureHouse API key configured)"
                })

            # Guard: only proceed for genuine molecule design objectives
            if not is_molecule_design_objective(objective, self.orch.planner.model):
                return json.dumps({
                    "status": "skipped",
                    "message": "Objective does not appear to involve molecular design or synthesis planning. Skipping molecules query."
                })

            print(f"  ⚡ Tool: Querying MOLECULES agent for '{objective[:80]}...'")

            try:
                mol_res = self.orch.mol_agent.query(objective)

                if mol_res['status'] != 'success':
                    return json.dumps({
                        "status": mol_res['status'],
                        "message": mol_res.get('message', 'Molecules query did not succeed')
                    })

                # Save to file
                mol_path = self._output_dir() / "molecule_design.md"
                with open(mol_path, 'w') as f:
                    f.write("# Molecular Design & Synthesis Planning Results\n\n")
                    f.write(mol_res['content'])

                print(f"  ✅ Molecules query completed. Saved to {mol_path.name}")

                return json.dumps({
                    "status": "success",
                    "file_path": str(mol_path),
                    "content_preview": mol_res['content'][:500] + "..." if len(mol_res['content']) > 500 else mol_res['content'],
                    "hint": "Pass file_path as molecule_context to generate_initial_plan()"
                })

            except Exception as e:
                logging.error(f"Molecules query error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=query_molecules,
            name="query_molecules",
            description=(
                "Queries the FutureHouse Molecules agent for molecular design, synthesis planning, "
                "or cheminformatics. Call BEFORE generate_initial_plan() when the objective involves "
                "molecule design or discovery. Pass the returned file_path as molecule_context."
            ),
            parameters={
                "objective": {"type": "string", "description": "Molecular design or synthesis objective"}
            },
            required=["objective"]
        )

        # 1. GENERATE INITIAL PLAN
        def generate_initial_plan(
            specific_objective: str = None,
            knowledge_paths: str = None,
            primary_data_set: str = None,
            additional_context: str = None,
            skill: str = None,
            literature_context: str = None,
            molecule_context: str = None,
            n_candidates: int = None,
            selection_profile: str = None,
            white_paper: bool = None,
            new_campaign: bool = None,
            kind: str = "experiment"
        ):
            """
            Generates experimental plan (science strategy only, no code).

            Note: code_paths parameter is deprecated. Use generate_implementation_code()
            as a separate step to add code after plan approval.
            """
            obj = specific_objective if specific_objective else self.orch.objective
            print(f"  ⚡ Tool: Generating Initial Plan for '{obj}'...")

            # Campaign boundary (issue #396): decided up front so the
            # best-of-N default and the stale-literature guard below see it.
            _snc = getattr(self.orch.planner, "starts_new_campaign", None)
            starts_new = bool(_snc(obj, new_campaign)) if _snc else False

            # Guard: a NEW campaign fed literature that the registry
            # attributes only to PREVIOUS campaigns is cross-topic
            # contamination (live-confirmed: a chat-visible file path from
            # an earlier brainstorm gets re-passed for an unrelated one) —
            # unless the caller declares BOTH signals explicitly
            # (new_campaign=true + the file), which marks deliberate reuse.
            if starts_new and new_campaign is not True:
                stale = self._prior_campaign_literature(literature_context)
                if stale:
                    return json.dumps({
                        "status": "error",
                        "message": (
                            "literature_context points to literature from a "
                            "PREVIOUS campaign on a different topic: "
                            + ", ".join(Path(p).name for p in stale)
                            + ". This objective starts a NEW campaign, and "
                            "another topic's corpus must not ground it."
                        ),
                        "hint": (
                            "Omit literature_context (or run "
                            "search_literature for THIS topic first). If "
                            "reusing that literature is deliberate, pass "
                            "new_campaign=true together with the file."
                        ),
                    })

            # Resolve knowledge paths (with fallback to orchestrator dir)
            knowledge_list = self._resolve_knowledge_paths(knowledge_paths)
            if knowledge_list:
                # Validate paths
                invalid_paths = [p for p in knowledge_list if not Path(p).exists()]
                if invalid_paths:
                    return json.dumps({
                        "status": "error",
                        "message": f"Knowledge paths not found: {', '.join(invalid_paths)}",
                        "hint": "Check folder names and spelling"
                    })
                print(f"    📚 Knowledge sources: {knowledge_list}")

            # Parse primary dataset - UPDATED LOGIC
            primary_dataset = None
            if primary_data_set:
                # Try to resolve the path
                resolved_path, error = self._resolve_data_path(primary_data_set)
                
                if error:
                    return error  # Return the error JSON with suggestions
                
                path = Path(resolved_path)
                
                # Now handle resolved path
                if path.is_file():
                    primary_dataset = {"file_path": str(path)}
                    print(f"    📊 Primary data: {path.name}")
                    
                elif path.is_dir():
                    # Directory - check how many data files
                    all_files = []
                    for ext in ['*.csv', '*.xlsx', '*.xls']:
                        all_files.extend(path.glob(ext))
                    
                    if not all_files:
                        return json.dumps({
                            "status": "error",
                            "message": f"No data files (.csv, .xlsx, .xls) found in: {primary_data_set}",
                            "hint": "Add data files to the folder or specify a different path"
                        })
                    
                    elif len(all_files) == 1:
                        # Only one file - use it automatically
                        primary_dataset = {"file_path": str(all_files[0])}
                        print(f"    📊 Primary data (auto-selected): {all_files[0].name}")
                        
                    else:
                        # Multiple files - require user to specify
                        file_list = sorted([f.name for f in all_files], key=_natural_sort_key)
                        return json.dumps({
                            "status": "error",
                            "message": f"Multiple data files found in '{primary_data_set}'",
                            "available_files": file_list,
                            "file_count": len(file_list),
                            "hint": f"Please specify which file to use. Example: primary_data_set='./experimental_results/{file_list[0]}'"
                        })
            
            # Build context
            context_parts = []
            
            if additional_context:
                context_parts.append(f"User Requirements: {additional_context}")
                print(f"    ℹ️  User context: {additional_context[:60]}...")
            
            # Auto-include TEA results
            if self.orch.latest_tea_results:
                tea_summary = self.orch.latest_tea_results.get('summary', '')
                context_parts.append(f"Economic Analysis Results: {tea_summary}")
                print(f"    💰 Including TEA results in context")
            
            context_dict = None
            if context_parts:
                context_dict = {"user_context": "\n\n".join(context_parts)}
            
            # Resolve skill: use provided value or fall back to orchestrator's active skill
            effective_skill = skill or getattr(self.orch, '_active_skill', None)

            # Build external_context from literature/molecule files or raw text
            external_context_parts = []
            saved_extras = []
            if literature_context:
                lit_text = self._resolve_context_text(literature_context)
                if lit_text:
                    external_context_parts.append(lit_text)
                    print(f"    📚 Literature context resolved "
                          f"({len(lit_text.split())} words)")
            if molecule_context:
                mol_text = self._resolve_context_text(molecule_context)
                if mol_text:
                    external_context_parts.append(
                        "## Molecular Design & Synthesis Planning\n" + mol_text
                    )
                    print("    🧪 Molecule context resolved")

            ext_ctx = "\n\n".join(external_context_parts) if external_context_parts else None

            # DEPRECATED: ideation through the experiment tool. It is what
            # the portfolio contract replaces, so forward rather than author
            # a portfolio into the wrong schema one more time. Removed once
            # no caller passes it.
            if selection_profile == "ideation" and kind != "portfolio":
                logging.warning(
                    "selection_profile='ideation' on generate_initial_plan is "
                    "deprecated — forwarding to the portfolio contract; call "
                    "generate_ideation_portfolio directly.")
                print("    ↪️  ideation profile is deprecated on this tool — "
                      "authoring a portfolio instead.")
                kind = "portfolio"

            try:
                n_cand = resolve_n_candidates(
                    n_candidates, self.orch.planner.state,
                    new_campaign=starts_new)
                if n_candidates is None and n_cand > 1:
                    print(f"    🧭 New campaign — defaulting to best-of-{n_cand} "
                          "candidate plans (pass n_candidates=1 for a single plan).")
                plan = self.orch.planner.generate_plan(
                    objective=obj,
                    knowledge_paths=knowledge_list,
                    primary_data_set=primary_dataset,
                    additional_context=context_dict,
                    enable_human_feedback=self._get_human_feedback_enabled(),
                    reset_state=False,
                    skill=effective_skill,
                    external_context=ext_ctx,
                    n_candidates=n_cand,
                    candidate_report_dir=(str(self._output_dir() / "plan_candidates")
                                          if n_cand > 1 else None),
                    selection_profile=(selection_profile
                                       if selection_profile in ("lab", "ideation")
                                       else "lab"),
                    new_campaign=new_campaign,
                    kind=kind,
                )

                # An explicit lab profile marks THIS plan a bench plan even
                # inside an ideation campaign: the campaign ideated, the user
                # picked a direction, and now wants the executable protocol
                # for it. Without a way to say so the campaign stamp would
                # hand a runnable protocol a research dossier. Omitting the
                # profile inherits the campaign; only the explicit word
                # overrides it. `_stamp_campaign` never overwrites a type
                # that is already set, so this survives the later passes.
                if selection_profile == "lab" and isinstance(plan, dict):
                    plan.setdefault("type", "lab")
                    _cur = (self.orch.planner.state or {}).get("current_plan")
                    if isinstance(_cur, dict):
                        _cur.setdefault("type", "lab")

                # Store skill on orchestrator for downstream tools
                if effective_skill:
                    self.orch._active_skill = effective_skill

                # generate_plan signals failure as either `error` (RAG / parse
                # failure) or `status="failed"` + `last_error` (a returned state
                # dict) — check both so a failed plan never reports as success.
                if plan.get("error") or plan.get("status") == "failed":
                    return json.dumps({
                        "status": "error",
                        "message": (plan.get("error") or plan.get("last_error")
                                    or "Plan generation failed")
                    })

                # Claim literature for the (possibly new) campaign: pending
                # searches plus explicitly passed files (issue #396).
                self._adopt_literature(literature_context)

                # Save
                output_path = self._output_dir() / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(plan, f, indent=2)

                # Persist external grounding that arrived without a literature
                # file (molecule context / raw text stamped into the plan's
                # literature_search), so the campaign registry can carry it
                # into later refines. (The internal literature fallback that
                # originally motivated this save is removed.)
                if not literature_context and plan.get("literature_search"):
                    lit_path = self._output_dir() / "literature_search.md"
                    with open(lit_path, 'w') as f:
                        f.write("# Literature Search Results\n\n")
                        f.write(plan["literature_search"])
                    self._record_literature_file(lit_path)
                    saved_extras.append(str(lit_path))

                # Generate HTML — except for ideation runs: the report
                # template renders the plan as an executable protocol (an
                # ordered experimental-steps list, pipe-rows merged into
                # tables) with no ideation vocabulary, which misrepresents a
                # free-form research dossier; there the white paper is the
                # human-facing artifact and plan.json keeps the record.
                # (see the deprecation forward above)
                # Ask the PLAN what it is, not the best-of-N judge knob.
                # `selection_profile` weights candidate selection and is a
                # documented no-op with n_candidates=1 — so a consolidation
                # call inside an ideation campaign (live: "consolidate the
                # session's threads into one cross-cutting class") scored as
                # a lab run and got a protocol report, no dossier, and no
                # white paper, while the console it printed still used the
                # ideation vocabulary. The plan's own `type` stamp, inherited
                # from the campaign, is the honest signal; the profile clause
                # remains for the first call, which establishes the stamp.
                _ideation_run = (
                    plan.get("type") == "ideation"
                    or (selection_profile == "ideation" and n_cand > 1)
                )
                html_path = self._emit_plan_report(ideation=_ideation_run)

                num_experiments = len(plan.get('proposed_experiments', []))

                result = {
                    "status": "success",
                    "iteration": plan.get('iteration'),
                    "num_experiments": num_experiments,
                    "output_path": str(output_path),
                    "knowledge_used": knowledge_list is not None,
                    "primary_data_used": primary_dataset is not None,
                    "tea_context_included": self.orch.latest_tea_results is not None,
                    "hint": "Use generate_implementation_code() to add executable code"
                }
                if html_path is not None:
                    result["html_report"] = str(html_path)
                if saved_extras:
                    result["external_results_files"] = saved_extras
                pc = (self.orch.planner.state or {}).get("plan_candidates")
                if n_cand > 1 and pc:
                    result["best_of_n"] = {
                        "requested": n_cand,
                        "produced": len(pc.get("candidates", [])),
                        "selected_candidate": pc.get("selected_index"),
                        "human_override": pc.get("human_override", False),
                        "tier": pc.get("tier"),
                        "judge_reasoning": (pc.get("judge") or {}).get("reasoning", ""),
                        "candidate_reports": pc.get("reports", []),
                    }

                # Ideation runs get a detailed all-candidates dossier —
                # runner-ups are deliverables there, and rendering it
                # deterministically from the same state the white paper uses
                # keeps the two consistent by construction.
                # ...but the DOSSIER is a report over the candidate set, so it
                # needs one from THIS call. `plan_candidates` survives in state
                # across delegations, so a single-plan follow-up would render
                # its own flagship beside an EARLIER question's runner-ups — a
                # dossier answering a question the user has moved on from.
                if _ideation_run and n_cand > 1:
                    try:
                        result["ideation_report"] = self._write_ideation_report()
                    except Exception as e:  # noqa: BLE001 - plan result survives
                        logging.warning(f"Ideation report failed: {e}")

                # Ideation runs additionally produce a sponsor-facing white
                # paper by default (white_paper=False opts out; =True forces
                # one for any profile). Non-fatal: the plan already saved.
                _wp_auto = (white_paper is None and _ideation_run)
                if white_paper or _wp_auto:
                    try:
                        wp_path = self._write_white_paper()
                        result["white_paper"] = str(wp_path)
                        result["hint"] = (
                            "White paper saved alongside the plan — adapt it "
                            "for a pitch/pre-proposal; regenerate with "
                            "generate_white_paper(audience_context=...) to "
                            "target a specific sponsor. "
                            + result.get("hint", "")
                        )
                    except Exception as e:  # noqa: BLE001 - plan result survives
                        logging.warning(f"White paper generation failed: {e}")
                        result["white_paper_error"] = str(e)
                return json.dumps(result)
                
            except Exception as e:
                logging.error(f"Plan generation error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        # Register it
        self._register_tool(
            func=generate_initial_plan,
            name="generate_initial_plan",
            description=(
                "Generates an EXPERIMENTAL plan — a testable hypothesis with "
                "the measurements that would test it (science strategy only, "
                "no implementation code). "
                "NOT for a document that merely contains the word plan: a "
                "build/staging roadmap, a cost or footprint estimate, a "
                "consolidation memo or a summary is authored with "
                "write_technical_document. If the request has no hypothesis to "
                "test and nothing to measure, it is not this tool. "
                "Automatically includes previous TEA results if available. "
                "Can use: papers/reports, experimental data, lab constraints."
            ),
            parameters={
                "specific_objective": {"type": "string", "description": "Research objective"},
                "knowledge_paths": {"type": "string", "description": (
                    "Comma-separated paths to papers/reports/docs folders — "
                    "for document CORPORA too large to read directly "
                    "(triggers a full embedding knowledge-base build for "
                    "retrieval). A handful of documents you have ALREADY "
                    "read this session belongs in additional_context, not "
                    "here.")},
                "primary_data_set": {
                    "type": "string",
                    "description": (
                        "Path to a RAW EXPERIMENTAL data file or folder — "
                        "instrument measurements, composition assays, isotherms "
                        "the user collected. NOT for intermediate artifacts "
                        "produced by upstream tool calls in this session "
                        "(screening CSVs, scalarizer outputs, BO logs); those "
                        "belong in `knowledge_paths` or `additional_context`."
                    ),
                },
                "additional_context": {"type": "string", "description": "Lab constraints, equipment, reagents, budget, etc."},
                "skill": {
                    "type": "string",
                    "description": _build_planning_skill_description(
                        getattr(self.orch, "_custom_skills", None)
                    ),
                },
                "literature_context": {
                    "type": ["string", "array"],
                    "items": {"type": "string"},
                    "description": (
                        "Literature from search_literature(): a file path, a "
                        "LIST of file paths (or comma-separated), or raw "
                        "text. File contents are read and concatenated."
                    ),
                },
                "molecule_context": {"type": "string", "description": "File path or text from query_molecules() tool. Provides molecular design / synthesis context."},
                "new_campaign": {
                    "type": "boolean",
                    "description": (
                        "Campaign boundary. Set true when this plan starts "
                        "a NEW research topic unrelated to the current "
                        "campaign (e.g. a second brainstorm in the same "
                        "session) — the previous campaign's plans and "
                        "literature stay archived and are NOT carried into "
                        "the new topic's plans, white papers, or "
                        "refinements. Set false to force continuation of "
                        "the current campaign despite a reworded "
                        "objective. Omit to auto-detect from objective "
                        "similarity."
                    ),
                },
                "n_candidates": {
                    "type": "integer",
                    "description": (
                        "Best-of-N width (1-4). OMITTED: a campaign's FIRST "
                        "plan defaults to best-of-3 (distinct candidate "
                        "strategies, LLM judge picks, human can override); "
                        "later plans default to 1. Pass 1 explicitly when "
                        "the user wants a single plan; pass 2-4 to set the "
                        "width. A cap, not a quota: generation stops early "
                        "when the evidence supports no further distinct "
                        "approach. Keep specific_objective about ONE "
                        "scientific goal — do NOT ask for multiple "
                        "plans/strategies in the objective text; this "
                        "parameter provides the multiplicity."
                    ),
                },
                "selection_profile": {
                    "type": "string",
                    "enum": ["lab", "ideation"],
                    "description": (
                        "How the best-of-N judge weights its pick. 'lab' "
                        "(default): feasibility/actionability first — use "
                        "when the plan will actually be executed on stated "
                        "equipment. 'ideation': information gain and "
                        "mechanistic novelty first, feasibility only as a "
                        "tiebreaker — SWITCH to this when the user is "
                        "brainstorming, ideating, or asking for the most "
                        "scientifically interesting direction rather than "
                        "tomorrow's runnable protocol. Same candidates and "
                        "scores either way; only the authoring latitude and "
                        "the pick's weighting change, and the human can "
                        "still override. NOTE: ideation weighting requires "
                        "n_candidates >= 2 — with a single plan there is "
                        "nothing to select — so never pass n_candidates=1 "
                        "together with ideation. OMIT this parameter inside "
                        "an established campaign: the plan then inherits "
                        "that campaign's kind, so a follow-up or "
                        "consolidation in an ideation campaign still gets "
                        "the dossier and white paper rather than a bench "
                        "protocol report. Pass 'lab' EXPLICITLY to force a "
                        "bench/engineering plan even inside an ideation "
                        "campaign — the case is the user picking one ideated "
                        "direction and asking for its runnable bench "
                        "protocol. Note this tool is for EXPERIMENTAL "
                        "design and measurements; a roadmap, estimate or "
                        "summary document is not a plan and belongs in "
                        "save_file. Ideation plans also produce a "
                        "sponsor-facing white paper by default (see "
                        "white_paper)."
                    ),
                },
                "white_paper": {
                    "type": "boolean",
                    "description": (
                        "Also distill the plan into a sponsor-facing "
                        "technical white paper (pitch / pre-proposal). "
                        "OMITTED: automatic for ideation-profile runs, off "
                        "otherwise. Pass false to skip it on an ideation "
                        "run; pass true to force one for any profile."
                    ),
                },
            },
            required=[]
        )

        # --- WHITE PAPER TOOL ---
        def generate_white_paper(audience_context: str = None):
            """Distill the current campaign plan into a sponsor-facing
            white paper (technical pre-proposal)."""
            print("  ⚡ Tool: Generating white paper...")
            try:
                wp_path = self._write_white_paper(audience_context)
                text = Path(wp_path).read_text()
                return json.dumps({
                    "status": "success",
                    "white_paper": str(wp_path),
                    "word_count": len(text.split()),
                    "preview": text[:600],
                })
            except Exception as e:
                logging.error(f"White paper error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=generate_white_paper,
            name="generate_white_paper",
            description=(
                "Distill the CURRENT campaign plan into a technical white "
                "paper aimed at sponsors/program managers with technical "
                "backgrounds — a pitch / pre-proposal: significance and "
                "payoff forward, mechanisms rigorous, no bench-level "
                "protocol detail. After a best-of-N run it weaves distinct "
                "runner-up strategies in as secondary thrusts and turns the "
                "reviewer caveats into a risks-and-mitigation section. "
                "Requires an existing plan (generate_initial_plan first). "
                "Ideation runs already produce one automatically — call "
                "this to REGENERATE with sponsor targeting via "
                "audience_context, or to add one to a lab-profile plan."
            ),
            parameters={
                "audience_context": {
                    "type": "string",
                    "description": (
                        "Optional sponsor targeting, e.g. 'emphasize "
                        "fundamental-science significance and milestones' "
                        "or 'lead with cost and scalability impact'."
                    ),
                },
            },
            required=[],
        )

        # --- WORKFLOW DIAGRAM TOOL ---
        def generate_workflow_diagram(detail: str = "simple",
                                      focus: str = None):
            """QC'd Mermaid workflow diagram of the current campaign plan."""
            print("  ⚡ Tool: Generating workflow diagram...")
            from ...utils.mermaid_render import (
                mermaid_available, INSTALL_HINT)
            if not mermaid_available():
                return json.dumps({"status": "error",
                                   "message": INSTALL_HINT})
            plan = (self.orch.planner.state or {}).get("current_plan") or {}
            if not plan:
                return json.dumps({
                    "status": "error",
                    "message": "No campaign plan yet",
                    "hint": "Run generate_initial_plan first."})
            try:
                # Detail-suffixed stem so an elaborate rerun does not
                # silently overwrite the simple diagram (or vice versa).
                _stem = ("campaign_workflow" if detail == "simple"
                         else f"campaign_workflow_{detail}")
                res = self._get_diagram_agent().generate_workflow_diagram(
                    plan=plan, out_dir=self._output_dir(), detail=detail,
                    stem=_stem, extra_instructions=focus)
                if res.get("status") != "success":
                    return json.dumps({"status": "error",
                                       "message": res.get("error")})
                from .user_interface import record_deliverable
                record_deliverable(self.orch.base_dir, res["png_path"],
                                   "Campaign workflow diagram")
                return json.dumps({
                    "status": "success",
                    "diagram": res["png_path"],
                    "mermaid_source": res["mmd_path"],
                    "render_attempts": res["attempts"],
                    "visual_qc_rounds": res["qc_rounds"],
                })
            except Exception as e:
                logging.error(f"Workflow diagram error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=generate_workflow_diagram,
            name="generate_workflow_diagram",
            description=(
                "Draw a workflow diagram (rendered PNG) of the CURRENT "
                "campaign plan — generated as Mermaid code, retried on "
                "render errors, and passed through a visual quality gate "
                "before acceptance. Diagrams are COMPACT stage-level "
                "overviews by default; white papers already embed one "
                "automatically. Call this when the user explicitly asks "
                "for a (new or different) workflow/campaign diagram. "
                "Requires an existing plan."
            ),
            parameters={
                "detail": {
                    "type": "string",
                    "enum": ["simple", "elaborate"],
                    "description": (
                        "'simple' (default): compact stage-level overview. "
                        "'elaborate' ONLY when the user explicitly asks "
                        "for a detailed/step-by-step diagram."
                    ),
                },
                "focus": {
                    "type": "string",
                    "description": (
                        "Optional emphasis, e.g. 'highlight the "
                        "optimization loop' or 'show the decision gates'."
                    ),
                },
            },
            required=[],
        )

        # 2. GENERATE IMPLEMENTATION CODE
        def generate_implementation_code(code_paths: str = None):
            """
            Adds implementation code to the most recent experimental plan.
            Use after generate_initial_plan() to map experiments to executable code.
            
            Args:
                code_paths: Comma-separated paths to code folders. 
                        Optional if Code KB already loaded at startup.
            """
            
            if not self.orch.planner.state or not self.orch.planner.state.get("current_plan"):
                return json.dumps({
                    "status": "error",
                    "message": "No active plan. Generate a plan first using generate_initial_plan()"
                })
            
            current_plan = self.orch.planner.state["current_plan"]
            
            # Check if already has code
            if current_plan.get("proposed_experiments"):
                has_code = any(exp.get("implementation_code") for exp in current_plan["proposed_experiments"])
                if has_code:
                    return json.dumps({
                        "status": "warning",
                        "message": "Plan already has implementation code",
                        "hint": "Generate a new plan if you want to change the code source"
                    })
            
            print(f"  ⚡ Tool: Generating implementation code for existing plan...")

            kb_available = bool(self.orch.planner.kb_code.chunks)

            if not kb_available and not code_paths:
                return json.dumps({
                    "status": "error",
                    "message": "No Code Knowledge Base available",
                    "hint": "Provide code_paths parameter (e.g., code_paths='./opentrons_api,./automation_lib')",
                    "available_options": [
                        "Option 1: Specify code_paths='./your_code_folder'",
                        "Option 2: If code exists, check folder name and path"
                    ]
                })
            
            # Parse code paths
            code_list = []
            if code_paths:
                code_list = [p.strip() for p in code_paths.split(',') if p.strip()]
                
                # Validate paths (only if code_paths was provided)
                invalid_paths = []
                for path in code_list:
                    if not Path(path).exists():
                        invalid_paths.append(path)
                
                if invalid_paths:
                    # Check for common typos
                    suggestions = []
                    for invalid in invalid_paths:
                        parent = Path(invalid).parent
                        if parent.exists():
                            similar = [f.name for f in parent.iterdir() 
                                    if f.is_dir() and invalid.lower() in f.name.lower()]
                            if similar:
                                suggestions.append(f"Did you mean './{similar[0]}'?")
                    
                    hint = "Check folder names and spelling."
                    if suggestions:
                        hint += " " + " ".join(suggestions)
                    
                    return json.dumps({
                        "status": "error",
                        "message": f"Code paths not found: {', '.join(invalid_paths)}",
                        "hint": hint
                    })
                
                print(f"    💻 Code sources: {code_list}")
            elif kb_available:
                print(f"    💻 Using existing Code KB "
                      f"({self.orch.planner.kb_code.index.ntotal if self.orch.planner.kb_code.index else 0} vectors, "
                      f"{len(self.orch.planner.kb_code.chunks)} chunks)")
            
            try:
                updated_plan = self.orch.planner.generate_implementation_code(
                    plan=current_plan,
                    code_paths=code_list,
                    enable_human_feedback=self._get_human_feedback_enabled()
                )
                
                if updated_plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": updated_plan.get("error")
                    })
                
                # Save
                output_path = self._output_dir() / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(updated_plan, f, indent=2)
                
                # Regenerate HTML
                html_path = self._emit_plan_report()
                
                # Check if any experiments actually got code
                experiments = updated_plan.get("proposed_experiments", [])
                has_code = any(
                    exp.get("implementation_code")
                    for exp in experiments
                )

                if not has_code:
                    return json.dumps({
                        "status": "error",
                        "message": "Code generation failed — no executable code was produced for any experiment.",
                        "hint": "This may be due to an LLM API timeout or error. Try again.",
                        "output_path": str(output_path),
                        "html_report": str(html_path) if html_path else None
                    })

                # Save scripts to output folder
                final_out = str(self._output_dir() / "output_scripts")
                print(f"\n--- Saving Scripts to: {final_out} ---")
                write_experiments_to_disk(updated_plan, final_out)

                return json.dumps({
                    "status": "success",
                    "message": "Implementation code added to plan",
                    "output_path": str(output_path),
                    "html_report": str(html_path) if html_path else None,
                    "scripts_saved_to": final_out,
                    "code_sources_used": code_list
                })
                
            except Exception as e:
                logging.error(f"Code generation error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        # Register it
        self._register_tool(
            func=generate_implementation_code,
            name="generate_implementation_code",
            description=(
                "Generates executable implementation code for the most recent experimental plan. "
                "Maps experimental steps to code using API documentation and example repositories. "
                "Use after generate_initial_plan() once the scientific strategy is approved. "
                "If Code KB already loaded, code_paths is optional."
            ),
            parameters={
                "code_paths": {
                    "type": "string",
                    "description": (
                        "Comma-separated paths to SOURCE CODE or API documentation folders "
                        "(e.g., './opentrons_api,./automation_lib'). "
                        "Must contain .py, .js, or other code files — NOT scientific papers, "
                        "PDFs, or literature. Do NOT pass the knowledge directory here. "
                        "OPTIONAL if Code Knowledge Base is already loaded. "
                        "REQUIRED if no Code KB exists."
                    )
                }
            },
            required=[]
        )
        
        # 3. RUN ECONOMIC ANALYSIS
        def run_economic_analysis(
            focus_topic: str = None,
            knowledge_paths: str = None,
            primary_data_set: str = None,
            additional_context: str = None,
            literature_context: str = None
        ):
            """Performs Technoeconomic Analysis (TEA)."""
            obj = focus_topic if focus_topic else self.orch.objective
            print(f"  ⚡ Tool: Running TEA for '{obj}'...")

            # Resolve knowledge paths (with fallback to orchestrator dir)
            knowledge_list = self._resolve_knowledge_paths(knowledge_paths)
            if knowledge_list:
                print(f"    📚 Knowledge sources: {knowledge_list}")

            # Parse primary dataset
            primary_dataset = None
            if primary_data_set:
                # Try to resolve the path
                resolved_path, error = self._resolve_data_path(primary_data_set)
                
                if error:
                    return error  # Return the error JSON with suggestions
                
                path = Path(resolved_path)
                
                # Now handle resolved path
                if path.is_file():
                    primary_dataset = {"file_path": str(path)}
                    print(f"    📊 Primary data: {path.name}")
                    
                elif path.is_dir():
                    # Directory - check how many data files
                    all_files = []
                    for ext in ['*.csv', '*.xlsx', '*.xls']:
                        all_files.extend(path.glob(ext))
                    
                    if not all_files:
                        return json.dumps({
                            "status": "error",
                            "message": f"No data files (.csv, .xlsx, .xls) found in: {primary_data_set}",
                            "hint": "Add data files to the folder or specify a different path"
                        })
                    
                    elif len(all_files) == 1:
                        # Only one file - use it automatically
                        primary_dataset = {"file_path": str(all_files[0])}
                        print(f"    📊 Primary data (auto-selected): {all_files[0].name}")
                        
                    else:
                        # Multiple files - require user to specify
                        file_list = sorted([f.name for f in all_files], key=_natural_sort_key)
                        return json.dumps({
                            "status": "error",
                            "message": f"Multiple data files found in '{primary_data_set}'",
                            "available_files": file_list,
                            "file_count": len(file_list),
                            "hint": f"Please specify which file to use. Example: primary_data_set='./experimental_results/{file_list[0]}'"
                        })
            
            try:
                # Resolve literature context
                ext_ctx = None
                if literature_context:
                    lp = Path(literature_context)
                    ext_ctx = lp.read_text() if lp.is_file() else literature_context
                    print(f"    📚 Literature context from: {lp.name if lp.is_file() else 'inline text'}")

                res = self.orch.planner.perform_technoeconomic_analysis(
                    objective=obj,
                    knowledge_paths=knowledge_list,
                    primary_data_set=primary_dataset,
                    output_json_path=str(self._output_dir() / "tea_analysis.json"),
                    external_context=ext_ctx
                )
                
                if res.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": res.get("error")
                    })
                
                summary = res.get('technoeconomic_assessment', {}).get('summary', 'No summary')
                
                # Store TEA results in orchestrator state
                self.orch.latest_tea_results = {
                    "summary": summary,
                    "full_analysis": res.get('technoeconomic_assessment'),
                    "timestamp": datetime.now().isoformat()
                }
                print(f"    ✅ TEA results stored for future planning")
                
                return json.dumps({
                    "status": "success",
                    "summary": summary,
                    "output_path": str(self._output_dir() / "tea_analysis.json"),
                    "html_report": str(self._output_dir() / "tea_analysis.html"),
                    "hint": "These results will automatically inform future generate_initial_plan calls"
                })
                
            except Exception as e:
                logging.error(f"TEA error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        self._register_tool(
            func=run_economic_analysis,
            name="run_economic_analysis",
            description=(
                "Performs Technoeconomic Analysis (TEA) to assess economic viability, costs, market fit. "
                "Can incorporate papers and experimental data."
            ),
            parameters={
                "focus_topic": {
                    "type": "string",
                    "description": "Specific technology/process to analyze"
                },
                "knowledge_paths": {
                    "type": "string",
                    "description": "Comma-separated folder paths with papers/PDFs"
                },
                "primary_data_set": {
                    "type": "string",
                    "description": "Path to experimental data file or folder"
                },
                "additional_context": {
                    "type": "string",
                    "description": "Any other relevant context (constraints, requirements, etc.)"
                },
                "literature_context": {
                    "type": "string",
                    "description": "File path or text from search_literature(search_type='economic_data'). Provides external economic literature context."
                }
            },
            required=[]
        )

        # 4. REFINE PLAN (based on results)
        def refine_plan_with_results(
            result_data: str,
            use_literature_rag: bool = False,
            literature_context: str = None,
            molecule_context: str = None,
            additional_context: str = None
        ):
            """
            Refines the experimental plan (science strategy only) based on results.

            Use this for:
            - Strategic pivots or failures
            - Qualitative observations
            - Visual analysis of plots/images
            - When experiments didn't go as expected

            Supports multiple input formats:
            - Text: "Yield was 12%, precipitation observed"
            - File path: "./data.csv" or "./plot.png"
            - Comma-separated files: "./data.csv,./plot.png"
            """
            print(f"  ⚡ Tool: Refining Plan based on Results...")

            # Parse input - handle both single paths and comma-separated lists
            payload = self._parse_result_input(result_data)

            # Enrich with scalarizer metrics and plot if the file was already analyzed
            extras = self._collect_scalarizer_context(payload)
            if extras:
                if isinstance(payload, list):
                    payload.extend(extras)
                else:
                    payload = [payload] + extras

            # Build external context. Literature is tracked SEPARATELY from
            # molecule/additional context so provenance stamping downstream
            # never mistakes a critique or constraint note for literature.
            lit_text = None
            if literature_context:
                lit_text = self._resolve_context_text(literature_context)
                print(f"    📚 Literature context provided")
            else:
                # Auto-load ALL saved literature from the CURRENT campaign
                # (campaign-scoped registry, issue #396; plural union,
                # issue #425 — the newest-file singular load silently
                # dropped 77% of a live session's corpus). A campaign
                # that supplied no literature refines without any; another
                # topic's corpus is never injected.
                lit = self._load_campaign_literature()
                if lit is not None:
                    lit_text = lit["text"]
                    print(f"    📚 Auto-loaded literature context from "
                          f"session ({lit['n_files']} file(s), "
                          f"{len(lit_text):,} chars: "
                          f"{', '.join(lit['files'])})")
            ext_parts = [lit_text] if lit_text else []
            if molecule_context:
                mol_text = self._resolve_context_text(molecule_context)
                ext_parts.append("## Molecular Design & Synthesis Planning\n" + mol_text)
                print(f"    🧪 Molecule context provided")
            if additional_context:
                ext_parts.append(f"## Additional Context\n{additional_context}")
                print(f"    ℹ️  Additional context provided")
            ext_ctx = "\n\n".join(ext_parts) if ext_parts else None

            try:
                plan = self.orch.planner.refine_plan(
                    results=payload,
                    enable_human_feedback=self._get_human_feedback_enabled(),
                    use_literature_rag=use_literature_rag,
                    external_context=ext_ctx,
                    literature_text=lit_text
                )
                
                if plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": plan.get("error")
                    })

                # Explicitly supplied literature files now belong to this
                # campaign's corpus (issue #396).
                self._adopt_literature(literature_context)

                # (literature_search provenance is stamped inside
                # refine_plan itself — the single point where the refined
                # plan enters current_plan and history — from the
                # literature_text passed above.)

                # Save
                output_path = self._output_dir() / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(plan, f, indent=2)

                # Generate HTML
                html_path = self._emit_plan_report()

                return json.dumps({
                    "status": "success",
                    "iteration": plan.get('iteration'),
                    "num_experiments": len(plan.get('proposed_experiments', [])),
                    "output_path": str(output_path),
                    "html_report": str(html_path) if html_path else None,
                    "hint": "Use refine_implementation_code() to update executable code"
                })
                
            except Exception as e:
                logging.error(f"Plan refinement error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=refine_plan_with_results,
            name="refine_plan_with_results",
            description=(
                "Refines experimental plan (science strategy only) based on results. "
                "Handles text descriptions, single file paths, or comma-separated files. "
                "Use for: failures, pivots, qualitative observations, or visual analysis. "
                "Does NOT update implementation code - use refine_implementation_code() for that."
            ),
            parameters={
                "result_data": {
                    "type": "string",
                    "description": "Experimental results (text, file path, or comma-separated files)"
                },
                "use_literature_rag": {
                    "type": "boolean",
                    "description": "Search local knowledge base for relevant context. Default: false."
                },
                "literature_context": {
                    "type": "string",
                    "description": (
                        "External scientific literature for refinement: "
                        "file path(s) or '<path>#qN' section refs "
                        "(comma-separated; see list_literature_searches), "
                        "or raw text. OMITTED: all of the current "
                        "campaign's saved literature is auto-loaded — the "
                        "right default. When the campaign holds several "
                        "literature files and the refinement is narrow, "
                        "consider selecting via list_literature_searches "
                        "first."
                    )
                },
                "molecule_context": {
                    "type": "string",
                    "description": "File path or text from query_molecules() tool. Provides molecular design / synthesis context for refinement."
                },
                "additional_context": {
                    "type": "string",
                    "description": "Extra context (e.g., reference data from query_knowledge_data, constraints, observations) to inform refinement."
                }
            },
            required=["result_data"]
        )
        
        # 4b. ADJUST PLAN FOR CONSTRAINTS (pre-execution)
        def adjust_plan_for_constraints(constraint_description: str):
            """
            Adjusts the experimental plan for implementation or instrument
            constraints discovered during protocol/code generation.
            Does NOT increment the iteration — the experiment hasn't run yet.
            """
            print(f"  ⚡ Tool: Adjusting plan for implementation constraints...")

            try:
                plan = self.orch.planner.adjust_plan_for_constraints(
                    constraint_description=constraint_description,
                    enable_human_feedback=self._get_human_feedback_enabled()
                )

                if plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": plan.get("error")
                    })

                # Save
                output_path = self._output_dir() / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(plan, f, indent=2)

                # Generate HTML
                html_path = self._emit_plan_report()

                return json.dumps({
                    "status": "success",
                    "iteration": plan.get('iteration'),
                    "num_experiments": len(plan.get('proposed_experiments', [])),
                    "output_path": str(output_path),
                    "html_report": str(html_path) if html_path else None,
                    "hint": "Use generate_implementation_code() or refine_implementation_code() to update executable code for the adjusted plan."
                })

            except Exception as e:
                logging.error(f"Plan adjustment error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        self._register_tool(
            func=adjust_plan_for_constraints,
            name="adjust_plan_for_constraints",
            description=(
                "Adjusts the experimental plan when implementation or instrument "
                "constraints make the current plan impractical BEFORE running the experiment. "
                "Use when protocol generation reveals incompatibilities (e.g., pipette type "
                "vs plate layout, equipment limitations, reagent availability). "
                "Does NOT increment iteration or log as experimental results. "
                "Use refine_plan_with_results() instead when adjusting based on actual experimental outcomes."
            ),
            parameters={
                "constraint_description": {
                    "type": "string",
                    "description": (
                        "Description of the implementation constraint or instrument "
                        "incompatibility that requires plan adjustment. Include what "
                        "the constraint is, why it conflicts with the current plan, "
                        "and any proposed resolution if known."
                    )
                }
            },
            required=["constraint_description"]
        )

        # 5. REFINE IMPLEMENTATION CODE (based on refined plan)
        def refine_implementation_code():
            """
            Updates implementation code for the most recently refined plan.
            Use after refine_plan_with_results() to add/update executable code.
            """
            
            if not self.orch.planner.state or not self.orch.planner.state.get("current_plan"):
                return json.dumps({
                    "status": "error",
                    "message": "No active plan. Refine a plan first using refine_plan_with_results()"
                })
            
            current_plan = self.orch.planner.state["current_plan"]
            
            print(f"  ⚡ Tool: Refining implementation code for iteration {current_plan.get('iteration')}...")
            
            try:
                updated_plan = self.orch.planner.refine_implementation_code(
                    plan=current_plan,
                    enable_human_feedback=self._get_human_feedback_enabled()
                )
                
                if updated_plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": updated_plan.get("error")
                    })
                
                # Save
                output_path = self._output_dir() / "plan_refined.json"
                with open(output_path, 'w') as f:
                    json.dump(updated_plan, f, indent=2)
                
                # Regenerate HTML
                html_path = self._emit_plan_report("plan_refined.html")
                
                # Save scripts
                final_out = str(self._output_dir() / "output_scripts")
                print(f"\n--- Saving Scripts to: {final_out} ---")
                write_experiments_to_disk(updated_plan, final_out)
                
                return json.dumps({
                    "status": "success",
                    "message": "Implementation code updated",
                    "output_path": str(output_path),
                    "html_report": str(html_path) if html_path else None,
                    "scripts_saved_to": final_out
                })
                
            except Exception as e:
                logging.error(f"Code refinement error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=refine_implementation_code,
            name="refine_implementation_code",
            description=(
                "Updates implementation code for the most recently refined plan. "
                "Maps refined experimental steps to executable code. "
                "Use after refine_plan_with_results() once the scientific strategy is approved."
            ),
            parameters={},
            required=[]
        )
        
        def analyze_file(
                file_path: str,
                extraction_goal: str = None,
                force_regenerate: bool = False,
                inputs: list[str] = None,
                targets: list[str] = None):
            """
            Analyzes a raw data file (CSV/XLSX) to extract metrics.
            
            Args:
                file_path: Path to data file
                extraction_goal: What to extract
                force_regenerate: If True, regenerates analysis script even if one exists.
                inputs: List of column names to treat as INPUT parameters for optimization
                targets: List of column names to treat as OPTIMIZATION TARGETS
            """
            print(f"  ⚡ Tool: Analyzing {file_path}...")

            resolved_path, error = self._resolve_data_path(file_path)
            if error:
                return error
            file_path = resolved_path

            # Resolve absolute path for tracking
            file_path_abs = str(Path(file_path).resolve())
            
            #  Build schema-aware extraction goal
            enhanced_objective = extraction_goal or ""
            # Always include the campaign objective so the scalarizer knows
            # what physically meaningful targets to derive
            if self.orch.objective and self.orch.objective != "Undefined Research Goal":
                enhanced_objective = (
                    f"Research objective: {self.orch.objective}\n\n{enhanced_objective}"
                ).strip()

            if inputs and targets:
                # User explicitly specified schema - incorporate into the objective query
                schema_instruction = f"""
        REQUIRED OUTPUT SCHEMA:
        - INPUT PARAMETERS (for optimization): {inputs}
        - TARGET METRICS (to optimize): {targets}

        Extract EXACTLY these columns from the data. Each row should contain values for all input parameters and all target metrics.
        For multi-objective optimization, we need BOTH targets: {targets}
        """
                enhanced_objective = f"{enhanced_objective}\n\n{schema_instruction}".strip()
                print(f"    📊 User-specified schema:")
                print(f"       Inputs: {inputs}")
                print(f"       Targets: {targets}")
            
            # Determine script to use.
            # Strategy: if a locked script exists, always try it first. If it
            # fails on the new data (e.g., different columns), auto-regenerate.
            # This prevents false schema-change triggers from LLM-supplied
            # target names that don't exactly match stored names.
            has_locked_script = (
                self.orch.active_scalarizer_script
                and Path(self.orch.active_scalarizer_script).exists()
            )
            if force_regenerate:
                script_to_use = None
                print(f"    🔄 Force regenerate: Creating new analysis script")
            elif has_locked_script:
                script_to_use = self.orch.active_scalarizer_script
                print(f"    (Consistency Mode: Using cached script)")
            else:
                script_to_use = None
                print(f"    (Discovery Mode: Generating new script)")

            # Pass schema to experiment context
            current_plan = self.orch.planner.state.get("current_plan", {})
            # `or [{}]` — the dict default only fires on a MISSING key, so a plan
            # carrying an empty experiments list would IndexError here.
            _exps = (current_plan or {}).get("proposed_experiments") or [{}]
            exp_context = _exps[0]

            # Inject schema requirements into context (only when generating new script)
            role_hints = None
            if inputs and targets and not has_locked_script:
                exp_context = exp_context.copy() if exp_context else {}
                exp_context["_schema_requirements"] = {
                    "input_columns": inputs,
                    "target_columns": targets,
                    "optimization_type": "multi-objective" if len(targets) > 1 else "single-objective"
                }
                role_hints = {"inputs": inputs, "targets": targets}

            try:
                res = self.orch.scalarizer.scalarize(
                    data_path=file_path,
                    objective_query=enhanced_objective,
                    reuse_script_path=script_to_use,
                    experiment_context=exp_context,
                    enable_human_review=self._get_human_feedback_enabled(),
                    column_role_hints=role_hints
                )

                if res["status"] != "success":
                    hint = "Try force_regenerate=True if the data format has changed"
                    if script_to_use:
                        hint = (
                            "The locked analysis script failed on this file. "
                            "If the data format has changed, use force_regenerate=True "
                            "to create a new script."
                        )
                    return json.dumps({
                        "status": "error",
                        "message": res.get('error', 'Analysis failed'),
                        "hint": hint
                    })

                # Validate sidecar conditions match script output
                # If the script hardcoded values from a different file's sidecar,
                # the output will have wrong condition values. Detect and auto-regenerate.
                if script_to_use and not force_regenerate:
                    sidecar_path = Path(file_path).with_suffix('.json')
                    if sidecar_path.exists():
                        try:
                            with open(sidecar_path) as _sc:
                                sidecar_data = json.load(_sc)
                            metrics_to_check = res["metrics"]
                            if isinstance(metrics_to_check, list):
                                metrics_to_check = metrics_to_check[0] if metrics_to_check else {}
                            mismatched = []
                            for key, expected_val in sidecar_data.items():
                                if key in metrics_to_check:
                                    actual_val = metrics_to_check[key]
                                    if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
                                        if abs(actual_val - expected_val) > 1e-6:
                                            mismatched.append(f"{key}: expected {expected_val}, got {actual_val}")
                            if mismatched:
                                print(f"    ⚠️  Sidecar mismatch detected (script has hardcoded values):")
                                for m in mismatched:
                                    print(f"       {m}")
                                print(f"    🔄 Auto-regenerating script...")
                                res = self.orch.scalarizer.scalarize(
                                    data_path=file_path,
                                    objective_query=enhanced_objective,
                                    reuse_script_path=None,
                                    experiment_context=exp_context,
                                    enable_human_review=self._get_human_feedback_enabled(),
                                    column_role_hints=role_hints
                                )
                                if res["status"] != "success":
                                    return json.dumps({
                                        "status": "error",
                                        "message": res.get('error', 'Regeneration failed'),
                                    })
                                force_regenerate = True  # ensure script lock updates below
                        except Exception:
                            pass  # If sidecar can't be read, skip validation

                # Pass-through ingests (#366) carry no script to lock.
                if res.get("source_script") and (
                        not self.orch.active_scalarizer_script or force_regenerate):
                    self.orch.active_scalarizer_script = res["source_script"]
                    print(f"    ✅ Analysis Logic Locked: {Path(self.orch.active_scalarizer_script).name}")

                # Merge sidecar conditions into scalarizer output.
                # analyze_batch merges conditions externally (line ~1584),
                # but analyze_file did not — so a batch-generated script
                # that only outputs target metrics (e.g. Peak_Absorbance)
                # would miss input parameters (e.g. temperature, pH) that
                # live in the sidecar JSON, causing a column mismatch on
                # CSV append.
                sidecar_merge_path = Path(file_path).with_suffix('.json')
                if sidecar_merge_path.exists():
                    try:
                        with open(sidecar_merge_path, 'r') as _sc:
                            sidecar_conds = json.load(_sc)
                        if isinstance(sidecar_conds, dict):
                            scalar_conds = {
                                k: v for k, v in sidecar_conds.items()
                                if isinstance(v, (int, float, str))
                            }
                            if scalar_conds:
                                raw = res["metrics"]
                                if isinstance(raw, list):
                                    res["metrics"] = [{**row, **scalar_conds} for row in raw]
                                elif isinstance(raw, dict):
                                    res["metrics"] = {**raw, **scalar_conds}
                                print(f"    📎 Merged sidecar conditions: {list(scalar_conds.keys())}")
                    except Exception as e:
                        logging.warning(f"Could not merge sidecar conditions: {e}")

                # Handle both single-row and multi-row results
                metrics = res["metrics"]
                
                if isinstance(metrics, list):
                    df_new = pd.DataFrame(metrics)
                    print(f"    📊 Processing {len(df_new)} data points from multi-well experiment")
                elif isinstance(metrics, dict):
                    # A multi-row table pass-through (#366) returns columns
                    # as equal-length lists; expand to one row per
                    # experiment, broadcasting scalar sidecar conditions.
                    _lens = {len(v) for v in metrics.values()
                             if isinstance(v, list)}
                    if res.get("passthrough") and len(_lens) == 1 and _lens != {1}:
                        _n = next(iter(_lens))
                        df_new = pd.DataFrame(
                            {k: (v if isinstance(v, list) else [v] * _n)
                             for k, v in metrics.items()})
                        print(f"    📊 Processing {_n} data points from table pass-through")
                    else:
                        df_new = pd.DataFrame([metrics])
                else:
                    return json.dumps({
                        "status": "error",
                        "message": f"Unexpected metrics format: {type(metrics)}"
                    })
                
                # DEDUPLICATION - Content-based tracking
                # Compute current file hash
                current_hash = self._compute_file_hash(file_path)
                current_row_count = len(df_new)

                # Get previous tracking for this file (handle both old and new format)
                prev_tracking = self.orch.analyzed_files.get(file_path_abs, {})
                if isinstance(prev_tracking, dict):
                    prev_hash = prev_tracking.get('hash')
                    prev_row_count = prev_tracking.get('row_count', 0)
                else:
                    # Legacy format: just row count as int
                    prev_hash = None
                    prev_row_count = prev_tracking

                # Check for duplicate content across different filenames
                for tracked_path, tracking_info in self.orch.analyzed_files.items():
                    if tracked_path == file_path_abs:
                        continue  # Skip self
                    tracked_hash = tracking_info.get('hash') if isinstance(tracking_info, dict) else None
                    if tracked_hash and tracked_hash == current_hash:
                        print(f"    ⚠️  Duplicate content detected - matches: {Path(tracked_path).name}")
                        df_final = pd.read_csv(self.orch.bo_data_path) if self.orch.bo_data_path.exists() else pd.DataFrame()
                        return json.dumps({
                            "status": "warning",
                            "message": f"This file's content was already analyzed from '{Path(tracked_path).name}'",
                            "data_points_collected": len(df_final),
                            "rows_added": 0,
                            "optimization_ready": len(df_final) >= 3,
                            "hint": "Data already in optimization set. No action needed unless this is different data with identical content."
                        })

                # Determine what to process based on hash and row count
                if prev_hash is None:
                    # FIRST TIME analyzing this file
                    print(f"    ✨ First time analyzing this file")
                    df_to_append = df_new
                    num_new = len(df_new)

                elif prev_hash != current_hash:
                    # FILE CONTENT CHANGED - reprocess entirely
                    print(f"    🔄 File content changed (hash mismatch) - reprocessing entirely")
                    
                    # Remove old data from optimization_data.csv if it exists
                    if self.orch.bo_data_path.exists() and prev_row_count > 0:
                        try:
                            df_existing = pd.read_csv(self.orch.bo_data_path)
                            # Remove the last prev_row_count rows (assumes they're from this file)
                            if len(df_existing) >= prev_row_count:
                                df_existing = df_existing.iloc[:-prev_row_count]
                                df_existing.to_csv(self.orch.bo_data_path, index=False)
                                print(f"    🗑️  Removed {prev_row_count} old rows from optimization data")
                        except Exception as e:
                            logging.warning(f"Could not clean old data: {e}")
                    
                    df_to_append = df_new
                    num_new = len(df_new)

                elif current_row_count > prev_row_count:
                    # ROWS APPENDED - process only new rows
                    df_new_only = df_new.iloc[prev_row_count:]
                    num_skipped = prev_row_count
                    num_new = len(df_new_only)
                    
                    if num_skipped > 0:
                        print(f"    🔍 Skipped {num_skipped} previously analyzed row(s)")
                    print(f"    ✅ Adding {num_new} NEW row(s)")
                    
                    df_to_append = df_new_only

                elif current_row_count == prev_row_count:
                    # prev_hash == current_hash (guaranteed by earlier elif)
                    schema_changed = (
                        inputs and targets and (
                            set(inputs) != set(self.orch.expected_input_columns or [])
                            or set(targets) != set(self.orch.expected_target_columns or [])
                        )
                    )
                    if schema_changed:
                        # Same data, new schema — reprocess with updated columns
                        print(f"    🔄 Schema changed — reprocessing with new inputs/targets")
                        df_to_append = df_new
                        num_new = len(df_new)
                        # Clear existing optimization data (schema mismatch)
                        if self.orch.bo_data_path.exists():
                            backup = self.orch.bo_data_path.with_suffix('.csv.backup')
                            self.orch.bo_data_path.rename(backup)
                            print(f"    ⚠️  Old optimization data backed up (schema change)")
                    else:
                        # TRULY UNCHANGED
                        print(f"    ℹ️  File unchanged (same content hash)")
                        df_final = pd.read_csv(self.orch.bo_data_path) if self.orch.bo_data_path.exists() else pd.DataFrame()

                        return json.dumps({
                            "status": "success",
                            "message": "File already analyzed - no changes detected",
                            "data_points_collected": len(df_final),
                            "rows_added": 0,
                            "optimization_ready": len(df_final) >= 3
                        })

                else:
                    # FEWER ROWS - file was truncated/replaced
                    print(f"    ⚠️  File has fewer rows ({current_row_count} < {prev_row_count}) - reprocessing")
                    
                    # Remove old data
                    if self.orch.bo_data_path.exists() and prev_row_count > 0:
                        try:
                            df_existing = pd.read_csv(self.orch.bo_data_path)
                            if len(df_existing) >= prev_row_count:
                                df_existing = df_existing.iloc[:-prev_row_count]
                                df_existing.to_csv(self.orch.bo_data_path, index=False)
                                print(f"    🗑️  Removed {prev_row_count} old rows from optimization data")
                        except Exception as e:
                            logging.warning(f"Could not clean old data: {e}")
                    
                    df_to_append = df_new
                    num_new = len(df_new)

                # Schema enforcement BEFORE saving
                all_cols = list(df_to_append.columns)

                # Case 1: Agent explicitly provided schema (Enables MOO)
                if inputs and targets:
                    # Validate that requested columns exist in the extracted data
                    missing_inputs = [c for c in inputs if c not in all_cols]
                    missing_targets = [t for t in targets if t not in all_cols]
                    
                    if missing_inputs or missing_targets:
                        # Try fuzzy matching for column names
                        available_cols = all_cols
                        suggestions = {}
                        
                        for missing in missing_inputs + missing_targets:
                            # Simple fuzzy match: find columns containing similar substrings
                            matches = [c for c in available_cols if missing.lower().replace('_', '') in c.lower().replace('_', '') 
                                    or c.lower().replace('_', '') in missing.lower().replace('_', '')]
                            if matches:
                                suggestions[missing] = matches
                        
                        return json.dumps({
                            "status": "schema_mismatch",
                            "message": (
                                f"The analysis script could not produce the requested columns. "
                                f"Missing inputs: {missing_inputs or 'none'}. "
                                f"Missing targets: {missing_targets or 'none'}. "
                                f"Available columns from extraction: {all_cols}."
                            ),
                            "missing_inputs": missing_inputs,
                            "missing_targets": missing_targets,
                            "available_columns": all_cols,
                            "suggestions": suggestions if suggestions else None,
                            "recovery_options": [
                                "Retry with corrected column names from available_columns",
                                "Use force_regenerate=True with an updated extraction_goal",
                                "Choose different inputs/targets from the available columns"
                            ]
                        })
                    
                    self.orch.expected_input_columns = inputs
                    self.orch.expected_target_columns = targets
                    # Capture optimization direction and input types from scalarizer
                    column_roles = res.get("column_roles", {})
                    opt_dir = column_roles.get("optimization_direction", {})
                    if opt_dir:
                        self.orch.target_directions = opt_dir
                    self._capture_input_types(column_roles, inputs)
                    print(f"    📊 Schema Enforced (User-Specified):")
                    print(f"       Inputs: {self.orch.expected_input_columns}")
                    print(f"       Targets: {self.orch.expected_target_columns}")
                    if self.orch.target_directions:
                        print(f"       Directions: {self.orch.target_directions}")

                # Case 2: Schema already established from previous analysis
                elif self.orch.expected_input_columns and self.orch.expected_target_columns:
                    # Still capture direction if scalarizer provided it and we don't have one yet
                    column_roles = res.get("column_roles", {})
                    if not self.orch.target_directions:
                        opt_dir = column_roles.get("optimization_direction", {})
                        if opt_dir:
                            self.orch.target_directions = opt_dir
                    if not getattr(self.orch, "expected_input_types", None):
                        self._capture_input_types(column_roles, self.orch.expected_input_columns)
                    print(f"    📊 Schema Enforced (From Previous Analysis):")
                    print(f"       Inputs: {self.orch.expected_input_columns}")
                    print(f"       Targets: {self.orch.expected_target_columns}")
                
                # Case 3: No user schema — use scalarizer's column_roles classification
                else:
                    column_roles = res.get("column_roles", {})
                    proposed_inputs = column_roles.get("inputs", [])
                    proposed_targets = column_roles.get("targets", [])

                    if proposed_inputs and proposed_targets:
                        # Validate proposed columns exist in extracted data
                        missing = [c for c in proposed_inputs + proposed_targets if c not in all_cols]
                        if missing:
                            print(f"    ⚠️  Scalarizer classification references missing columns: {missing}")
                            # Fall through to schema_required below
                            proposed_inputs, proposed_targets = [], []
                        else:
                            reasoning = column_roles.get("reasoning", "")
                            opt_dir = column_roles.get("optimization_direction", {})
                            print(f"    🔬 Scalarizer classified columns:")
                            print(f"       Inputs: {proposed_inputs}")
                            print(f"       Targets: {proposed_targets}")
                            if opt_dir:
                                print(f"       Directions: {opt_dir}")
                            if reasoning:
                                print(f"       Reasoning: {reasoning}")

                            if self.orch.autonomy_level == "CO_PILOT":
                                # Return proposal for user confirmation
                                n_data = len(df_to_append)
                                return json.dumps({
                                    "status": "schema_proposed",
                                    "inputs": proposed_inputs,
                                    "targets": proposed_targets,
                                    "optimization_direction": opt_dir,
                                    "reasoning": reasoning,
                                    "data_points": n_data,
                                    "message": "Scalarizer proposes this classification. Confirm or adjust.",
                                    "available_columns": all_cols
                                })
                            else:
                                # AUTOPILOT/AUTONOMOUS: accept directly
                                self.orch.expected_input_columns = proposed_inputs
                                self.orch.expected_target_columns = proposed_targets
                                if opt_dir:
                                    self.orch.target_directions = opt_dir
                                self._capture_input_types(column_roles, proposed_inputs)
                                print(f"    ✅ Schema auto-accepted: inputs={proposed_inputs}, targets={proposed_targets}")

                    # Targets found but no inputs — measurement-only data (e.g., spectra)
                    if proposed_targets and not proposed_inputs:
                        n_data = len(df_to_append)
                        return json.dumps({
                            "status": "inputs_required",
                            "message": (
                                "The data file contains measurement data but no experimental conditions. "
                                "Input parameters (e.g., temperature, pH, concentration) are needed for optimization."
                            ),
                            "targets_found": proposed_targets,
                            "reasoning": column_roles.get("reasoning", ""),
                            "data_points": n_data,
                            "options": [
                                "Provide a metadata JSON sidecar file with experimental conditions",
                                "Manually specify input parameter values for this data file",
                                "Re-call analyze_file with inputs=[...] listing the parameter names to add"
                            ]
                        })

                    # Fallback: scalarizer didn't classify or classification was invalid
                    if not proposed_inputs or not proposed_targets:
                        print(f"    ⚠️  No schema specified. Extracted columns: {all_cols}")
                        data_preview = df_to_append.head(3).to_dict(orient='records')
                        col_stats = {}
                        for col in all_cols:
                            if pd.api.types.is_numeric_dtype(df_to_append[col]):
                                col_stats[col] = {
                                    "type": "numeric",
                                    "unique": int(df_to_append[col].nunique()),
                                    "min": float(df_to_append[col].min()),
                                    "max": float(df_to_append[col].max()),
                                }
                            else:
                                col_stats[col] = {"type": "non-numeric", "unique": int(df_to_append[col].nunique())}
                        n_data = len(df_to_append)
                        numeric_cols = [c for c in all_cols if pd.api.types.is_numeric_dtype(df_to_append[c])]
                        return json.dumps({
                            "status": "schema_required",
                            "message": "Could not auto-classify columns. Re-call analyze_file with explicit inputs and targets.",
                            "available_columns": all_cols,
                            "column_stats": col_stats,
                            "data_preview": data_preview,
                            "data_points": n_data,
                            "objective": self.orch.objective or "Not set",
                            "hint": (
                                "Use the objective, column names, and data preview to decide: "
                                "which columns are controllable INPUT parameters (experimentally set) "
                                "and which are measured TARGET metrics (outcomes to optimize). "
                                "Non-numeric columns (e.g., Sample_ID, Notes) should be excluded. "
                                "Then call: analyze_file(file_path=..., inputs=[...], targets=[...])"
                            ),
                            "objective_count_guidance": self._build_objective_guidance(n_data, numeric_cols)
                        })
                
                # FILTER TO MATCH CSV SCHEMA
                # The scalarizer may output extra metrics beyond what the
                # optimization CSV tracks.  Use the existing CSV columns as
                # ground truth (they may be wider than expected_target_columns
                # if targets were narrowed via run_optimization).  For new
                # CSVs, fall back to the expected schema.
                if self.orch.bo_data_path.exists():
                    df_existing = pd.read_csv(self.orch.bo_data_path)
                    ref_cols = list(df_existing.columns)
                else:
                    df_existing = None
                    if self.orch.expected_input_columns and self.orch.expected_target_columns:
                        ref_cols = self.orch.expected_input_columns + self.orch.expected_target_columns
                    else:
                        ref_cols = None

                if ref_cols:
                    available = [c for c in ref_cols if c in df_to_append.columns]
                    extra = [c for c in df_to_append.columns if c not in ref_cols]
                    missing = [c for c in ref_cols if c not in df_to_append.columns]
                    if extra:
                        print(f"    📎 Dropping extra columns: {extra}")
                    if missing:
                        print(f"    ⚠️  Missing expected columns: {missing}")
                    if available:
                        df_to_append = df_to_append[available]

                # SCHEMA ENFORCEMENT ON SAVE
                if df_existing is not None:
                    if set(df_to_append.columns) != set(df_existing.columns):
                        return json.dumps({
                            "status": "error",
                            "message": "Schema mismatch detected",
                            "expected_columns": list(df_existing.columns),
                            "received_columns": list(df_to_append.columns),
                            "hint": "All data must have same structure. Use reset_analysis_logic to start fresh."
                        })

                    df_to_append = df_to_append[df_existing.columns]
                    df_to_append.to_csv(self.orch.bo_data_path, mode='a', header=False, index=False)
                else:
                    df_to_append.to_csv(self.orch.bo_data_path, mode='w', header=True, index=False)
                
                # Update tracking
                self.orch.analyzed_files[file_path_abs] = {
                    'row_count': current_row_count,
                    'hash': current_hash,
                    'timestamp': datetime.now().isoformat()
                }
                with open(self.orch.analyzed_files_path, 'w') as f:
                    json.dump(self.orch.analyzed_files, f, indent=2)
                
                df_final = pd.read_csv(self.orch.bo_data_path)
                data_count = len(df_final)
                
                return json.dumps({
                    "status": "success",
                    "data_points_collected": data_count,
                    "rows_added": num_new,
                    "optimization_ready": data_count >= 3,
                    "inputs": self.orch.expected_input_columns,
                    "targets": self.orch.expected_target_columns
                })
                
            except Exception as e:
                logging.error(f"Analyze file error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=analyze_file,
            name="analyze_file",
            description=(
                "Analyzes raw data files (CSV/XLSX/TXT) to extract scalar metrics. "
                "Automatically generates analysis code on first use, then reuses it for consistency. "
                "Results are appended to optimization dataset."
            ),
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Path to the data file to analyze (e.g., 'results/run_001.csv')"
                },
                "extraction_goal": {
                    "type": "string",
                    "description": "Natural language description of what to extract (e.g., 'Calculate peak area and retention time')"
                },
                "force_regenerate": {
                    "type": "boolean",
                    "description": (
                        "If true, generates new analysis script even if one exists. "
                        "Use when analysis requirements change (e.g., switching from single-row to multi-row extraction, "
                        "or changing which metrics to extract). Default: false"
                    )
                },
                "inputs": {
                    "type": "array", 
                    "items": {"type": "string"},
                    "description": "List of column names to treat as INPUT parameters"
                },
                "targets": {
                    "type": "array", 
                    "items": {"type": "string"}, 
                    "description": "List of column names to treat as OPTIMIZATION TARGETS"
                }
            },
            required=["file_path"]
        )
        
        # 6b. ANALYZE BATCH
        def analyze_batch(
                file_paths: list[str],
                extraction_goal: str = None,
                conditions: str = None,
                conditions_file: str = None,
                inputs: list[str] = None,
                targets: list[str] = None,
                force_regenerate: bool = False):
            """
            Analyzes multiple raw data files in a single call.
            Runs the scalarizer on each file, merges with experimental conditions,
            and appends all results to the optimization dataset.
            """
            print(f"  ⚡ Tool: Analyzing batch of {len(file_paths)} files...")

            # --- 1. Resolve all file paths first ---
            resolved_paths = []
            resolve_errors = []
            for fp in file_paths:
                resolved, err = self._resolve_data_path(fp)
                if err:
                    resolve_errors.append({"file": fp, "error": "File not found"})
                else:
                    resolved_paths.append(resolved)

            if not resolved_paths:
                return json.dumps({
                    "status": "error",
                    "message": "None of the provided file paths could be resolved.",
                    "errors": resolve_errors
                })

            # --- 2. Parse conditions ---
            file_conditions = {}  # resolved_path -> {param: value}

            raw_conditions = None
            if conditions:
                try:
                    raw_conditions = json.loads(conditions)
                except json.JSONDecodeError as e:
                    return json.dumps({
                        "status": "error",
                        "message": f"Could not parse conditions JSON: {e}",
                        "expected_formats": [
                            '{"filename.csv": {"temperature": 300, "pH": 7.0}, ...}',
                            '[{"temperature": 300}, {"temperature": 350}]'
                        ]
                    })
            elif conditions_file:
                cf_resolved, cf_err = self._resolve_data_path(conditions_file)
                if cf_err:
                    return json.dumps({
                        "status": "error",
                        "message": f"Conditions file not found: {conditions_file}"
                    })
                try:
                    with open(cf_resolved, 'r') as f:
                        raw_conditions = json.load(f)
                except Exception as e:
                    return json.dumps({
                        "status": "error",
                        "message": f"Could not read conditions file: {e}"
                    })

            if raw_conditions is not None:
                if isinstance(raw_conditions, list):
                    # Positional matching
                    if len(raw_conditions) != len(resolved_paths):
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"Conditions list has {len(raw_conditions)} entries "
                                f"but {len(resolved_paths)} files were resolved. Counts must match."
                            )
                        })
                    for rpath, cond in zip(resolved_paths, raw_conditions):
                        if isinstance(cond, dict):
                            file_conditions[rpath] = cond
                elif isinstance(raw_conditions, dict):
                    # Key matching by filename or stem
                    unmatched_keys = []
                    for key, cond in raw_conditions.items():
                        matched = False
                        for rpath in resolved_paths:
                            fname = Path(rpath).name
                            stem = Path(rpath).stem
                            if key in (fname, stem):
                                file_conditions[rpath] = cond
                                matched = True
                                break
                        if not matched:
                            unmatched_keys.append(key)
                    if unmatched_keys and not file_conditions:
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"None of the condition keys matched any file. "
                                f"Unmatched keys: {unmatched_keys}. "
                                f"Resolved filenames: {[Path(p).name for p in resolved_paths]}"
                            )
                        })
                    if unmatched_keys:
                        print(f"    ⚠️  Unmatched condition keys (ignored): {unmatched_keys}")

            # --- 3. Process each file through scalarizer ---
            enhanced_objective = extraction_goal or ""
            if self.orch.objective and self.orch.objective != "Undefined Research Goal":
                enhanced_objective = (
                    f"Research objective: {self.orch.objective}\n\n{enhanced_objective}"
                ).strip()
            if inputs and targets:
                schema_instruction = (
                    f"\nREQUIRED OUTPUT SCHEMA:\n"
                    f"- INPUT PARAMETERS: {inputs}\n"
                    f"- TARGET METRICS: {targets}\n"
                    f"Extract EXACTLY these columns from the data."
                )
                enhanced_objective = f"{enhanced_objective}\n{schema_instruction}".strip()

            current_plan = self.orch.planner.state.get("current_plan", {})
            # `or [{}]` — the dict default only fires on a MISSING key, so a plan
            # carrying an empty experiments list would IndexError here.
            _exps = (current_plan or {}).get("proposed_experiments") or [{}]
            exp_context = _exps[0]

            if inputs and targets:
                exp_context = exp_context.copy() if exp_context else {}
                exp_context["_schema_requirements"] = {
                    "input_columns": inputs,
                    "target_columns": targets,
                    "optimization_type": "multi-objective" if len(targets) > 1 else "single-objective"
                }

            role_hints = {"inputs": inputs, "targets": targets} if inputs and targets else None

            script_to_use = None
            if not force_regenerate:
                script_to_use = self.orch.active_scalarizer_script if (
                    self.orch.active_scalarizer_script and Path(self.orch.active_scalarizer_script).exists()
                ) else None

            results = []
            errors = list(resolve_errors)
            first_column_roles = None

            for rpath in resolved_paths:
                fname = Path(rpath).name
                try:
                    res = self.orch.scalarizer.scalarize(
                        data_path=rpath,
                        objective_query=enhanced_objective,
                        reuse_script_path=script_to_use,
                        experiment_context=exp_context,
                        enable_human_review=False,
                        column_role_hints=role_hints
                    )

                    if res.get("status") != "success":
                        errors.append({"file": fname, "error": res.get("error", "Scalarizer failed")})
                        continue

                    # Lock script after first success
                    if not script_to_use and res.get("source_script"):
                        script_to_use = res["source_script"]
                        self.orch.active_scalarizer_script = script_to_use
                        print(f"    🔒 Script locked: {Path(script_to_use).name}")

                    if first_column_roles is None:
                        first_column_roles = res.get("column_roles", {})

                    # Extract metrics row
                    metrics = res.get("metrics", {})
                    if isinstance(metrics, list):
                        if len(metrics) == 1:
                            row = metrics[0]
                        else:
                            # Multi-row from a single spectrum file — take all rows
                            # but this is unusual; log a warning
                            print(f"    ⚠️  {fname}: scalarizer returned {len(metrics)} rows (expected 1)")
                            row = metrics[0]
                    elif isinstance(metrics, dict):
                        row = metrics
                    else:
                        errors.append({"file": fname, "error": f"Unexpected metrics type: {type(metrics)}"})
                        continue

                    # Merge conditions with scalarizer output.
                    # External conditions take priority — they overwrite any
                    # values the script may have hardcoded from the first sidecar.
                    cond = file_conditions.get(rpath, {})
                    merged = {**row, **cond}
                    results.append({"file": fname, "row": merged, "path": rpath})

                    print(f"    ✅ {fname}: {len(merged)} columns extracted")

                except Exception as e:
                    logging.error(f"Error processing {fname}: {e}", exc_info=True)
                    errors.append({"file": fname, "error": str(e)})

            if not results:
                return json.dumps({
                    "status": "error",
                    "message": "All files failed during processing.",
                    "errors": errors
                })

            # --- 4. Determine schema ---
            all_keys = list(results[0]["row"].keys())
            condition_keys = set()
            if file_conditions:
                for cond in file_conditions.values():
                    condition_keys.update(cond.keys())
                condition_keys = sorted(condition_keys)

            if inputs and targets:
                # User-specified schema
                self.orch.expected_input_columns = inputs
                self.orch.expected_target_columns = targets
                self._capture_input_types(first_column_roles or {}, inputs)
            elif condition_keys:
                # Derive: inputs = condition keys, targets = remaining keys
                proposed_targets = first_column_roles.get("targets", []) if first_column_roles else []
                scalarizer_keys = [k for k in all_keys if k not in condition_keys]
                if not proposed_targets:
                    proposed_targets = scalarizer_keys
                self.orch.expected_input_columns = list(condition_keys)
                self.orch.expected_target_columns = proposed_targets
                self._capture_input_types(first_column_roles or {}, list(condition_keys))
            else:
                # No conditions at all — check if scalarizer found inputs
                proposed_inputs = first_column_roles.get("inputs", []) if first_column_roles else []
                proposed_targets = first_column_roles.get("targets", []) if first_column_roles else []
                if proposed_inputs and proposed_targets:
                    self.orch.expected_input_columns = proposed_inputs
                    self.orch.expected_target_columns = proposed_targets
                    self._capture_input_types(first_column_roles or {}, proposed_inputs)
                else:
                    # Return inputs_required with example template
                    file_names = [r["file"] for r in results]
                    targets_found = proposed_targets or all_keys
                    example = {fn: {"parameter_1": "value", "parameter_2": "value"} for fn in file_names}
                    return json.dumps({
                        "status": "inputs_required",
                        "message": (
                            f"Extracted targets from {len(results)} files but no experimental "
                            f"conditions were found. Input parameters are needed for optimization."
                        ),
                        "files": file_names,
                        "targets_found": targets_found,
                        "example_conditions": example,
                        "options": [
                            "Re-call analyze_batch with conditions='{...}' mapping filenames to parameter values",
                            "Re-call analyze_batch with conditions_file='path/to/conditions.json'",
                            "Place sidecar JSON files next to each data file (e.g., spectrum_300C.json)"
                        ]
                    })

            print(f"    📊 Batch schema: inputs={self.orch.expected_input_columns}, targets={self.orch.expected_target_columns}")

            # --- 5. Validate condition key consistency across files ---
            files_missing_conditions = []
            if condition_keys:
                for r in results[:]:  # iterate over copy
                    missing_cond_keys = [k for k in condition_keys if k not in r["row"] or r["row"][k] is None]
                    if missing_cond_keys:
                        files_missing_conditions.append(r["file"])
                        results.remove(r)

            if not results and files_missing_conditions:
                example = {fn: {k: "value" for k in condition_keys} for fn in files_missing_conditions}
                return json.dumps({
                    "status": "inputs_required",
                    "message": "All files are missing experimental conditions.",
                    "files_missing_conditions": files_missing_conditions,
                    "example_conditions": example,
                    "options": [
                        "Re-call analyze_batch with conditions='{...}' providing values for all files",
                        "Re-call analyze_batch with conditions_file='path/to/conditions.json'"
                    ]
                })

            # --- 6. Build DataFrame and append to optimization_data.csv ---
            expected_cols = self.orch.expected_input_columns + self.orch.expected_target_columns

            # Filter rows to expected columns only
            clean_rows = []
            for r in results:
                row_data = {}
                for col in expected_cols:
                    row_data[col] = r["row"].get(col)
                clean_rows.append(row_data)

            df_batch = pd.DataFrame(clean_rows)

            if self.orch.bo_data_path.exists():
                df_existing = pd.read_csv(self.orch.bo_data_path)
                if set(df_batch.columns) != set(df_existing.columns):
                    if inputs or targets:
                        # Schema intentionally changed by user — replace old data
                        print(f"    🔄 Schema changed (user-specified targets). Replacing old optimization data.")
                        backup = self.orch.bo_data_path.with_suffix('.csv.backup')
                        self.orch.bo_data_path.rename(backup)
                        # Clear file tracking so all files are re-counted
                        self.orch.analyzed_files = {}
                        df_batch.to_csv(self.orch.bo_data_path, mode='w', header=True, index=False)

                        for r in results:
                            file_path_abs = str(Path(r["path"]).resolve())
                            current_hash = self._compute_file_hash(r["path"])
                            self.orch.analyzed_files[file_path_abs] = {
                                'row_count': 1, 'hash': current_hash,
                                'timestamp': datetime.now().isoformat()
                            }
                        with open(self.orch.analyzed_files_path, 'w') as f:
                            json.dump(self.orch.analyzed_files, f, indent=2)

                        df_final = pd.read_csv(self.orch.bo_data_path)
                        return json.dumps({
                            "status": "success",
                            "files_processed": len(results),
                            "rows_added": len(results),
                            "data_points_collected": len(df_final),
                            "optimization_ready": len(df_final) >= 3,
                            "inputs": self.orch.expected_input_columns,
                            "targets": self.orch.expected_target_columns,
                            "note": "Schema changed — old data replaced with new target selection"
                        })
                    return json.dumps({
                        "status": "error",
                        "message": "Schema mismatch with existing optimization data.",
                        "expected_columns": list(df_existing.columns),
                        "received_columns": list(df_batch.columns),
                        "hint": "Use reset_analysis_logic to start fresh."
                    })
                df_batch = df_batch[df_existing.columns]
                df_batch.to_csv(self.orch.bo_data_path, mode='a', header=False, index=False)
            else:
                df_batch.to_csv(self.orch.bo_data_path, mode='w', header=True, index=False)

            # Track each file
            for r in results:
                file_path_abs = str(Path(r["path"]).resolve())
                current_hash = self._compute_file_hash(r["path"])
                self.orch.analyzed_files[file_path_abs] = {
                    'row_count': 1,
                    'hash': current_hash,
                    'timestamp': datetime.now().isoformat()
                }
            with open(self.orch.analyzed_files_path, 'w') as f:
                json.dump(self.orch.analyzed_files, f, indent=2)

            df_final = pd.read_csv(self.orch.bo_data_path)
            data_count = len(df_final)
            num_added = len(results)

            response = {
                "status": "success" if not errors and not files_missing_conditions else "partial_success",
                "files_processed": len(results),
                "rows_added": num_added,
                "data_points_collected": data_count,
                "optimization_ready": data_count >= 3,
                "inputs": self.orch.expected_input_columns,
                "targets": self.orch.expected_target_columns,
            }
            if errors:
                response["errors"] = errors
            if files_missing_conditions:
                example = {fn: {k: "value" for k in condition_keys} for fn in files_missing_conditions}
                response["files_missing_conditions"] = files_missing_conditions
                response["example_conditions"] = example

            return json.dumps(response)

        self._register_tool(
            func=analyze_batch,
            name="analyze_batch",
            description=(
                "Analyzes multiple data files (e.g., spectra, time series) in a single call. "
                "Runs the scalarizer on each file to extract target metrics, then merges with "
                "experimental conditions. All results are appended to the optimization dataset. "
                "Use instead of calling analyze_file repeatedly when files share the same structure."
            ),
            parameters={
                "file_paths": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": (
                        "List of data file paths to analyze "
                        "(e.g., ['data/spectrum_300C.csv', 'data/spectrum_350C.csv'])"
                    )
                },
                "extraction_goal": {
                    "type": "string",
                    "description": "What to extract from each file (e.g., 'Calculate peak area and FWHM')"
                },
                "conditions": {
                    "type": "string",
                    "description": (
                        "JSON string mapping filenames to experimental conditions. "
                        "Dict format: {\"spectrum_300C.csv\": {\"temperature\": 300}, ...}. "
                        "List format: [{\"temperature\": 300}, {\"temperature\": 350}] (positional). "
                        "Omit if sidecar JSONs exist next to each data file."
                    )
                },
                "conditions_file": {
                    "type": "string",
                    "description": (
                        "Path to a JSON file containing the conditions mapping "
                        "(same format as the conditions parameter)"
                    )
                },
                "inputs": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of column names to treat as INPUT parameters (overrides auto-detection)"
                },
                "targets": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of column names to treat as OPTIMIZATION TARGETS (overrides auto-detection)"
                },
                "force_regenerate": {
                    "type": "boolean",
                    "description": "If true, regenerates analysis script even if one exists. Default: false"
                }
            },
            required=["file_paths"]
        )

        # 7. RESET ANALYSIS LOGIC
        def reset_analysis_logic():
            """Resets the analysis script, optimization data, AND file tracking."""
            self.orch.active_scalarizer_script = None
            self.orch.expected_input_columns = None
            self.orch.expected_target_columns = []
            
            # Clear file tracking completely
            self.orch.analyzed_files = {}
            if self.orch.analyzed_files_path.exists():
                try:
                    self.orch.analyzed_files_path.unlink()
                    print(f"    🗑️  Cleared file tracking history")
                except Exception as e:
                    logging.warning(f"Could not delete analyzed_files.json: {e}")
            
            if self.orch.bo_data_path.exists():
                backup_path = self.orch.bo_data_path.with_suffix('.csv.backup')
                self.orch.bo_data_path.rename(backup_path)
                print(f"    ⚠️  Old data backed up to: {backup_path.name}")

            bo_history = self.orch.bo.history_file
            if bo_history.exists():
                backup = bo_history.with_suffix('.json.backup')
                bo_history.rename(backup)
                print(f"    ⚠️  BO history backed up to: {backup.name}")

            return json.dumps({
                "status": "success",
                "message": "Analysis logic reset. All files will be reprocessed fresh on next analyze_file call.",
                "hint": "Previous optimization data was backed up"
            })
        
        self._register_tool(
            func=reset_analysis_logic,
            name="reset_analysis_logic",
            description=(
                "Resets the locked analysis script and clears optimization data. "
                "Use this when the current analysis approach is fundamentally wrong. "
                "Previous data is backed up before deletion."
            ),
            parameters={},
            required=[]
        )
        
        # 8. RUN OPTIMIZATION
        def run_optimization(
            parallel_capable: bool = False,
            batch_size: int = None,
            physical_constraints: str = None,
            experimental_budget: int = None,
            targets: list[str] = None,
            strategy_hint: str = None,
            skill: str = None,
            candidate_pool: str = None
        ):
            """
            Runs Bayesian Optimization to suggest next parameters.
            Supports optional physical constraints for realizable batch design
            and optional experimental budget for exploration/exploitation control.
            """
            print(f"  ⚡ Tool: Running Bayesian Optimization...")
            
            # --- PRE-FLIGHT CHECKS ---
            # A campaign ingested purely via table pass-through (#366) has
            # data and schema but never locks a script; only refuse when
            # nothing has been ingested at all.
            if (not self.orch.active_scalarizer_script
                    and not self.orch.bo_data_path.exists()):
                return json.dumps({
                    "status": "error",
                    "message": "No analysis script locked yet",
                    "hint": "Run analyze_file on at least 3 data files first",
                    "workflow": "analyze_file (×3) → run_optimization"
                })
            
            if not self.orch.bo_data_path.exists():
                return json.dumps({
                    "status": "error",
                    "message": "No optimization_data.csv found",
                    "hint": "Run analyze_file to collect data points first"
                })
            
            try:
                df = pd.read_csv(self.orch.bo_data_path)
            except Exception as e:
                return json.dumps({
                    "status": "error",
                    "message": f"Failed to read optimization data: {e}",
                    "hint": "CSV may be corrupted. Check optimization_data.csv"
                })
            
            if len(df) < 3:
                return json.dumps({
                    "status": "error", 
                    "message": f"Insufficient data points: {len(df)}/3",
                    "hint": "Collect at least 3 experimental results before optimizing",
                    "current_data_count": len(df)
                })
            
            if not self.orch.expected_target_columns or not self.orch.expected_input_columns:
                return json.dumps({
                    "status": "error",
                    "message": "Schema not established",
                    "hint": "This shouldn't happen. Try reset_analysis_logic."
                })

            # TARGET NARROWING — allows switching from MOO to SOO
            # without re-running analyze_batch on every file.
            if targets:
                missing = [t for t in targets if t not in df.columns]
                if missing:
                    return json.dumps({
                        "status": "error",
                        "message": f"Requested targets not in data: {missing}",
                        "available_columns": list(df.columns)
                    })
                old_targets = self.orch.expected_target_columns
                self.orch.expected_target_columns = targets
                print(f"    🎯 Targets narrowed: {old_targets} → {targets}")

            # SCHEMA VALIDATION
            missing_targets = [t for t in self.orch.expected_target_columns if t not in df.columns]
            if missing_targets:
                return json.dumps({
                    "status": "error",
                    "message": f"Target columns missing from data: {missing_targets}",
                    "available_columns": list(df.columns)
                })
            
            missing_inputs = [c for c in self.orch.expected_input_columns if c not in df.columns]
            if missing_inputs:
                return json.dumps({
                    "status": "error",
                    "message": f"Input columns missing: {missing_inputs}",
                    "available_columns": list(df.columns)
                })
            
            critical_cols = self.orch.expected_input_columns + self.orch.expected_target_columns
            
            if df[critical_cols].isnull().any().any():
                return json.dumps({
                    "status": "error",
                    "message": "Missing values detected in optimization data",
                    "hint": "Ensure all data files were analyzed successfully",
                    "affected_rows": df[df[critical_cols].isnull().any(axis=1)].index.tolist()
                })
            
            # ============================================
            # BOUNDS & CONSTRAINTS CALCULATION
            # ============================================
            # Pull continuous-parameter bounds and categorical-parameter levels
            # from the planner's current_plan. The planner schema supports both
            # parameter_type="continuous" (with min_value/max_value) and
            # parameter_type="categorical" (with levels). Missing parameter_type
            # is treated as continuous for backward compatibility.
            scientific_bounds: Dict[str, tuple] = {}
            planner_levels: Dict[str, List[str]] = {}
            current_plan = self.orch.planner.state.get("current_plan", {})

            if current_plan and "proposed_experiments" in current_plan:
                for exp in current_plan["proposed_experiments"]:
                    for param in exp.get("optimization_params", []):
                        name = param.get("parameter_name")
                        if not name:
                            continue
                        ptype = param.get("parameter_type", "continuous")
                        if ptype == "categorical":
                            levels = param.get("levels") or []
                            if levels:
                                planner_levels[name] = [str(lv) for lv in levels]
                                print(f"  🔬 Scientific Constraint Found: {name} ∈ {planner_levels[name]}")
                            continue
                        min_v = param.get("min_value")
                        max_v = param.get("max_value")
                        if min_v is not None and max_v is not None:
                            scientific_bounds[name] = (float(min_v), float(max_v))
                            print(f"  🔬 Scientific Constraint Found: {name} must be between {min_v} and {max_v}")

            # Resolve input types: scalarizer is the source of truth on type;
            # planner is the source of truth on the level universe (so BO can
            # recommend levels not yet observed in the data).
            input_types_state = getattr(self.orch, "expected_input_types", None) or {}

            optimization_inputs: List[str] = []
            level_maps: Dict[str, List[str]] = {}
            type_conflict_warnings: List[str] = []

            for col in self.orch.expected_input_columns:
                if col not in df.columns:
                    print(f"  ⚠️ Skipping missing input column: {col}")
                    continue

                declared_type = input_types_state.get(col)
                planner_says_cat = col in planner_levels

                # Type resolution: scalarizer wins on type because it sees the
                # actual data file. The planner's declaration of `levels` is
                # only honored when the scalarizer agrees the column is
                # categorical (or hasn't classified it). This avoids silently
                # corrupting a continuous knob whose values happen to look
                # discrete in the observed data.
                if declared_type == "continuous":
                    if planner_says_cat:
                        type_conflict_warnings.append(
                            f"{col}: scalarizer classified as continuous, planner declared "
                            "categorical — honoring scalarizer (data shape wins). To force "
                            "categorical, fix the scalarizer's input_types classification."
                        )
                    is_categorical = False
                elif declared_type == "categorical" or planner_says_cat:
                    is_categorical = True
                else:
                    is_categorical = False

                if is_categorical:
                    observed = sorted(df[col].dropna().astype(str).unique().tolist())
                    if planner_levels.get(col):
                        # Planner is authoritative on the level universe.
                        # Observed values that aren't in the planner-declared
                        # set are a real misalignment (data needs re-encoding,
                        # or the planner's levels are wrong) — fail loudly
                        # rather than silently append spurious levels.
                        levels = list(planner_levels[col])
                        unknown = [v for v in observed if v not in levels]
                        if unknown:
                            return json.dumps({
                                "status": "error",
                                "message": (
                                    f"Input column '{col}' has values {unknown} that are not "
                                    f"in the planner-declared level universe {levels}. Either "
                                    f"re-encode the data to use the declared level names, or "
                                    f"correct the planner's optimization_params.levels for "
                                    f"this parameter."
                                ),
                            })
                    else:
                        levels = observed
                    level_maps[col] = levels
                    optimization_inputs.append(col)
                else:
                    if not pd.api.types.is_numeric_dtype(df[col]):
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"Input column '{col}' is non-numeric but not declared "
                                f"categorical. Either fix the data or set its input_type "
                                f"to 'categorical' in the scalarizer output."
                            ),
                            "available_columns": list(df.columns),
                        })
                    optimization_inputs.append(col)

            for w in type_conflict_warnings:
                print(f"  ⚠️ {w}")

            if not optimization_inputs:
                return json.dumps({
                    "status": "error",
                    "message": "No usable input parameters found."
                })

            self.orch.expected_input_columns = optimization_inputs
            self.orch.expected_input_levels = level_maps if level_maps else None

            # Build bounds (continuous: scientific or data-derived; categorical: [0, n-1])
            input_bounds = []
            for col in optimization_inputs:
                if col in level_maps:
                    n = len(level_maps[col])
                    input_bounds.append([0.0, float(n - 1)])
                    print(f"     -> Bound for '{col}': [0, {n - 1}] (Source: CATEGORICAL — {n} levels)")
                elif col in scientific_bounds:
                    sci_min, sci_max = scientific_bounds[col]
                    input_bounds.append([sci_min, sci_max])
                    print(f"     -> Bound for '{col}': [{sci_min}, {sci_max}] (Source: PLANNER)")
                else:
                    data_min = float(df[col].min())
                    data_max = float(df[col].max())

                    if data_min == data_max:
                        margin = 1.0 if data_min == 0 else abs(data_min * 0.1)
                    else:
                        margin = (data_max - data_min) * 0.1

                    safe_min = data_min - margin
                    safe_max = data_max + margin

                    input_bounds.append([safe_min, safe_max])
                    print(f"     -> Bound for '{col}': [{safe_min:.2f}, {safe_max:.2f}] (Source: DATA)")

            # Build cat_dims (positional indices) and integer-encode CSV
            cat_dims = [
                i for i, c in enumerate(optimization_inputs) if c in level_maps
            ]
            bo_data_path_for_run = str(self.orch.bo_data_path)
            if level_maps:
                df_encoded = df.copy()
                for col, levels in level_maps.items():
                    idx = {lv: i for i, lv in enumerate(levels)}
                    encoded = df[col].astype(str).map(idx)
                    if encoded.isnull().any():
                        unknown = sorted(
                            df.loc[encoded.isnull(), col].astype(str).unique().tolist()
                        )
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"Unknown levels for categorical input '{col}': {unknown}. "
                                f"Known levels: {levels}."
                            )
                        })
                    df_encoded[col] = encoded.astype(float)
                encoded_dir = self.orch.base_dir / "bo_artifacts"
                encoded_dir.mkdir(exist_ok=True, parents=True)
                encoded_path = encoded_dir / "optimization_data_encoded.csv"
                df_encoded.to_csv(encoded_path, index=False)
                bo_data_path_for_run = str(encoded_path)
                print(f"  🔡 Encoded {len(level_maps)} categorical column(s) → {encoded_path.name}")
            
            # ============================================
            # BATCH SIZE DETERMINATION
            # ============================================
            if not parallel_capable:
                final_batch_size = 1
                mode_desc = "sequential (single experiment)"
            else:
                if batch_size is None:
                    return json.dumps({
                        "status": "batch_size_required",
                        "message": "Batch size must be specified for parallel optimization.",
                        "instruction": (
                            "Analyze the experimental plan to determine appropriate batch_size "
                            "(e.g., plate format, number of conditions, equipment capacity), "
                            "then call: run_optimization(parallel_capable=True, batch_size=N)"
                        ),
                        "hint": "Common values: 8, 12, 24, 96, 384 for plate-based experiments"
                    })
                
                if batch_size < 1:
                    return json.dumps({
                        "status": "error", 
                        "message": f"Invalid batch_size: {batch_size}. Must be at least 1."
                    })
                
                final_batch_size = batch_size
                mode_desc = f"parallel (batch of {batch_size})"
                print(f"    ℹ️  Using batch_size: {batch_size}")
            
            # ============================================
            # CONSTRAINT-AWARE & BUDGET-AWARE LOGGING
            # ============================================
            if physical_constraints:
                mode_desc += " + constraint-aware"
                print(f"    📐 Physical constraints provided — will use LLM-guided batch design")
            
            if experimental_budget is not None:
                mode_desc += f" + budget={experimental_budget}"
                print(f"    💰 Experimental budget: {experimental_budget} iteration(s) remaining")
            
            print(f"    📊 Optimization Setup:")
            print(f"       Mode: {mode_desc}")
            print(f"       Data points: {len(df)}")
            print(f"       Inputs: {self.orch.expected_input_columns}")
            print(f"       Targets: {self.orch.expected_target_columns}")
            print(f"       Bounds: {input_bounds}")
            if physical_constraints:
                print(f"       Constraints: {physical_constraints[:100]}...")
            
            # ============================================
            # DATA SUFFICIENCY CHECK FOR MOO
            # ============================================
            n_targets = len(self.orch.expected_target_columns)
            n_inputs = len(self.orch.expected_input_columns)
            n_data = len(df)
            if n_targets > 1:
                min_recommended = 5 * n_inputs * n_targets
                if n_data < min_recommended:
                    print(f"    ⚠️  MOO data sufficiency: {n_data} points for "
                          f"{n_inputs} inputs × {n_targets} targets (recommend ≥{min_recommended})")
                    return json.dumps({
                        "status": "warning",
                        "message": (
                            f"Multi-objective optimization with {n_targets} targets and "
                            f"{n_inputs} inputs ideally needs ≥{min_recommended} data points, "
                            f"but only {n_data} are available. The Pareto recommendations "
                            f"will be unreliable."
                        ),
                        "suggestion": (
                            f"Re-call run_optimization with targets=[\"chosen_target\"] to "
                            f"narrow to single-objective. No need to re-analyze files. "
                            f"You can switch to MOO once more data is collected."
                        ),
                        "current_targets": self.orch.expected_target_columns,
                        "data_points": n_data,
                        "recommended_minimum": min_recommended,
                    })

            try:
                # ============================================
                # DISTILL OBJECTIVE & CALL BO
                # ============================================
                bo_objective = self._distill_objective_for_bo(
                    self.orch.expected_target_columns
                )

                # Multi-fidelity: if the scalarizer declared a fidelity column
                # (and it survives into the final input set), translate it to a
                # fidelity_config keyed by the column's index in input_cols. None
                # otherwise -> standard single-fidelity BO (unchanged behavior).
                fidelity_config = None
                fspec = getattr(self.orch, "fidelity_spec", None)
                if fspec and fspec.get("column") in (self.orch.expected_input_columns or []):
                    fidelity_config = {
                        "fidelity_col": self.orch.expected_input_columns.index(fspec["column"])
                    }
                    if fspec.get("target_fidelity") is not None:
                        fidelity_config["target_fidelity"] = fspec["target_fidelity"]
                    if fspec.get("costs"):
                        fidelity_config["fidelity_costs"] = fspec["costs"]
                    print(f"    📶  Multi-fidelity active: column '{fspec['column']}' "
                          f"(index {fidelity_config['fidelity_col']})")

                # Candidate pool: resolve the CSV path here (fuzzy matching,
                # clear error to the LLM); column matching / measured-row
                # exclusion / degradation live in BOAgent._resolve_candidate_pool.
                resolved_pool = None
                if candidate_pool:
                    resolved_pool, pool_err = self._resolve_data_path(candidate_pool)
                    if pool_err:
                        return pool_err
                    print(f"    🎯 Candidate pool: {Path(resolved_pool).name}")

                res = self.orch.bo.run_optimization_loop(
                    data_path=bo_data_path_for_run,
                    objective_text=bo_objective,
                    input_cols=self.orch.expected_input_columns,
                    input_bounds=input_bounds,
                    target_cols=self.orch.expected_target_columns,
                    target_directions=self.orch.target_directions,
                    output_dir=str(self.orch.base_dir / "bo_artifacts"),
                    batch_size=int(final_batch_size),
                    physical_constraints=physical_constraints,
                    experimental_budget=experimental_budget,
                    strategy_hint=strategy_hint,
                    plot_acq=True,
                    save_acq=True,
                    cat_dims=cat_dims if cat_dims else None,
                    skill=skill,
                    fidelity_config=fidelity_config,
                    candidate_pool=resolved_pool,
                )
                
                if res.get("status") != "success":
                    return json.dumps({
                        "status": "error", 
                        "message": res.get("error", "Optimization failed"),
                        "bo_output": res
                    })
                
                # Format response
                next_params = res.get('next_parameters')

                # Decode categorical recommendations back to human-readable levels
                if level_maps and next_params is not None:
                    next_params = self._decode_categorical_recs(next_params, level_maps)

                if parallel_capable:
                    hint = f"Run all {final_batch_size} experiments in parallel, then use analyze_file on each result file."
                    params_summary = f"Generated {final_batch_size} parameter sets"
                else:
                    hint = "Run this experiment, then use analyze_file on the result to continue."
                    params_summary = "Generated next experiment parameters"
                
                response = {
                    "status": "success",
                    "mode": "parallel" if parallel_capable else "sequential",
                    "batch_size": final_batch_size,
                    "recommended_parameters": next_params,
                    "params_summary": params_summary,
                    "strategy_used": res.get('strategy', {}).get('acquisition_strategy', {}).get('type'),
                    "plot_path": res.get('plot_path'),
                    "hint": hint
                }
                if res.get("acq_plot_path"):
                    response["acq_plot_path"] = res["acq_plot_path"]
                if res.get("acq_data_path"):
                    response["acq_data_path"] = res["acq_data_path"]
                
                # Include visual inspection results
                if res.get("inspection"):
                    response["inspection"] = res["inspection"]

                # Include constrained planning metadata
                if res.get("constrained_planning"):
                    cp = res["constrained_planning"]
                    response["constraint_aware"] = True
                    response["coverage_summary"] = cp.get("coverage_summary", "")
                    response["trade_offs"] = cp.get("trade_offs", "")
                    if cp.get("validation_errors"):
                        response["constraint_warnings"] = cp["validation_errors"]

                # Candidate-pool provenance: how many library points were
                # provided / still unmeasured (absent if the pool was
                # ignored — the BO log warning explains why).
                if res.get("candidate_pool"):
                    response["candidate_pool"] = res["candidate_pool"]

                # Include budget context
                if res.get("budget"):
                    response["budget"] = res["budget"]
                
                return json.dumps(response)
                
            except Exception as e:
                logging.error(f"Optimization error: {e}")
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=run_optimization,
            name="run_optimization",
            description=(
                "Runs Bayesian Optimization to suggest next experimental parameters. "
                "Requires at least 3 data points from analyze_file. "
                "For parallel mode, batch_size must be specified. "
                "Supports optional physical_constraints for constraint-aware batch design — "
                "when provided, the agent evaluates the acquisition landscape and uses LLM "
                "reasoning to design a batch that maximizes information gain while respecting "
                "physical experimental limitations (e.g., plate layouts, discrete reagent stocks, "
                "shared equipment channels). "
                "Supports optional experimental_budget for exploration/exploitation control — "
                "pass the number of remaining optimization iterations to shift strategy from "
                "exploration (high budget) to exploitation (low budget)."
            ),
            parameters={
                "parallel_capable": {
                    "type": "boolean",
                    "description": "True if experiments can run in parallel. False for sequential (default)."
                },
                "batch_size": {
                    "type": "integer",
                    "description": (
                        "Number of parallel experiments (required if parallel_capable=True). "
                        "Infer from experimental plan (e.g., plate format, grid size, equipment capacity)."
                    )
                },
                "candidate_pool": {
                    "type": "string",
                    "description": (
                        "Path to a CSV of the finite candidate library to recommend from "
                        "(columns must include the input parameter columns; extra columns are "
                        "ignored). Recommendations are then restricted to unmeasured rows of "
                        "this library — the right mode when experiments can only be drawn from "
                        "a fixed set (a compound catalog, pre-made formulations, a measured "
                        "design library). Pass ONLY when the user or calling system explicitly "
                        "provided such a candidate file; otherwise omit — recommendations are "
                        "continuous within the plan- or data-derived bounds. Single-objective "
                        "campaigns only (ignored with a warning for multi-objective)."
                    )
                },
                "physical_constraints": {
                    "type": "string",
                    "description": (
                        "Natural language description of physical experimental constraints that "
                        "prevent arbitrary parameter combinations. When provided, the optimizer "
                        "evaluates the full acquisition landscape and uses LLM reasoning to design "
                        "a realizable batch. Examples:\n"
                        "- '96-well plate: rows share temperature (8 values), columns share pH (12 values)'\n"
                        "- 'Only 5 catalyst concentrations available: 0.1, 0.5, 1.0, 2.0, 5.0 mM'\n"
                        "- 'Reactor has 4 zones with independent temp but shared pressure'\n"
                        "- 'Gradient limited to linear ramp: min at well A1, max at well H12'\n"
                        "If not provided, standard unconstrained BO is used."
                    )
                },
                "experimental_budget": {
                    "type": "integer",
                    "description": (
                        "Number of remaining optimization iterations (including this one). "
                        "Controls exploration-vs-exploitation balance:\n"
                        "- 1 = final shot (pure exploitation, no exploration)\n"
                        "- 2-3 = critical budget (strongly favor exploitation)\n"
                        "- Higher values = scaled based on campaign progress\n"
                        "- Omit for no budget constraint (default behavior).\n"
                        "Pass when user mentions remaining experiments, budget, 'last round', "
                        "or 'N more tries'. This counts iterations (calls to run_optimization), "
                        "not individual experiments within a batch."
                    )
                },
                "targets": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": (
                        "Subset of target columns to optimize. Use to narrow from "
                        "multi-objective to single-objective without re-analyzing files. "
                        "Example: targets=['Peak_Area'] to optimize only peak area when "
                        "the scalarizer extracted multiple metrics. The specified targets "
                        "must exist in the optimization data."
                    )
                },
                "strategy_hint": {
                    "type": "string",
                    "description": (
                        "Optional user preference for BO strategy. Pass when the user "
                        "requests a specific kernel, acquisition function, or noise prior. "
                        "Examples: 'use RBF kernel', 'try Thompson sampling', "
                        "'switch to Matern-1.5', 'use UCB with high exploration'. "
                        "The hint is respected unless it conflicts with budget constraints."
                    )
                },
                "skill": {
                    "type": "string",
                    "description": _build_optimization_skill_description()
                }
            },
            required=[]
        )


        # 9. SAVE FILE
        def save_file(filename: str, content: str, subfolder: str = "",
                      deliverable: bool = False, title: str = ""):
            """
            Save text content (code, protocols, notes) to a file in the session
            directory.
            """
            print(f"  ⚡ Tool: Saving file '{filename}'...")

            # Sanitise: strip path separators from filename to prevent traversal.
            safe_name = Path(filename).name
            if not safe_name:
                return json.dumps({
                    "status": "error",
                    "message": "Invalid filename.",
                })

            # Delegation-scoped, like every other artifact this turn writes
            # (plan.json, white_paper.md, ...). Rooting at base_dir instead
            # scattered LLM-saved files OUTSIDE the delegation directory —
            # live, a white paper landed in planning/<slug>/ as a sibling of
            # delegations/, duplicating the copy already written inside it.
            target_dir = self._output_dir()
            if subfolder:
                safe_sub = Path(subfolder).name
                target_dir = target_dir / safe_sub
            target_dir.mkdir(parents=True, exist_ok=True)
            dest = target_dir / safe_name

            try:
                dest.write_text(content, encoding="utf-8")
                from .user_interface import format_path, record_deliverable
                record_deliverable(self.orch.base_dir, dest, title,
                                   deliverable)
                print(f"    💾 Saved{' (deliverable)' if deliverable else ''}: "
                      f"{format_path(dest)}")
                # Same invariant as edit_file and the revision branch: a
                # markdown write never leaves a stale PDF twin beside it
                # (live, an agent rewrote a document via save_file and its
                # forwarded PDF kept serving the old content).
                pdf_refreshed = self._refresh_pdf_twin(dest)
                return json.dumps({
                    "status": "success",
                    "path": str(dest),
                    "size_bytes": dest.stat().st_size,
                    "deliverable": bool(deliverable),
                    "pdf_refreshed": pdf_refreshed,
                })
            except Exception as e:
                logging.error(f"save_file failed: {e}")
                return json.dumps({
                    "status": "error",
                    "message": str(e),
                })

        self._register_tool(
            func=save_file,
            name="save_file",
            description=(
                "Save text content (notes, small scripts, content you have "
                "already composed) to a NEW file "
                "in the session directory. To change an EXISTING document, "
                "do not rewrite it with save_file — use edit_file for a "
                "snippet swap, rename_file to change its filename, or "
                "write_technical_document with revise_path "
                "for a content revision; those paths keep the content "
                "byte-safe and guard against accidental truncation. "
                "Large content may not survive the "
                "trip as a single tool-call argument — for anything long "
                "(roughly >100 lines), save the first chunk with save_file and "
                "the rest with append_file. For executable analysis/"
                "optimization code, prefer generate_implementation_code, which "
                "generates and persists the script itself."
            ),
            parameters={
                "filename": {
                    "type": "string",
                    "description": (
                        "Name of the file to create, e.g. 'extraction_protocol.py' "
                        "or 'notes.txt'."
                    ),
                },
                "content": {
                    "type": "string",
                    "description": "The text content to write to the file.",
                },
                "subfolder": {
                    "type": "string",
                    "description": (
                        "Optional subfolder within the session directory, "
                        "e.g. 'protocols' or 'scripts'. Created if it doesn't exist."
                    ),
                },
                "deliverable": {
                    "type": "boolean",
                    "description": (
                        "Set TRUE when this file IS the artifact the user "
                        "asked for (a brief, a report, a protocol they "
                        "requested) rather than a working note or "
                        "intermediate. Deliverables are shown to the user "
                        "directly — starred in the file list and previewed "
                        "in the chat — so they do not have to hunt through "
                        "the session folder for them."
                    ),
                },
                "title": {
                    "type": "string",
                    "description": (
                        "Short human label for the file, e.g. 'Top-3 "
                        "priority brief'. Shown beside it in the file list "
                        "and as the preview heading."
                    ),
                },
            },
            required=["filename", "content"],
        )

        # 9b. APPEND FILE
        def append_file(filename: str, content: str, subfolder: str = ""):
            """
            Append text content to a file in the session directory (created
            if it doesn't exist). Companion to save_file for chunked writes
            of large content.
            """
            print(f"  ⚡ Tool: Appending to file '{filename}'...")

            safe_name = Path(filename).name
            if not safe_name:
                return json.dumps({
                    "status": "error",
                    "message": "Invalid filename.",
                })

            # Delegation-scoped, like every other artifact this turn writes
            # (plan.json, white_paper.md, ...). Rooting at base_dir instead
            # scattered LLM-saved files OUTSIDE the delegation directory —
            # live, a white paper landed in planning/<slug>/ as a sibling of
            # delegations/, duplicating the copy already written inside it.
            target_dir = self._output_dir()
            if subfolder:
                safe_sub = Path(subfolder).name
                target_dir = target_dir / safe_sub
            target_dir.mkdir(parents=True, exist_ok=True)
            dest = target_dir / safe_name

            try:
                with open(dest, "a", encoding="utf-8") as f:
                    f.write(content)
                print(f"    💾 Appended: {dest}")
                return json.dumps({
                    "status": "success",
                    "path": str(dest),
                    "size_bytes": dest.stat().st_size,
                })
            except Exception as e:
                logging.error(f"append_file failed: {e}")
                return json.dumps({
                    "status": "error",
                    "message": str(e),
                })

        self._register_tool(
            func=append_file,
            name="append_file",
            description=(
                "Append text content to a file in the session directory "
                "(created if it doesn't exist). Use together with save_file "
                "to write large files in chunks — save_file for the first "
                "chunk, then append_file for each subsequent chunk — keeping "
                "every chunk small enough to pass reliably as a tool argument."
            ),
            parameters={
                "filename": {
                    "type": "string",
                    "description": "Name of the file to append to.",
                },
                "content": {
                    "type": "string",
                    "description": "The text content to append to the file.",
                },
                "subfolder": {
                    "type": "string",
                    "description": (
                        "Optional subfolder within the session directory, "
                        "e.g. 'protocols' or 'scripts'. Created if it doesn't exist."
                    ),
                },
            },
            required=["filename", "content"],
        )

        # 9c. EDIT FILE — surgical in-place replacement (single or batched)
        def edit_file(path: str, old_text: str = None, new_text: str = None,
                      replace_all: bool = False, edits: list = None):
            """
            Mechanical in-place edit of an existing session document:
            replace exact text snippets — one old/new pair, or a batched
            `edits` list applied atomically in one call. Content revisions
            go through write_technical_document(revise_path=...).
            """
            print(f"  ⚡ Tool: Editing file '{path}'...")
            try:
                from ...utils.file_edit import apply_surgical_edits
                from .user_interface import format_path, record_deliverable
                if not edits:
                    if old_text is None or new_text is None:
                        return json.dumps({
                            "status": "error",
                            "message": ("Provide either old_text+new_text "
                                        "or a non-empty edits list.")})
                    edits = [{"old_text": old_text, "new_text": new_text,
                              "replace_all": replace_all}]
                rp = Path(path)
                if not rp.is_absolute():
                    rp = self._output_dir() / rp
                rp = rp.resolve()
                root = Path(self.orch.base_dir).resolve()
                d = self._output_dir()
                # Guards + backup live in the shared core; this wrapper owns
                # path resolution, the planning-specific routing hint below,
                # deliverable recording, and the PDF-twin invariant.
                out = apply_surgical_edits(
                    rp, edits,
                    root=root, backup_dir=d,
                    too_large_message=(
                        "Edit too large for edit_file. For a content "
                        "revision use write_technical_document with "
                        "revise_path (whole-document rewrite under its "
                        "length guard). For a VERBATIM insertion, split "
                        "the text at a unique boundary into consecutive "
                        "smaller edit_file calls — do not rewrite the "
                        "whole file with save_file, which keeps no "
                        "backup and has no truncation guard."),
                )
                if out["status"] != "success":
                    return json.dumps(out)
                # Audit trail: ONE chain-origin backup per file per
                # delegation, recorded once (live: an 18-edit chain filed
                # 18 near-identical backups, each re-embedded by the UI).
                if out.get("backup_created") and out.get("previous_version"):
                    who = ("" if d.resolve() == root
                           else f" (edited by {d.name})")
                    record_deliverable(self.orch.base_dir,
                                       Path(out["previous_version"]),
                                       f"Pre-edit copy of {rp.name}{who}")
                n = out.get("n_edits", 1)
                print(f"    ✏️  Edited in place"
                      f"{f' ({n} edits)' if n > 1 else ''}: "
                      f"{format_path(rp)}")
                out["pdf_refreshed"] = self._refresh_pdf_twin(rp)
                return json.dumps(out)
            except Exception as e:
                logging.error(f"edit_file failed: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=edit_file,
            name="edit_file",
            description=(
                "Surgically edit an existing session document IN PLACE by "
                "replacing exact text snippets — an image reference, a "
                "path, a caption, a typo, a parameter value. Mechanical "
                "edits only: copy old_text VERBATIM from the file (read_file "
                "first), and each snippet is capped at 2000 characters. "
                "Several changes to ONE file go in a single call as the "
                "`edits` list (applied atomically, in order) — never as a "
                "chain of one-edit calls, which burns the turn's tool "
                "budget. A document-wide STYLE pass (readability, tone, "
                "consistency) is a content revision even though each "
                "change looks mechanical — use write_technical_document "
                "with revise_path for that, as for any prose rewrite or "
                "restructuring. If the document has a PDF twin beside it, "
                "the PDF is re-exported automatically so it never goes "
                "stale."
            ),
            parameters={
                "path": {
                    "type": "string",
                    "description": (
                        "Path of the file to edit — absolute, or relative "
                        "to the current output directory. Must be inside "
                        "the session directory."
                    ),
                },
                "old_text": {
                    "type": "string",
                    "description": (
                        "Single-edit form: the exact snippet to replace, "
                        "copied VERBATIM from the file including "
                        "whitespace. Must match exactly one place unless "
                        "replace_all is true. Omit when passing `edits`."
                    ),
                },
                "new_text": {
                    "type": "string",
                    "description": ("Single-edit form: the replacement "
                                    "text. Omit when passing `edits`."),
                },
                "replace_all": {
                    "type": "boolean",
                    "description": (
                        "Single-edit form: replace every occurrence "
                        "instead of requiring the snippet to be unique. "
                        "Default false."
                    ),
                },
                "edits": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "old_text": {"type": "string"},
                            "new_text": {"type": "string"},
                            "replace_all": {"type": "boolean"},
                        },
                        "required": ["old_text", "new_text"],
                    },
                    "description": (
                        "Batched form: ALL the changes for this file in "
                        "one call. Applied in order, each against the "
                        "already-edited text; all-or-nothing — if one "
                        "edit fails to match, nothing is applied and the "
                        "error names which edit failed."
                    ),
                },
            },
            required=["path"],
        )

        # 9d. RENAME FILE — byte-exact, never through the LLM
        def rename_file(path: str, new_name: str, copy: bool = False,
                        deliverable: bool = False, title: str = ""):
            """
            Rename an existing session file byte-exactly in its own
            directory, or COPY one into the current output directory.
            Exists because the alternative was observed live: with no
            rename tool, the agent reconstructed a 30 KB document via
            save_file + append_file chunks, dropping content and leaving
            divergent duplicates.
            """
            verb_now = "Copying" if copy else "Renaming"
            print(f"  ⚡ Tool: {verb_now} file '{path}' → '{new_name}'...")
            try:
                from ...utils.file_edit import rename_or_copy_file
                from .user_interface import format_path, record_deliverable
                out_dir = Path(self._output_dir()).resolve()
                rp = Path(path)
                if not rp.is_absolute():
                    rp = out_dir / rp
                rp = rp.resolve()
                # A bare target name keeps a RENAME where the file lives —
                # it changes identity, not location (moving between
                # delegation folders is how phantom nested copies happen).
                # A COPY lands in the CURRENT output directory instead, so
                # a consolidating delegation can bring a figure written by
                # an earlier one alongside its own document and have the
                # embed resolve. The destination is never agent-chosen:
                # it is this delegation's folder or nowhere, which also
                # stops a copy from writing into a sibling delegation.
                safe_name = Path(new_name).name
                if not safe_name:
                    return json.dumps({"status": "error",
                                       "message": "Invalid new_name."})
                dest = ((out_dir if copy else rp.parent) / safe_name).resolve()
                # An error here is a routing decision point: the agent
                # asked to bring a file somewhere and needs to know which
                # tool does that, not merely that two paths matched.
                if rp == dest:
                    if not copy and rp.parent != out_dir:
                        return json.dumps({"status": "error", "message": (
                            f"A rename keeps the file in its own folder "
                            f"({rp.parent}), so this call changes nothing. "
                            f"To bring it alongside the document you are "
                            f"writing, call again with copy=true — it lands "
                            f"in {out_dir}. To leave it where it is, "
                            f"reference it by a relative path instead.")})
                    return json.dumps({"status": "error", "message": (
                        f"'{safe_name}' is already this file's name in "
                        f"{dest.parent} — nothing to do. Give a different "
                        f"new_name, or reference the file as it is.")})
                out = rename_or_copy_file(
                    rp, dest,
                    root=Path(self.orch.base_dir).resolve(), copy=copy)
                if out["status"] != "success":
                    return json.dumps(out)
                verb = "Copied" if copy else "Renamed"
                print(f"    📛 {verb}: {format_path(Path(out['path']))}")
                if out["pdf_twin_followed"]:
                    print("    📄 PDF twin followed the "
                          f"{'copy' if copy else 'rename'}")
                record_deliverable(
                    self.orch.base_dir, Path(out["path"]),
                    title or f"{verb} from {rp.name}", deliverable)
                return json.dumps(out)
            except Exception as e:
                logging.error(f"rename_file failed: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=rename_file,
            name="rename_file",
            description=(
                "Rename or copy an existing session file BYTE-EXACTLY — "
                "the content never passes through the model. Two uses: "
                "RENAME (default) lands a document under its intended "
                "filename, in its own folder; COPY (copy=true) brings a "
                "file into the folder you are writing in now, which is "
                "how you embed a figure an earlier delegation produced "
                "and have the reference resolve beside your document. "
                "Never reconstruct a file with save_file/append_file to "
                "rename or move it: that loses content and leaves "
                "divergent duplicates. A markdown document's PDF twin "
                "follows automatically."
            ),
            parameters={
                "path": {
                    "type": "string",
                    "description": (
                        "File to rename — absolute, or relative to the "
                        "current output directory. Must be inside the "
                        "session directory."
                    ),
                },
                "new_name": {
                    "type": "string",
                    "description": (
                        "Target filename (bare name, no directories). "
                        "Keep the source's own name when copying a figure "
                        "you are about to embed."
                    ),
                },
                "copy": {
                    "type": "boolean",
                    "description": (
                        "TRUE keeps the original and puts a copy in the "
                        "folder you are writing in now — use it to bring "
                        "an asset from an earlier delegation alongside "
                        "your document. FALSE (default) renames in place; "
                        "prefer a true rename for documents, since "
                        "duplicates diverge."
                    ),
                },
                "deliverable": {
                    "type": "boolean",
                    "description": (
                        "Set TRUE when the renamed file IS the artifact "
                        "the user asked for, so it is starred in the "
                        "files list."
                    ),
                },
                "title": {
                    "type": "string",
                    "description": (
                        "Short human label for the file in the files "
                        "list (e.g. its document title)."
                    ),
                },
            },
            required=["path", "new_name"],
        )

        # 10. SAVE CHECKPOINT
        def save_checkpoint():
            """
            Saves complete orchestrator state including conversation and agent state.
            Use this periodically during long campaigns.
            """
            checkpoint_path = self.orch.base_dir / "checkpoint.json"
            
            # Calculate data points
            data_points = 0
            if self.orch.bo_data_path.exists():
                try:
                    df = pd.read_csv(self.orch.bo_data_path)
                    data_points = len(df)
                except Exception:
                    pass
            
            # Get message count (handle both OpenAI and Gemini)
            if self.orch.use_openai:
                # OpenAI: messages is a list attribute
                message_count = len(self.orch.messages)
            else:
                # Gemini: history is in chat_session
                try:
                    message_count = len(self.orch.chat_session.history) if hasattr(self.orch.chat_session, 'history') else 0
                except Exception:
                    message_count = 0
            
            state = {
                "timestamp": datetime.now().isoformat(),
                "objective": self.orch.objective,
                "active_scalarizer_script": self.orch.active_scalarizer_script,
                "expected_input_columns": self.orch.expected_input_columns,
                "expected_target_columns": self.orch.expected_target_columns,
                "data_points_collected": data_points,
                "message_count": message_count,
                "planner_state": self.orch.planner.state if hasattr(self.orch.planner, 'state') else None,
                "latest_tea_results": self.orch.latest_tea_results,
                "autonomy_level": self.orch.autonomy_level.value if hasattr(self.orch, 'autonomy_level') and self.orch.autonomy_level else None,
                "data_dir": str(self.orch.data_dir) if self.orch.data_dir else None,
                "knowledge_dir": str(self.orch.knowledge_dir) if self.orch.knowledge_dir else None,
                "code_dir": str(self.orch.code_dir) if self.orch.code_dir else None,
            }
            
            try:
                with open(checkpoint_path, 'w') as f:
                    json.dump(state, f, indent=2)
                
                print(f"    💾 Checkpoint saved: {checkpoint_path}")
                
                result = {
                    "status": "success",
                    "checkpoint_path": str(checkpoint_path),
                    "data_points": data_points,
                    "message_count": message_count,
                    "timestamp": state["timestamp"]
                }

                # Check if knowledge synthesis might be valuable
                planner_state = self.orch.planner.state if self.orch.planner.state else {}
                plan_history = planner_state.get("plan_history", [])
                iterations_with_results = len(planner_state.get("experimental_results", []))
                existing_knowledge = len(self.orch.active_knowledge)

                if iterations_with_results >= 2 and existing_knowledge == 0:
                    result["knowledge_synthesis_available"] = True
                    result["plan_iterations_with_results"] = iterations_with_results
                elif existing_knowledge > 0 and iterations_with_results > existing_knowledge:
                    result["knowledge_update_available"] = True
                    result["unsynthesized_iterations"] = iterations_with_results - existing_knowledge

                return json.dumps(result)
                
            except Exception as e:
                logging.error(f"Checkpoint save failed: {e}")
                return json.dumps({
                    "status": "error",
                    "message": f"Failed to save checkpoint: {e}"
                })
        
        # Register the tool
        self._register_tool(
            func=save_checkpoint,
            name="save_checkpoint",
            description=(
                "Saves complete campaign state including conversation history, "
                "analysis scripts, and optimization data. Use this periodically "
                "during long campaigns (every 3-5 experiments) to enable resumption "
                "after crashes or breaks."
            ),
            parameters={},
            required=[]
        )

        # 10. DISCARD PLAN
        def discard_plan(reason: str = ""):
            """
            Discards the most recent experimental plan (marks it as superseded).
            The plan remains in history for transparency but won't appear in reports.
            
            Args:
                reason: Why the plan is being discarded
            """
            if not self.orch.planner.state:
                return json.dumps({
                    "status": "error",
                    "message": "No active planning session"
                })
            
            history = self.orch.planner.state.get("plan_history", [])
            
            if not history:
                return json.dumps({
                    "status": "error",
                    "message": "No plans in history to discard"
                })
            
            # Find last non-TEA, non-superseded entry
            for i in range(len(history) - 1, -1, -1):
                plan = history[i]
                if (plan.get("type") != "technoeconomic_analysis" and 
                    plan.get("status") != "superseded"):
                    
                    # Mark as superseded instead of deleting
                    plan["status"] = "superseded"
                    plan["superseded_reason"] = reason if reason else "Plan replaced with corrected version"
                    plan["superseded_at"] = datetime.now().isoformat()
                    
                    print(f"    🗑️  Discarded plan: iteration {plan.get('iteration')}")
                    if reason:
                        print(f"       Reason: {reason}")
                    
                    return json.dumps({
                        "status": "success",
                        "message": f"Plan from iteration {plan.get('iteration')} discarded",
                        "reason": plan["superseded_reason"],
                        "hint": "The discarded plan remains in history for transparency"
                    })
            
            return json.dumps({
                "status": "error",
                "message": "No active experimental plans to discard"
            })

        # Register the tool
        self._register_tool(
            func=discard_plan,
            name="discard_plan",
            description=(
                "Discards the most recent experimental plan (marks it as superseded). "
                "The plan remains in full history for transparency but won't appear in final reports. "
                "Use when correcting a wrong plan before generating the corrected version."
            ),
            parameters={
                "reason": {
                    "type": "string",
                    "description": (
                        "Why the plan is being discarded. Be specific about the mismatch. "
                        "Examples: 'Wrong material - data has Mg not Mn', "
                        "'User requested different equipment', 'Incorrect objective interpretation'"
                    )
                }
            },
            required=["reason"]
        )

        def show_directory_guide():
            """
            Shows the recommended directory structure for optimal agent performance.
            """
            guide = """
        ╔══════════════════════════════════════════════════════════════════════════╗
        ║                  RECOMMENDED DIRECTORY STRUCTURE                         ║
        ╚══════════════════════════════════════════════════════════════════════════╝

        📁 my_research_project/          ← Run orchestrator from here
        │
        ├── 📚 papers/                    ← Scientific papers & literature
        │   ├── separation_methods_2024.pdf
        │   ├── lithium_extraction_review.pdf
        │   └── rare_earth_recovery.pdf
        │
        ├── 📊 experimental_results/      ← Raw experimental data files
        │   ├── batch_001.csv
        │   ├── batch_002.csv
        │   ├── batch_003.csv
        │   └── pilot_run_*.xlsx
        │
        ├── 💻 code/                      ← Analysis scripts & API docs (optional)
        │   ├── analysis_pipeline.py
        │   ├── visualization.py
        │   └── api_documentation/
        │
        ├── 📁 campaign_session/          ← Created automatically by orchestrator
        │   ├── optimization_data.csv    (collected metrics)
        │   ├── analysis_artifacts/      (generated analysis scripts)
        │   ├── bo_artifacts/            (optimization plots)
        │   ├── plan.json                (experimental plans)
        │   └── checkpoint.json          (saved state)
        │
        └── 🗂️ kb_storage/                ← Created automatically
            ├── default_kb_docs/         (knowledge base from papers)
            └── default_kb_code/         (knowledge base from code)

        ╔══════════════════════════════════════════════════════════════════════════╗
        ║                           QUICK START GUIDE                              ║
        ╚══════════════════════════════════════════════════════════════════════════╝

        CHAT EXAMPLES:

        📋 Generate plan with papers:
        "Generate a plan for lithium extraction using ./papers/ and ./code/"

        📊 Analyze experimental data:
        "Analyze ./experimental_results/batch_001.csv and extract yield"

        🔬 Run optimization:
        "Run optimization to suggest next experiments"

        💾 Save progress:
        "Save checkpoint"
        """
            
            print(guide)
            
            # Also return as JSON for the LLM
            return json.dumps({
                "status": "success",
                "message": "Directory structure guide displayed",
                "recommended_folders": ["papers/", "experimental_results/", "code/"],
                "auto_created_folders": ["campaign_session/", "kb_storage/"]
            })

        # Register the tool
        self._register_tool(
            func=show_directory_guide,
            name="show_directory_guide",
            description=(
                "Shows recommended directory structure for optimal agent performance. "
                "Use when user asks about setup, organization, or how to structure their project."
            ),
            parameters={},
            required=[]
        )

        # =====================================================================
        # READ FILE (non-destructive inspection)
        # =====================================================================
        # A literature report is read for its whole content, never its head:
        # its questions sit in sequence, so a head-only read hides every
        # question but the first. Live, an agent read one seven times, decided
        # the file "repeats" Q1, and wrote the note without Q2-Q4 — 70% of the
        # retrieval unused, recorded as [TBD] as though the literature were
        # thin. These files get the whole body up to a cap.
        # Cap sized against the real corpus: session lit reports run 89k-262k
        # chars (median 169k), so a smaller cap would exclude precisely the
        # multi-question reports that most need a whole read. Past the cap the
        # read truncates but emits the section outline, so the agent can
        # offset= to each question instead of being stuck at the head.
        _FULL_READ_STEMS = ("literature_search",)
        _FULL_READ_MAX_CHARS = 250_000

        def read_file(file_path: str, max_lines: int = 200,
                      tail: bool = False, search: str = None,
                      offset: int = None) -> str:
            """
            Read and return the contents of a file. Use this to inspect
            plans, protocols, configs, logs, or any text/JSON file without
            triggering analysis pipelines.

            Reading from the top was the only mode, and the truncation notice
            named what was missing without naming a way to get it — so a
            question about the END of a long document ("does this paper close
            with a References section?") had no answer at any parameter value.
            Live, an agent asked it five times and gave up. `tail` and
            `search` answer those questions in one call.
            """
            print(f"  ⚡ Tool: Reading file '{file_path}'...")

            # Resolve path
            resolved, error = self._resolve_data_path(file_path)
            if error:
                return error

            path = Path(resolved)
            if not path.is_file():
                return json.dumps({
                    "status": "error",
                    "message": f"Not a file: {file_path}"
                })

            try:
                ext = path.suffix.lower()

                # Size guard — skip for Excel/CSV since we cap at 50 rows × 40 cols.
                # Documents get more headroom: extraction is page-based and a
                # figure-heavy PDF is megabytes of images, not of text.
                if ext not in ('.xlsx', '.xls', '.csv'):
                    size_mb = path.stat().st_size / (1024 * 1024)
                    cap_mb = 25 if ext in ('.pdf', '.docx') else 5
                    if size_mb > cap_mb:
                        return json.dumps({
                            "status": "error",
                            "message": f"File too large ({size_mb:.1f} MB)."
                        })

                if ext == ".json":
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    content = json.dumps(data, indent=2)
                elif ext in ('.xlsx', '.xls', '.csv'):
                    MAX_PREVIEW_ROWS = 100
                    MAX_PREVIEW_COLS = 40
                    MAX_PREVIEW_CHARS = 30000
                    if ext == '.csv':
                        df_preview = pd.read_csv(path, nrows=MAX_PREVIEW_ROWS)
                        with open(path) as _f:
                            total_rows = sum(1 for _ in _f) - 1
                    else:
                        df_preview = pd.read_excel(path, nrows=MAX_PREVIEW_ROWS)
                        try:
                            import openpyxl
                            _wb = openpyxl.load_workbook(path, read_only=True)
                            total_rows = _wb.active.max_row - 1
                            _wb.close()
                        except Exception:
                            total_rows = len(df_preview)
                    total_cols = len(df_preview.columns)
                    display_df = df_preview.iloc[:, :MAX_PREVIEW_COLS]
                    preview_text = display_df.to_string()
                    # Adaptive row reduction if output exceeds char budget
                    if len(preview_text) > MAX_PREVIEW_CHARS and len(display_df) > 5:
                        ratio = MAX_PREVIEW_CHARS / len(preview_text)
                        fewer_rows = max(5, int(len(display_df) * ratio))
                        display_df = display_df.iloc[:fewer_rows]
                        preview_text = display_df.to_string()
                        if len(preview_text) > MAX_PREVIEW_CHARS:
                            preview_text = preview_text[:MAX_PREVIEW_CHARS] + "\n... (truncated)"
                    shown_rows = len(display_df)
                    shown_cols = len(display_df.columns)
                    trunc_parts = []
                    if shown_rows < total_rows:
                        trunc_parts.append(f"first {shown_rows} rows")
                    if shown_cols < total_cols:
                        trunc_parts.append(f"first {shown_cols} columns")
                    trunc = f" (showing {', '.join(trunc_parts)})" if trunc_parts else ""
                    content = f"Shape: {total_rows} rows × {total_cols} columns{trunc}\n\n{preview_text}"
                else:
                    doc_meta = {}
                    if ext in ('.pdf', '.docx'):
                        # #397 phase 0: opened as text, a PDF returns its
                        # compressed byte streams — live, a delegation handed
                        # a proposal PDF could not ground on it (one run
                        # proceeded from the caller's paraphrase; another
                        # refused an adversarial review outright). Extraction
                        # is ALREADY shared infrastructure; route through it.
                        from ...parsers.extract import extract_text
                        info = extract_text(
                            str(path),
                            ocr_model=getattr(self.orch.planner, "model",
                                              None))
                        raw = info.get("text") or ""
                        if not raw.strip():
                            return json.dumps({
                                "status": "error",
                                "message": (
                                    f"No extractable text in {path.name} "
                                    "(empty or image-only document).")})
                        lines = raw.splitlines(keepends=True)
                        doc_meta = {k: info[k] for k in
                                    ("n_pages", "n_ocr_pages", "n_paragraphs")
                                    if info.get(k) is not None}
                        doc_meta["extracted"] = ext.lstrip(".")
                    else:
                        with open(path, 'r', encoding='utf-8',
                                  errors='replace') as f:
                            lines = f.readlines()
                    total = len(lines)

                    if search:
                        # The real question behind most repeat reads is "is X
                        # in here, and where" — a search, not a read. Answering
                        # it directly costs one call and stays cheap however
                        # long the file is.
                        try:
                            rx = re.compile(search, re.I)
                        except re.error as e:
                            return json.dumps({
                                "status": "error",
                                "message": f"Invalid search pattern: {e}"})
                        hits = [i for i, ln in enumerate(lines) if rx.search(ln)]
                        CAP = 40
                        shown, out = hits[:CAP], []
                        for i in shown:
                            lo, hi = max(0, i - 1), min(total, i + 2)
                            out.append(f"@@ line {i + 1}\n"
                                       + "".join(lines[lo:hi]).rstrip("\n"))
                        body = "\n\n".join(out) if out else "(no matches)"
                        note = (f"{len(hits)} matching line(s) in {total} total"
                                + (f"; showing the first {CAP}" if len(hits) > CAP
                                   else ""))
                        return json.dumps({
                            "status": "success",
                            "file_path": str(path),
                            "mode": "search",
                            "pattern": search,
                            "matches": len(hits),
                            "match_lines": [i + 1 for i in shown],
                            "total_lines": total,
                            "content": f"{note}\n\n{body}",
                            **doc_meta,
                        })

                    # A truncated read must say what it is missing and where,
                    # or the agent cannot tell a short file from a short read.
                    def _outline():
                        heads = [(i + 1, ln.strip()) for i, ln in
                                 enumerate(lines) if ln.startswith('#')]
                        if len(heads) < 2:
                            return ""
                        return "\nSections: " + " · ".join(
                            f"{h.lstrip('# ')[:44]} @ line {n}"
                            for n, h in heads[:12]) + (
                            " …" if len(heads) > 12 else "")

                    # Files read for their whole content, not their head.
                    whole = (any(s in path.name.lower()
                                 for s in _FULL_READ_STEMS)
                             and offset is None and not tail
                             and len("".join(lines)) <= _FULL_READ_MAX_CHARS)

                    truncated = True
                    if whole or total <= max_lines:
                        first, last, content = 1, total, "".join(lines)
                        truncated = False
                    elif offset is not None:
                        start = max(0, offset - 1)
                        shown = lines[start:start + max_lines]
                        first, last = start + 1, min(total, start + max_lines)
                        content = "".join(shown) + (
                            f"\n... (showing lines {first}-{last} of {total}."
                            f"{_outline()})")
                    elif tail:
                        shown = lines[-max_lines:]
                        first, last = total - max_lines + 1, total
                        more = (f"... ({first - 1} earlier lines not shown; "
                                f"omit tail to read from the top)")
                        content = more + "\n" + "".join(shown)
                    else:
                        shown = lines[:max_lines]
                        first, last = 1, max_lines
                        content = "".join(shown) + (
                            f"\n... ({total - max_lines} more lines not "
                            f"shown — this is a TRUNCATED READ, not the whole "
                            f"file. Jump to any part with offset=<line>; read "
                            f"the END with tail=true; find something with "
                            f"search='<pattern>'; or raise max_lines."
                            f"{_outline()})")

                    return json.dumps({
                        "status": "success",
                        "file_path": str(path),
                        "mode": "tail" if tail else "head",
                        "total_lines": total,
                        "shown_lines": f"{first}-{last}",
                        "truncated": truncated,
                        "content": content,
                        **doc_meta,
                    })

                return json.dumps({
                    "status": "success",
                    "file_path": str(path),
                    "content": content
                })

            except Exception as e:
                return json.dumps({
                    "status": "error",
                    "message": f"Failed to read file: {e}"
                })

        self._register_tool(
            func=read_file,
            name="read_file",
            description=(
                "Read a text or JSON file — plans, protocols, scripts, "
                "configs, logs, documents. PDF and Word documents are "
                "extracted to text automatically (tables preserved, scanned "
                "pages OCR'd), so uploaded papers and proposals read like "
                "any text file. Reads from the TOP by default; "
                "literature-search reports are returned WHOLE. "
                "For a long file do not read it repeatedly hoping to see more — "
                "you will get the same lines back. A truncated read lists the "
                "section headings and their line numbers: use offset=<line> to "
                "jump to one, search='<pattern>' to find where something is "
                "(and whether it is there at all), or tail=true to read the "
                "END. Do NOT use analyze_file for reading — that triggers the "
                "scalarizer pipeline."
            ),
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Path to the file to read"
                },
                "max_lines": {
                    "type": "integer",
                    "description": "Maximum lines to return (default: 200)"
                },
                "tail": {
                    "type": "boolean",
                    "description": (
                        "Read the LAST max_lines lines instead of the first. "
                        "Use to check how a long document ENDS — that it "
                        "closes with the References section you added, that a "
                        "log ends in success, that a file was not truncated."
                    ),
                },
                "search": {
                    "type": "string",
                    "description": (
                        "Case-insensitive regex. Returns every matching line "
                        "with its line number and one line of context either "
                        "side, plus the total match count — instead of the "
                        "file body. The right tool for 'is there a References "
                        "section', 'which lines cite Boettiger', 'did this log "
                        "raise'. Far cheaper than reading a long file, and it "
                        "answers presence/absence definitively."
                    ),
                },
                "offset": {
                    "type": "integer",
                    "description": (
                        "1-based line to start reading from — the way to read "
                        "the MIDDLE of a file. A truncated read lists the "
                        "section headings with their line numbers; pass one "
                        "here to jump straight there (e.g. offset=198 to read "
                        "a section beginning at line 198). Do NOT re-read from "
                        "the top hoping to see more — you will get the same "
                        "lines back."
                    ),
                },
            },
            required=["file_path"]
        )

        # =====================================================================
        # KNOWLEDGE DATA QUERY TOOL
        # =====================================================================

        QUERYABLE_EXTENSIONS = {'.xlsx', '.xls', '.csv'}

        DIR_DB_MIN_FILES = 10  # minimum same-extension files to treat as database

        def _summarize_json_value(value, max_str_len=200):
            """Return a compact string representation of a JSON value."""
            if value is None:
                return "null"
            if isinstance(value, bool):
                return str(value).lower()
            if isinstance(value, (int, float)):
                return str(value)
            if isinstance(value, str):
                if len(value) > max_str_len:
                    return json.dumps(value[:100]) + f"... ({len(value)} chars)"
                return json.dumps(value)
            if isinstance(value, list):
                if len(value) == 0:
                    return "list (0 items)"
                first = _summarize_json_value(value[0], max_str_len=80)
                return f"list ({len(value)} items, first: {first})"
            if isinstance(value, dict):
                keys = list(value.keys())
                return f"dict ({len(keys)} keys: {keys[:5]})"
            return repr(value)[:100]

        def _summarize_json_file(file_path: str) -> str:
            """Parse a JSON file and return a compact summary showing all keys."""
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, dict):
                lines = ["{"]
                for k, v in data.items():
                    lines.append(f"  {json.dumps(k)}: {_summarize_json_value(v)},")
                lines.append("}")
                return "\n".join(lines)
            elif isinstance(data, list):
                if len(data) == 0:
                    return "[]  (empty list)"
                summary = f"list of {len(data)} items\n"
                if isinstance(data[0], dict):
                    summary += "First item:\n"
                    lines = ["{"]
                    for k, v in data[0].items():
                        lines.append(f"  {json.dumps(k)}: {_summarize_json_value(v)},")
                    lines.append("}")
                    summary += "\n".join(lines)
                else:
                    summary += f"First item: {_summarize_json_value(data[0])}"
                return summary
            return repr(data)[:2000]

        from functools import lru_cache as _lru_cache

        @_lru_cache(maxsize=32)
        def _inspect_directory(dir_path: str) -> dict:
            """Summarize directory contents for LLM-driven querying.

            Memoized: per-orchestrator-instance cache keyed on `dir_path`.
            The directory listing is static for the lifetime of a planning
            session, so repeated qkd/screen_database calls on the same
            database reuse the inspection (saves ~1-3s per call on large
            directories). Callers must treat the return dict as read-only.
            """
            from collections import Counter

            p = Path(dir_path)
            files_by_ext = Counter()
            all_files = {}  # ext -> sorted list of filenames

            for f in p.iterdir():
                if f.is_file() and not f.name.startswith('.'):
                    ext = f.suffix.lower()
                    files_by_ext[ext] += 1
                    all_files.setdefault(ext, []).append(f.name)

            # Sort filenames for each extension
            for ext in all_files:
                all_files[ext].sort()

            total = sum(files_by_ext.values())

            # Pick dominant extension for sampling
            if not files_by_ext:
                return {
                    "directory": str(p),
                    "files_by_extension": {},
                    "total_files": 0,
                    "sample_files": [],
                    "all_filenames_sample": [],
                }

            # Sample files from each extension that has >= DIR_DB_MIN_FILES
            # For smaller groups, sample one file; this gives the LLM visibility
            # into all file types in the directory.
            sample_files = []  # list of {"name", "content", "ext"}
            sampled_exts = []
            for ext, count in files_by_ext.most_common():
                if count < DIR_DB_MIN_FILES:
                    continue
                sampled_exts.append(ext)
                names_for_ext = all_files[ext]
                # Pick one representative file per extension
                sample_name = names_for_ext[len(names_for_ext) // 2]
                fp = p / sample_name
                try:
                    if ext == '.json':
                        content = _summarize_json_file(str(fp))
                    else:
                        size = fp.stat().st_size
                        with open(fp, 'r', encoding='utf-8', errors='replace') as fh:
                            if size > 50_000:
                                content = fh.read(50_000)
                                content += f"\n... (truncated, {size} bytes total)"
                            else:
                                content = fh.read()
                except Exception as e:
                    content = f"(error reading file: {e})"
                sample_files.append({"name": sample_name, "content": content, "ext": ext})

            # Collect first 20 filenames from the most common extension
            dominant_ext = files_by_ext.most_common(1)[0][0]

            return {
                "directory": str(p),
                "files_by_extension": dict(files_by_ext.most_common()),
                "total_files": total,
                "extensions": sampled_exts,
                "sample_files": sample_files,
                "all_filenames_sample": all_files[dominant_ext][:20],
            }

        def _inspect_knowledge_file(file_path: str) -> dict:
            """Return a format-agnostic diagnostic snapshot of a data file."""
            p = Path(file_path)
            ext = p.suffix.lower()

            if ext in ('.xlsx', '.xls'):
                df = pd.read_excel(file_path, nrows=10)
                read_instr = f"pd.read_excel('{file_path}')"
                fmt = "excel"
            elif ext == '.csv':
                df = pd.read_csv(file_path, nrows=10)
                read_instr = f"pd.read_csv('{file_path}')"
                fmt = "csv"
            else:
                return {"error": f"Unsupported format: {ext}"}

            # Get actual row count without loading full file
            if ext in ('.xlsx', '.xls'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    total_rows = wb.active.max_row - 1
                    wb.close()
                except Exception:
                    total_rows = "unknown"
            else:
                with open(file_path) as f:
                    total_rows = sum(1 for _ in f) - 1

            return {
                "format": fmt,
                "shape": (total_rows, len(df.columns)),
                "columns": list(df.columns),
                "dtypes": df.dtypes.to_string(),
                "head": df.to_string(),
                "read_instruction": read_instr,
            }

        def _discover_queryable_files() -> list:
            """Scan knowledge directories for queryable data files and directory databases."""
            from collections import Counter

            search_dirs = set()
            # User-specified knowledge folder (or kb_storage/ default)
            if self.orch.knowledge_dir and Path(self.orch.knowledge_dir).exists():
                search_dirs.add(Path(self.orch.knowledge_dir))
            # Session knowledge dir (where UI uploads go)
            session_kdir = Path(self.orch.base_dir) / "knowledge"
            if session_kdir.exists():
                search_dirs.add(session_kdir)

            if not search_dirs:
                return []

            found = {}
            dir_db_dirs = set()  # directories detected as databases

            # --- Detect directory databases ---
            # Check each search dir and its immediate subdirectories
            dirs_to_check = list(search_dirs)
            for kdir in search_dirs:
                for child in kdir.iterdir():
                    if child.is_dir() and not child.name.startswith('.'):
                        dirs_to_check.append(child)

            for d in dirs_to_check:
                ext_counts = Counter()
                for f in d.iterdir():
                    if f.is_file() and not f.name.startswith('.'):
                        ext_counts[f.suffix.lower()] += 1
                # If any extension has enough files, treat whole directory as database
                db_exts = {ext: count for ext, count in ext_counts.items()
                           if count >= DIR_DB_MIN_FILES}
                if db_exts:
                    parts = ", ".join(f"{c} {e}" for e, c in sorted(db_exts.items()))
                    display_name = f"{d.name} ({parts} files)"
                    found[display_name] = {
                        "name": display_name,
                        "path": str(d),
                        "type": "directory",
                    }
                    dir_db_dirs.add(d)

            # --- Discover single queryable files ---
            for kdir in search_dirs:
                for f in kdir.rglob("*"):
                    if f.is_file() and f.suffix.lower() in QUERYABLE_EXTENSIONS:
                        # Skip files inside a detected database directory
                        if any(f.parent == dd for dd in dir_db_dirs):
                            continue
                        found[f.name] = {"name": f.name, "path": str(f), "type": "file"}

            return sorted(found.values(), key=lambda x: x["name"])

        def _resolve_knowledge_data_file(file_name: str):
            """Resolve a file name to a queryable target.

            An absolute / relative path to an existing file or directory is
            used directly — so a meta-delegated task, whose data lives outside
            the knowledge directory, works (mirroring read_file / analyze_file).
            Otherwise the name is matched within the knowledge directory.

            Returns (target, error) where target is either:
            - a file path string (for single files)
            - a dict with "type": "directory" (for directory databases)
            """
            from difflib import get_close_matches

            # Direct path: an existing file/directory given by path is used
            # as-is, without requiring knowledge-directory discovery.
            p = Path(file_name).expanduser()
            if p.is_file() and p.suffix.lower() in QUERYABLE_EXTENSIONS:
                return str(p.resolve()), None
            if p.is_dir():
                return {"name": p.name, "path": str(p.resolve()),
                        "type": "directory"}, None

            candidates = _discover_queryable_files()
            if not candidates:
                return None, json.dumps({
                    "status": "error",
                    "message": "No queryable data files or directories found in knowledge directory."
                })
            names = [c["name"] for c in candidates]
            entry_map = {c["name"]: c for c in candidates}

            def _return_entry(name):
                entry = entry_map[name]
                if entry.get("type") == "directory":
                    return entry, None
                return entry["path"], None

            # Exact match
            if file_name in entry_map:
                return _return_entry(file_name)

            # Stem match (without extension) — only for file entries
            stem = Path(file_name).stem
            for n in names:
                if Path(n).stem == stem:
                    return _return_entry(n)

            # Fuzzy match
            matches = get_close_matches(file_name, names, n=3, cutoff=0.5)
            if matches:
                suggestion = ", ".join(matches)
                return None, json.dumps({
                    "status": "error",
                    "message": f"'{file_name}' not found. Did you mean: {suggestion}?",
                    "available_files": names
                })

            return None, json.dumps({
                "status": "error",
                "message": f"'{file_name}' not found.",
                "available_files": names
            })

        def _extract_code_block(text: str) -> str:
            """Extract Python code from LLM response."""
            # Try ```python blocks
            match = re.search(r'```python\s*\n(.*?)```', text, re.DOTALL)
            if match:
                return match.group(1).strip()
            # Try generic ``` blocks
            match = re.search(r'```\s*\n(.*?)```', text, re.DOTALL)
            if match:
                return match.group(1).strip()
            # Prompt ends with open ```python — response may be code followed by ```
            match = re.search(r'^(.*?)```', text, re.DOTALL)
            if match and 'import ' in match.group(1):
                return match.group(1).strip()
            # Try raw code starting with import
            for line in text.split('\n'):
                if line.strip().startswith('import '):
                    idx = text.index(line)
                    # Strip any trailing ``` if present
                    code = text[idx:].strip()
                    code = re.sub(r'\n```\s*$', '', code)
                    return code
            # Last resort: if entire response looks like code (has import somewhere)
            if 'import ' in text and 'print(' in text:
                return re.sub(r'\n```\s*$', '', text.strip())
            return ""

        def _build_directory_scaffold(info: dict) -> str:
            """Build scaffold code with readers for each file type."""
            dir_path = info["directory"]
            lines = [
                "import json, glob, os",
                "from pathlib import Path",
                "",
                f"directory = {json.dumps(dir_path)}",
            ]
            # Add file lists and readers per extension
            reader_map = {
                '.json': (
                    "def read_json(filepath):\n"
                    "    with open(filepath, 'r') as f:\n"
                    "        return json.load(f)"
                ),
                '.csv': (
                    "import pandas as pd\n"
                    "def read_csv(filepath):\n"
                    "    return pd.read_csv(filepath)"
                ),
                '.tsv': (
                    "import pandas as pd\n"
                    "def read_tsv(filepath):\n"
                    "    return pd.read_csv(filepath, sep='\\t')"
                ),
            }
            default_reader = (
                "def read_text(filepath):\n"
                "    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:\n"
                "        return f.read()"
            )

            for ext in info["extensions"]:
                count = info["files_by_extension"].get(ext, 0)
                var_name = ext.lstrip('.') + "_files"
                lines.append(f'{var_name} = sorted(glob.glob(os.path.join(directory, "*{ext}")))')
                lines.append(f"# {count} {ext} files")
            lines.append("")

            added_readers = set()
            for ext in info["extensions"]:
                if ext in reader_map and ext not in added_readers:
                    lines.append(reader_map[ext])
                    added_readers.add(ext)
                elif ext not in added_readers:
                    lines.append(default_reader)
                    added_readers.add('_text')
            lines.append("")

            return "\n".join(lines)

        def _query_directory(dir_entry: dict, query: str) -> str:
            """Query a directory database using LLM-generated code."""
            import subprocess

            dir_path = dir_entry["path"]
            print(f"    - Inspecting directory: {dir_entry['name']}")

            try:
                info = _inspect_directory(dir_path)
            except Exception as e:
                return json.dumps({"status": "error", "message": f"Failed to inspect directory: {e}"})

            if not info["sample_files"]:
                return json.dumps({"status": "error", "message": "Directory is empty or unreadable."})

            # Build scaffold and sample sections for prompt
            scaffold = _build_directory_scaffold(info)
            sample_sections = []
            for s in info["sample_files"]:
                sample_sections.append(f"--- {s['name']} ({s['ext']}) ---\n{s['content']}")
            sample_text = "\nSAMPLE FILE CONTENTS:\n" + "\n\n".join(sample_sections) if sample_sections else ""

            prompt = KNOWLEDGE_QUERY_DIRECTORY_CODEGEN_PROMPT.format(
                directory=info["directory"],
                files_by_extension=info["files_by_extension"],
                total_files=info["total_files"],
                filenames=info["all_filenames_sample"],
                sample_sections=sample_text,
                scaffold=scaffold,
                query=query,
            )

            scripts_dir = Path(self.orch.base_dir) / "knowledge_query_scripts"
            scripts_dir.mkdir(parents=True, exist_ok=True)
            debug_dir = scripts_dir / "_debug"
            debug_dir.mkdir(parents=True, exist_ok=True)

            last_error = None
            for attempt in range(2):
                current_prompt = prompt
                if last_error:
                    current_prompt += f"\n\n**PREVIOUS ERROR:** {last_error}\nFix the script."

                try:
                    from scilink.knowledge import parse_json_from_response
                    response = self.orch.planner.model.generate_content(
                        [current_prompt],
                        generation_config={"max_output_tokens": 8192, "temperature": 0.0},
                    )
                    # Log raw response for debugging
                    raw_log_path = debug_dir / f"kq_dir_raw_{abs(hash(query)) % 10000:04d}_a{attempt}.txt"
                    raw_log_path.write_text(response.text)
                    # LLM returns JSON: {"code": "...TODO lines..."}
                    result, parse_error = parse_json_from_response(response)
                    if parse_error or not result or "code" not in result:
                        # Fallback: treat response as raw code
                        body = _extract_code_block(response.text) or response.text.strip()
                        body = re.sub(r'^```\w*\s*', '', body)
                        body = re.sub(r'\s*```$', '', body)
                    else:
                        body = result["code"]
                    code = (
                        f"{scaffold}\n"
                        f"{body}\n"
                        f"print(json.dumps({{\"answer\": answer, \"summary\": summary}}))\n"
                    )
                except Exception as e:
                    return json.dumps({"status": "error", "message": f"Code generation failed: {e}"})

                script_path = scripts_dir / f"kq_dir_{Path(dir_path).name}_{abs(hash(query)) % 10000:04d}.py"
                script_path.write_text(code)
                print(f"    - Running: {script_path.name} (attempt {attempt + 1})")

                try:
                    result = subprocess.run(
                        ["python", str(script_path)],
                        capture_output=True, text=True, timeout=120,
                    )
                    if result.returncode != 0:
                        last_error = result.stderr.strip()[-500:]
                        continue

                    # Parse the last valid JSON object from stdout
                    # (LLM may print extra output before the json.dumps line)
                    answer_data = None
                    for line in reversed(result.stdout.strip().splitlines()):
                        line = line.strip()
                        if line.startswith('{') and line.endswith('}'):
                            try:
                                answer_data = json.loads(line)
                                if "answer" in answer_data or "summary" in answer_data:
                                    break
                            except json.JSONDecodeError:
                                continue
                    if not answer_data:
                        last_error = f"No valid JSON in output: {result.stdout[-300:]}"
                        continue

                    answer_str = json.dumps(answer_data.get("answer", ""))
                    if len(answer_str) > 5000:
                        answer_data["answer"] = str(answer_data["answer"])[:5000] + "... (truncated)"

                    print(f"    - ✅ Directory query answered successfully.")
                    return json.dumps({
                        "status": "success",
                        "query": query,
                        "file": dir_entry["name"],
                        "answer": answer_data.get("answer"),
                        "summary": answer_data.get("summary", ""),
                        "details": answer_data.get("details"),
                        "script_path": str(script_path),
                    })

                except subprocess.TimeoutExpired:
                    last_error = "Script timed out (120s limit)."
                    continue
                except json.JSONDecodeError as e:
                    last_error = f"Invalid JSON in output: {e}"
                    continue

            return json.dumps({
                "status": "error",
                "message": "Directory query failed after 2 attempts.",
                "last_error": last_error,
            })

        def query_knowledge_data(query: str, file_name: str = None) -> str:
            """Query a knowledge data file or directory database with natural language."""
            import subprocess

            print(f"  ⚡ Tool: Querying knowledge data: '{query}'")

            # 1. Discover queryable files / directory databases. Informational
            #    only — an explicit `file_name` path is resolved directly below,
            #    so discovery being empty is not an error when a path is given.
            queryable = _discover_queryable_files()

            # 2. Resolve target
            if file_name is not None:
                target, error = _resolve_knowledge_data_file(file_name)
                if error:
                    return error
            elif not queryable:
                return json.dumps({
                    "status": "error",
                    "message": "No queryable data files or directories found by "
                               "discovery. Pass the file's absolute path as `file_name`."
                })
            elif len(queryable) == 1:
                target = queryable[0]
                print(f"    - Auto-selected: {target['name']}")
            else:
                return json.dumps({
                    "status": "file_selection_needed",
                    "message": "Multiple queryable sources found. Specify file_name.",
                    "available_files": [f["name"] for f in queryable]
                })

            # 2b. Branch: directory database vs single file
            if isinstance(target, dict) and target.get("type") == "directory":
                return _query_directory(target, query)

            # Single file path — extract path string
            target_path = target if isinstance(target, str) else target["path"]

            # 3. Inspect file
            try:
                info = _inspect_knowledge_file(target_path)
                if "error" in info:
                    return json.dumps({"status": "error", "message": info["error"]})
            except Exception as e:
                return json.dumps({"status": "error", "message": f"Failed to read file: {e}"})

            # 4. Build prompt
            prompt = KNOWLEDGE_QUERY_CODEGEN_PROMPT.format(
                file_path=target_path,
                file_format=info["format"],
                rows=info["shape"][0],
                cols=info["shape"][1],
                columns=info["columns"],
                dtypes=info["dtypes"],
                head=info["head"],
                read_instruction=info["read_instruction"],
                query=query,
            )

            # 5. Generate and execute (with 1 retry)
            scripts_dir = Path(self.orch.base_dir) / "knowledge_query_scripts"
            scripts_dir.mkdir(parents=True, exist_ok=True)

            last_error = None
            for attempt in range(2):
                current_prompt = prompt
                if last_error:
                    current_prompt += f"\n\n**PREVIOUS ERROR:** {last_error}\nFix the script."

                # Generate code
                try:
                    response = self.orch.planner.model.generate_content(
                        [current_prompt],
                        generation_config={"max_output_tokens": 1024, "temperature": 0.0},
                    )
                    # LLM returns JSON: {"code": "...TODO lines..."}
                    from scilink.knowledge import parse_json_from_response
                    result, parse_error = parse_json_from_response(response)
                    if parse_error or not result or "code" not in result:
                        # Fallback: treat response as raw code
                        body = _extract_code_block(response.text) or response.text.strip()
                        body = re.sub(r'^```\w*\s*', '', body)
                        body = re.sub(r'\s*```$', '', body)
                    else:
                        body = result["code"]
                    code = (
                        f"import pandas as pd, json\n"
                        f"df = {info['read_instruction']}\n"
                        f"{body}\n"
                        f"print(json.dumps({{\"answer\": answer, \"summary\": summary}}))\n"
                    )
                except Exception as e:
                    return json.dumps({"status": "error", "message": f"Code generation failed: {e}"})

                # Write and execute script
                script_path = scripts_dir / f"kq_{Path(target_path).stem}_{abs(hash(query)) % 10000:04d}.py"
                script_path.write_text(code)
                print(f"    - Running: {script_path.name} (attempt {attempt + 1})")

                try:
                    result = subprocess.run(
                        ["python", str(script_path)],
                        capture_output=True, text=True, timeout=60,
                    )
                    if result.returncode != 0:
                        last_error = result.stderr.strip()[-500:]
                        continue

                    json_match = re.search(r'\{.*\}', result.stdout.strip(), re.DOTALL)
                    if not json_match:
                        last_error = f"No JSON in output: {result.stdout[:300]}"
                        continue

                    answer_data = json.loads(json_match.group(0))

                    # Truncate large results
                    answer_str = json.dumps(answer_data.get("answer", ""))
                    if len(answer_str) > 5000:
                        answer_data["answer"] = str(answer_data["answer"])[:5000] + "... (truncated)"

                    print(f"    - ✅ Query answered successfully.")
                    return json.dumps({
                        "status": "success",
                        "query": query,
                        "file": Path(target_path).name,
                        "answer": answer_data.get("answer"),
                        "summary": answer_data.get("summary", ""),
                        "details": answer_data.get("details"),
                        "script_path": str(script_path),
                    })

                except subprocess.TimeoutExpired:
                    last_error = "Script timed out (60s limit)."
                    continue
                except json.JSONDecodeError as e:
                    last_error = f"Invalid JSON in output: {e}"
                    continue

            return json.dumps({
                "status": "error",
                "message": f"Query failed after 2 attempts.",
                "last_error": last_error,
            })

        self._register_tool(
            func=query_knowledge_data,
            name="query_knowledge_data",
            description=(
                "Query tabular data with natural language. Works with single "
                "data files (CSV, XLSX) and directory databases (folders of "
                "uniformly-structured files like JSON records). Generates and "
                "executes a Python script to answer questions about the data. "
                "Accepts a file by an absolute path (preferred when the data "
                "lives outside the knowledge directory) or by name."
            ),
            parameters={
                "query": {
                    "type": "string",
                    "description": "Natural language question about the data"
                },
                "file_name": {
                    "type": "string",
                    "description": (
                        "The data file to query, given as an absolute path to "
                        "an existing file or directory, or as a bare file name "
                        "to look up in the knowledge directory. If omitted, "
                        "lists available files."
                    )
                }
            },
            required=["query"]
        )

        # ─── screen_database — production filter+rank, follow-up to qkd ────
        def _thread_all_exploration_results(max_items: int = 20) -> str:
            """Collect ALL prior `query_knowledge_data` tool results from the
            session's chat history (regardless of which file/directory they
            targeted) and format them as a PRIOR EXPLORATION FINDINGS block
            for the screen_database codegen prompt. Returns "" when no
            exploration history is available."""
            try:
                msgs = getattr(self.orch, "messages", None) or []
                qkd_call_ids = set()
                hits = []
                for m in msgs:
                    role = m.get("role") if isinstance(m, dict) else None
                    if role == "assistant":
                        for tc in (m.get("tool_calls") or []):
                            fn = tc.get("function") if isinstance(tc, dict) else None
                            fn_name = fn.get("name") if isinstance(fn, dict) else None
                            tc_id = tc.get("id") if isinstance(tc, dict) else None
                            if fn_name == "query_knowledge_data" and tc_id:
                                qkd_call_ids.add(tc_id)
                    elif role == "tool" and m.get("tool_call_id") in qkd_call_ids:
                        try:
                            content = m.get("content", "")
                            payload = json.loads(content) if isinstance(content, str) else content
                        except (json.JSONDecodeError, TypeError):
                            continue
                        if not isinstance(payload, dict) or payload.get("status") != "success":
                            continue
                        hits.append({
                            "query":   payload.get("query", ""),
                            "file":    payload.get("file", ""),
                            "answer":  payload.get("answer"),
                            "summary": payload.get("summary", ""),
                        })
                if not hits:
                    return ""
                hits = hits[-max_items:]
                lines = [
                    "",
                    "PRIOR EXPLORATION FINDINGS (from earlier query_knowledge_data",
                    "calls; treat as ground truth about schema and value ranges —",
                    "do not re-discover):",
                    "",
                ]
                for i, r in enumerate(hits, 1):
                    ans_str = json.dumps(r["answer"], default=str)
                    if len(ans_str) > 1500:
                        ans_str = ans_str[:1500] + "... (truncated)"
                    lines.append(f"  [{i}] file: {r['file']}")
                    lines.append(f"      Q: {r['query']}")
                    lines.append(f"      A: {ans_str}")
                    if r["summary"]:
                        lines.append(f"      summary: {r['summary']}")
                    lines.append("")
                return "\n".join(lines)
            except Exception:
                # Defensive — never let exploration-context wiring break screening.
                return ""

        def screen_database(query: str, database_path: str,
                            top_k: int = 50, max_retries: int = 3) -> str:
            """Production database screening — filter + rank → top-K JSON."""
            import subprocess
            from scilink.executors import require_sandbox_approval

            print(f"  ⚡ Tool: screen_database  '{query}'  top_k={top_k}")

            if not require_sandbox_approval(
                interactive=True, allow_override=True,
                context=f"Database screening over {database_path}",
            ):
                return json.dumps({"status": "error",
                                   "message": "Sandbox approval denied — screening cannot run."})

            # Resolve to an absolute directory path (direct → fallback discovery).
            direct = Path(database_path).expanduser()
            if direct.is_dir():
                dir_path, dir_name = str(direct.resolve()), direct.name
            else:
                target, err = _resolve_knowledge_data_file(database_path)
                if err:
                    return err
                if isinstance(target, dict) and target.get("type") == "directory":
                    dir_path, dir_name = target["path"], target["name"]
                elif isinstance(target, str) and Path(target).is_dir():
                    dir_path, dir_name = target, Path(target).name
                else:
                    return json.dumps({
                        "status": "error",
                        "message": (f"screen_database requires a directory database; "
                                    f"got '{database_path}'. For single CSV/XLSX "
                                    f"files use query_knowledge_data."),
                    })

            print(f"    - Inspecting directory: {dir_name}")
            try:
                info = _inspect_directory(dir_path)
            except Exception as e:
                return json.dumps({"status": "error", "message": f"Failed to inspect directory: {e}"})
            if not info["sample_files"]:
                return json.dumps({"status": "error", "message": "Directory is empty or unreadable."})

            scaffold = _build_directory_scaffold(info)
            sample_sections = []
            for s in info["sample_files"]:
                sample_sections.append(f"--- {s['name']} ({s['ext']}) ---\n{s['content']}")
            sample_text = "\nSAMPLE FILE CONTENTS:\n" + "\n\n".join(sample_sections) if sample_sections else ""

            prior_context = _thread_all_exploration_results()

            prompt = SCREEN_DATABASE_CODEGEN_PROMPT.format(
                directory=info["directory"],
                files_by_extension=info["files_by_extension"],
                total_files=info["total_files"],
                filenames=info["all_filenames_sample"],
                sample_sections=sample_text,
                prior_exploration_block=prior_context,
                scaffold=scaffold,
                query=query,
                top_k=top_k,
            )

            scripts_dir = Path(self.orch.base_dir) / "knowledge_query_scripts"
            scripts_dir.mkdir(parents=True, exist_ok=True)
            debug_dir = scripts_dir / "_debug"
            debug_dir.mkdir(parents=True, exist_ok=True)
            safe_name = re.sub(r'[^A-Za-z0-9._-]+', '_', dir_name)
            hash_str  = f"{abs(hash((dir_name, query))) % 10000:04d}"
            script_path = scripts_dir / f"screen_{safe_name}_{hash_str}.py"
            result_path = scripts_dir / f"screen_{safe_name}_{hash_str}_result.json"

            last_error = None
            for attempt in range(max_retries):
                current_prompt = prompt
                if last_error:
                    current_prompt += (
                        f"\n\n**PREVIOUS ERROR (attempt {attempt}):** {last_error}\n"
                        "Fix the script. If PRIOR EXPLORATION FINDINGS are above, "
                        "re-read them — schema/field-name mistakes are the most common cause."
                    )

                try:
                    from scilink.knowledge import parse_json_from_response
                    # No max_output_tokens cap — screening codegen needs room for
                    # legitimate complexity (filter + score + multiprocessing + sort).
                    response = self.orch.planner.model.generate_content(
                        [current_prompt],
                        generation_config={"temperature": 0.0},
                    )
                    (debug_dir / f"screen_{safe_name}_{hash_str}_a{attempt}.txt").write_text(response.text)
                    parsed, parse_error = parse_json_from_response(response)
                    if parse_error or not parsed or "code" not in parsed:
                        body = _extract_code_block(response.text) or response.text.strip()
                        body = re.sub(r'^```\w*\s*', '', body)
                        body = re.sub(r'\s*```$', '', body)
                    else:
                        body = parsed["code"]
                    # Force `fork` start method so the LLM-generated multiprocessing.Pool()
                    # works on macOS without an `if __name__` guard (the alternative —
                    # spawn — re-imports the script and recursively spawns workers).
                    mp_preamble = (
                        "import multiprocessing as _scilink_mp\n"
                        "try: _scilink_mp.set_start_method('fork', force=True)\n"
                        "except (RuntimeError, ValueError): pass\n"
                    )
                    code = (
                        f"{scaffold}\n"
                        f"{mp_preamble}"
                        f"{body}\n"
                        f"import json as _json\n"
                        f"print(_json.dumps({{\n"
                        f"    \"n_scanned\":       n_scanned,\n"
                        f"    \"n_passed\":        len(results),\n"
                        f"    \"results\":         results[:{top_k}],\n"
                        f"    \"filters_applied\": filters_applied,\n"
                        f"    \"ranking_metric\":  ranking_metric,\n"
                        f"    \"summary\":         summary,\n"
                        f"}}))\n"
                    )
                except Exception as e:
                    return json.dumps({"status": "error", "message": f"Codegen failed: {e}"})

                script_path.write_text(code)
                print(f"    - Running: {script_path.name} (attempt {attempt + 1}/{max_retries})")

                try:
                    proc = subprocess.run(
                        ["python", str(script_path)],
                        capture_output=True, text=True, timeout=120,
                    )
                    if proc.returncode != 0:
                        last_error = proc.stderr.strip()[-800:]
                        continue
                    result_obj = None
                    for line in reversed(proc.stdout.strip().splitlines()):
                        line = line.strip()
                        if line.startswith('{') and line.endswith('}'):
                            try:
                                cand = json.loads(line)
                                if "n_scanned" in cand and "results" in cand:
                                    result_obj = cand
                                    break
                            except json.JSONDecodeError:
                                continue
                    if not result_obj:
                        last_error = f"No valid screening-result JSON in output: {proc.stdout[-400:]}"
                        continue

                    full = {
                        "query": query,
                        "database_path": dir_path,
                        "n_scanned": result_obj.get("n_scanned"),
                        "n_passed":  result_obj.get("n_passed"),
                        "filters_applied": result_obj.get("filters_applied"),
                        "ranking_metric":  result_obj.get("ranking_metric"),
                        "results":   result_obj.get("results", []),
                        "summary":   result_obj.get("summary", ""),
                        "exploration_context_used": bool(prior_context),
                        "top_k_requested": top_k,
                        "script_path": str(script_path),
                        "attempts_used": attempt + 1,
                    }
                    result_path.write_text(json.dumps(full, indent=2, default=str))
                    print(
                        f"    - ✅ Screened: {full['n_passed']} of {full['n_scanned']} "
                        f"passed; saved {result_path.name}"
                    )
                    preview = full["results"][:min(10, len(full["results"]))]
                    return json.dumps({
                        "status": "success",
                        "output_path": str(result_path),
                        "script_path": str(script_path),
                        "n_scanned": full["n_scanned"],
                        "n_passed":  full["n_passed"],
                        "filters_applied": full["filters_applied"],
                        "ranking_metric":  full["ranking_metric"],
                        "summary":   full["summary"],
                        "top_k_preview": preview,
                    })
                except subprocess.TimeoutExpired:
                    last_error = "Script timed out (120s limit)."
                    continue

            return json.dumps({
                "status": "error",
                "message": f"Screening failed after {max_retries} attempts.",
                "last_error": last_error,
                "script_path": str(script_path),
            })

        self._register_tool(
            func=screen_database,
            name="screen_database",
            description=(
                "Production database screening — filter + rank an entire "
                "directory database (folder of JSON / CSV records) by criteria "
                "expressed in natural language, persist the ranked top-K to a "
                "JSON file, and return a preview. Use as a follow-up to a few "
                "exploratory `query_knowledge_data` calls — ALL prior "
                "`query_knowledge_data` answers from this session are "
                "auto-threaded into the screening codegen prompt as schema "
                "ground truth, so call exploration FIRST. Scope: deterministic "
                "local-file filter / rank only. Sandbox-gated."
            ),
            parameters={
                "query": {
                    "type": "string",
                    "description": (
                        "Filter criteria + ranking metric in natural language. "
                        "Be explicit about numeric thresholds, the ranking "
                        "direction, and any composite scoring formula."
                    ),
                },
                "database_path": {
                    "type": "string",
                    "description": (
                        "Absolute path to the directory database to screen, or "
                        "the bare directory name if listed under the knowledge "
                        "directory."
                    ),
                },
                "top_k": {
                    "type": "integer",
                    "description": "Number of top-ranked entries to retain. Default 50.",
                },
            },
            required=["query", "database_path"],
        )

        # =====================================================================
        # KNOWLEDGE & SKILL TOOLS
        # =====================================================================

        # 12. SYNTHESIZE KNOWLEDGE
        def synthesize_knowledge(plan_ids: list, focus: str, synthesis_type: str = "reference") -> str:
            """
            Distill findings from completed planning iterations into reusable knowledge.
            The synthesized knowledge can be graduated into a skill for future campaigns.
            """
            from scilink.knowledge import synthesize_knowledge as _synthesize

            print(f"  ⚡ Tool: Synthesizing knowledge ({synthesis_type}) from {len(plan_ids)} plan iterations...")

            planner_state = self.orch.planner.state if self.orch.planner.state else {}
            plan_history = planner_state.get("plan_history", [])
            experimental_results = planner_state.get("experimental_results", [])
            feedback_history = planner_state.get("human_feedback_history", [])
            results = []
            missing_ids = []

            for pid in plan_ids:
                found = False
                # Collect ALL plan_history entries for this iteration (draft, refined, constraint-adjusted)
                matching_plans = [p for p in plan_history if str(p.get("iteration")) == str(pid)]

                if matching_plans:
                    parts = []
                    for plan in matching_plans:
                        stage = plan.get("stage", "Unknown")
                        parts.append(f"--- Stage: {stage} ---")
                        for exp in plan.get("proposed_experiments", []):
                            parts.append(f"Experiment: {exp.get('experiment_name', '')}")
                            parts.append(f"Hypothesis: {exp.get('hypothesis', '')}")
                            steps = exp.get("experimental_steps", [])
                            if steps:
                                parts.append(f"Steps: {'; '.join(steps)}")
                            parts.append(f"Justification: {exp.get('justification', '')}")
                            parts.append(f"Expected outcome: {exp.get('expected_outcome', '')}")

                    # Include experimental results/outcomes for this iteration
                    matching_results = [
                        r for r in experimental_results
                        if str(r.get("iteration")) == str(pid)
                    ]
                    for exp_result in matching_results:
                        data_summary = exp_result.get("data_summary", "")
                        if data_summary:
                            parts.append(f"--- Experimental Outcome (iteration {pid}) ---")
                            parts.append(data_summary)

                    # Collect human feedback entries relevant to this iteration
                    user_feedback_parts = []
                    for fb in feedback_history:
                        feedback_text = fb.get("feedback", "")
                        phase = fb.get("phase", "")
                        if feedback_text:
                            user_feedback_parts.append(f"[{phase}] {feedback_text}")

                    result_dict = {
                        "detailed_analysis": "\n".join(parts),
                        "analysis_id": f"plan_iter_{pid}",
                        "status": matching_plans[-1].get("stage", ""),
                    }
                    if user_feedback_parts:
                        result_dict["human_feedback"] = {
                            "user_feedback": "\n".join(user_feedback_parts)
                        }
                    results.append(result_dict)
                    found = True

                if not found:
                    missing_ids.append(pid)

            if missing_ids:
                available = sorted(set(
                    str(p.get("iteration")) for p in plan_history if p.get("iteration") is not None
                ))
                return json.dumps({
                    "status": "error",
                    "message": f"Plan iteration(s) not found: {missing_ids}",
                    "available_iterations": available
                })

            if not results:
                return json.dumps({
                    "status": "error",
                    "message": "No plan history available. Generate a plan first."
                })

            # Synthesize via the standalone function
            counter = len(self.orch.active_knowledge) + 1
            try:
                entry = _synthesize(
                    results, focus,
                    model=self.orch.planner.model,
                    knowledge_id=f"knowledge_{counter:03d}",
                    synthesis_type=synthesis_type,
                )
            except (ValueError, RuntimeError) as e:
                return json.dumps({"status": "error", "message": str(e)})

            entry["source_plans"] = plan_ids
            self.orch.active_knowledge.append(entry)

            # Save to disk
            knowledge_dir = self.orch.base_dir / "knowledge"
            knowledge_dir.mkdir(parents=True, exist_ok=True)
            knowledge_file = knowledge_dir / f"{entry['id']}.json"
            with open(knowledge_file, 'w') as f:
                json.dump(entry, f, indent=2)

            response = {
                "status": "success",
                "knowledge_id": entry["id"],
                "focus": focus,
                "synthesis_type": synthesis_type,
                "summary": entry["summary"],
                "key_findings": entry["key_findings"],
                "saved_to": str(knowledge_file),
                "note": "Use graduate_to_skill to convert this knowledge into a reusable domain skill."
            }

            # Check if any graduated skill is linked to knowledge with same focus
            for skill_name, source_ids in self.orch._graduated_skill_sources.items():
                for kid in source_ids:
                    for k in self.orch.active_knowledge:
                        if k.get("id") == kid and k.get("focus", "").lower() == focus.lower():
                            response["skill_update_suggested"] = skill_name
                            response["skill_update_note"] = (
                                f"Graduated skill '{skill_name}' is linked to knowledge "
                                f"with the same focus area. Consider calling update_skill."
                            )
                            break
                    if "skill_update_suggested" in response:
                        break
                if "skill_update_suggested" in response:
                    break

            return json.dumps(response)

        self._register_tool(
            func=synthesize_knowledge,
            name="synthesize_knowledge",
            description=(
                "Distill findings from completed planning iterations into reusable knowledge. "
                "Use when the user wants to capture learnings from plan iterations — e.g., "
                "experimental design patterns, parameter ranges that worked, or failure modes."
            ),
            parameters={
                "plan_ids": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of plan iteration numbers (as strings) to synthesize knowledge from"
                },
                "focus": {
                    "type": "string",
                    "description": "What to extract/learn (e.g., 'optimal cycling protocol for NMC811', 'catalyst screening workflow')"
                },
                "synthesis_type": {
                    "type": "string",
                    "enum": ["reference", "trend", "failure", "method"],
                    "description": (
                        "Type of synthesis: 'reference' (default), 'trend', 'failure', or 'method'"
                    )
                }
            },
            required=["plan_ids", "focus"]
        )

        # 13. LIST KNOWLEDGE
        def list_knowledge() -> str:
            """List all active knowledge entries."""
            print(f"  ⚡ Tool: Listing active knowledge...")

            if not self.orch.active_knowledge:
                return json.dumps({
                    "status": "success",
                    "message": "No active knowledge entries.",
                    "entries": []
                })

            entries = []
            for entry in self.orch.active_knowledge:
                entries.append({
                    "id": entry["id"],
                    "focus": entry["focus"],
                    "source_count": len(entry.get("source_plans", entry.get("source_analyses", []))),
                    "findings_count": len(entry.get("key_findings", [])),
                    "timestamp": entry.get("timestamp")
                })

            return json.dumps({
                "status": "success",
                "total_entries": len(entries),
                "entries": entries
            })

        self._register_tool(
            func=list_knowledge,
            name="list_knowledge",
            description="Show all active knowledge entries synthesized from planning iterations.",
            parameters={},
            required=[]
        )

        # 14. CLEAR KNOWLEDGE
        def clear_knowledge(knowledge_id: str = None) -> str:
            """Remove active knowledge entries. If knowledge_id is None, removes all."""
            print(f"  ⚡ Tool: Clearing knowledge...")

            knowledge_dir = self.orch.base_dir / "knowledge"

            if knowledge_id is None:
                count = len(self.orch.active_knowledge)
                self.orch.active_knowledge.clear()
                if knowledge_dir.exists():
                    for f in knowledge_dir.glob("knowledge_*.json"):
                        f.unlink()
                return json.dumps({
                    "status": "success",
                    "message": f"Cleared all {count} knowledge entries."
                })

            for i, entry in enumerate(self.orch.active_knowledge):
                if entry["id"] == knowledge_id:
                    self.orch.active_knowledge.pop(i)
                    knowledge_file = knowledge_dir / f"{knowledge_id}.json"
                    if knowledge_file.exists():
                        knowledge_file.unlink()
                    return json.dumps({
                        "status": "success",
                        "message": f"Removed knowledge entry: {knowledge_id}"
                    })

            return json.dumps({
                "status": "error",
                "message": f"Knowledge ID not found: {knowledge_id}"
            })

        self._register_tool(
            func=clear_knowledge,
            name="clear_knowledge",
            description=(
                "Remove active knowledge entries. Specify a knowledge_id to remove a "
                "specific entry, or omit to clear all knowledge."
            ),
            parameters={
                "knowledge_id": {
                    "type": "string",
                    "description": "ID of knowledge entry to remove (omit to clear all)"
                }
            },
            required=[]
        )

        # 15. GRADUATE TO SKILL
        def graduate_to_skill(knowledge_id: str, skill_name: str, domain: str = "planning") -> str:
            """
            Convert a knowledge entry into a reusable planning skill (.md file).
            The skill is automatically registered for use in subsequent plan generation.
            """
            from .instruct import (
                PLANNING_KNOWLEDGE_TO_SKILL_INSTRUCTIONS,
                PLANNING_SKILL_UPDATE_INSTRUCTIONS,
            )
            from scilink.skills._shared._graduation import graduate_to_skill_file

            print(f"  ⚡ Tool: Graduating knowledge '{knowledge_id}' to skill '{skill_name}'...")

            # Find the knowledge entry
            knowledge_entry = None
            for entry in self.orch.active_knowledge:
                if entry.get("id") == knowledge_id:
                    knowledge_entry = entry
                    break

            if knowledge_entry is None:
                return json.dumps({
                    "status": "error",
                    "message": f"Knowledge ID not found: {knowledge_id}"
                })

            # Collect source planning details
            planning_details_parts = []
            source_ids = knowledge_entry.get("source_plans", [])
            plan_history = self.orch.planner.state.get("plan_history", []) if self.orch.planner.state else []

            for pid in source_ids:
                for plan in plan_history:
                    if str(plan.get("iteration")) == str(pid):
                        parts = [f"### Plan Iteration: {pid} (Stage: {plan.get('stage', 'N/A')})"]
                        for exp in plan.get("proposed_experiments", []):
                            parts.append(f"Experiment: {exp.get('experiment_name', '')}")
                            parts.append(f"Hypothesis: {exp.get('hypothesis', '')}")
                            steps = exp.get("experimental_steps", [])
                            if steps:
                                parts.append(f"Steps: {'; '.join(steps[:10])}")
                            parts.append(f"Expected outcome: {exp.get('expected_outcome', '')}")
                        planning_details_parts.append("\n".join(parts))
                        break

            # Also include feedback history if available
            feedback_history = self.orch.planner.state.get("human_feedback_history", []) if self.orch.planner.state else []
            for fb in feedback_history:
                planning_details_parts.append(
                    f"### User Feedback ({fb.get('phase', 'unknown')}):\n{fb.get('feedback', '')}"
                )

            planning_details = "\n\n".join(planning_details_parts) if planning_details_parts else "No source planning details available."

            # Fold focus/summary/findings + source detail into a single
            # knowledge_entry; the shared helper renders it into {knowledge_text}.
            findings = knowledge_entry.get("key_findings", []) or []
            distill_entry = {
                "focus": knowledge_entry.get("focus", ""),
                "summary": knowledge_entry.get("summary", ""),
                "key_findings": "\n".join(f"- {f}" for f in findings),
                "source_planning_details": planning_details,
            }

            def _llm_call(prompt: str) -> str:
                response = self.orch.planner.model.generate_content(
                    contents=[prompt],
                    generation_config=None,
                    safety_settings=None,
                )
                return response.text if hasattr(response, "text") else str(response)

            # Write to the persistent store (~/.scilink/graduated_skills) via
            # the shared structured-JSON helper so the skill survives the
            # session and a pip upgrade. The helper auto-detects
            # create-vs-update by whether the bundle already exists.
            try:
                result = graduate_to_skill_file(
                    knowledge_entry=distill_entry,
                    skill_name=skill_name,
                    domain=domain,
                    llm_call=_llm_call,
                    fresh_template=PLANNING_KNOWLEDGE_TO_SKILL_INSTRUCTIONS,
                    update_template=PLANNING_SKILL_UPDATE_INSTRUCTIONS,
                )
            except Exception as e:
                return json.dumps({"status": "error", "message": f"Graduation failed: {e}"})

            skill_path = result["skill_path"]

            # Register the skill so the live session can use it immediately.
            self.orch.register_skill(str(skill_path))

            # Track the link
            self.orch._graduated_skill_sources[skill_name] = [knowledge_id]

            return json.dumps({
                "status": "success",
                "skill_name": skill_name,
                "skill_path": str(skill_path),
                "method": result.get("method"),
                "source_knowledge_id": knowledge_id,
                "note": f"Skill '{skill_name}' has been registered (persistent memory) and will be applied to future plan generation."
            })

        self._register_tool(
            func=graduate_to_skill,
            name="graduate_to_skill",
            description=(
                "Convert a knowledge entry into a reusable planning skill (.md file). "
                "The skill is organized into 5 sections (overview, planning, implementation, "
                "interpretation, validation) and automatically registered for use in "
                "subsequent plan generation."
            ),
            parameters={
                "knowledge_id": {
                    "type": "string",
                    "description": "ID of the knowledge entry to graduate"
                },
                "skill_name": {
                    "type": "string",
                    "description": "Name for the new skill (used as filename and reference)"
                },
                "domain": {
                    "type": "string",
                    "description": "Domain area (default: 'planning')"
                }
            },
            required=["knowledge_id", "skill_name"]
        )

        # 16. UPDATE SKILL
        def update_skill(skill_name: str, knowledge_ids: list = None, domain: str = "planning") -> str:
            """
            Update a graduated skill with new knowledge entries.
            Merges into the persistent skill bundle in place.
            """
            from .instruct import (
                PLANNING_KNOWLEDGE_TO_SKILL_INSTRUCTIONS,
                PLANNING_SKILL_UPDATE_INSTRUCTIONS,
            )
            from scilink.skills._shared._graduation import graduate_to_skill_file
            from scilink.skills.loader import graduated_skills_dir

            print(f"  ⚡ Tool: Updating skill '{skill_name}'...")

            # The skill must already exist in the persistent store so the
            # helper takes its merge (update) branch.
            skill_path = graduated_skills_dir() / domain / skill_name / f"{skill_name}.md"
            if not skill_path.exists():
                return json.dumps({
                    "status": "error",
                    "message": f"Graduated skill not found: {domain}/{skill_name}"
                })

            # Determine source knowledge IDs
            tracked_ids = self.orch._graduated_skill_sources.get(skill_name, [])
            if knowledge_ids:
                new_ids = knowledge_ids
            else:
                # Use all knowledge entries with matching focus
                focus_areas = set()
                for kid in tracked_ids:
                    for k in self.orch.active_knowledge:
                        if k.get("id") == kid:
                            focus_areas.add(k.get("focus", "").lower())
                new_ids = [
                    k["id"] for k in self.orch.active_knowledge
                    if k["id"] not in tracked_ids and k.get("focus", "").lower() in focus_areas
                ]

            if not new_ids:
                return json.dumps({
                    "status": "error",
                    "message": "No new knowledge entries found to update the skill with."
                })

            # Collect new knowledge texts
            new_knowledge_parts = []
            for kid in new_ids:
                for k in self.orch.active_knowledge:
                    if k.get("id") == kid:
                        part = f"### {kid}\n**Focus:** {k.get('focus', '')}\n"
                        part += f"**Summary:** {k.get('summary', '')}\n"
                        part += "**Key Findings:**\n"
                        for f in k.get("key_findings", []):
                            part += f"- {f}\n"
                        new_knowledge_parts.append(part)
                        break

            new_knowledge = "\n\n".join(new_knowledge_parts)

            def _llm_call(prompt: str) -> str:
                response = self.orch.planner.model.generate_content(
                    contents=[prompt],
                    generation_config=None,
                    safety_settings=None,
                )
                return response.text if hasattr(response, "text") else str(response)

            # Delegate to the shared helper; since the bundle exists it
            # takes the structured-JSON merge (update) branch.
            try:
                result = graduate_to_skill_file(
                    knowledge_entry={"new_knowledge": new_knowledge},
                    skill_name=skill_name,
                    domain=domain,
                    llm_call=_llm_call,
                    fresh_template=PLANNING_KNOWLEDGE_TO_SKILL_INSTRUCTIONS,
                    update_template=PLANNING_SKILL_UPDATE_INSTRUCTIONS,
                )
            except Exception as e:
                return json.dumps({"status": "error", "message": f"Update failed: {e}"})

            # Update source tracking
            all_ids = list(set(tracked_ids + new_ids))
            self.orch._graduated_skill_sources[skill_name] = all_ids

            # Re-register the skill
            self.orch.register_skill(str(result["skill_path"]))

            return json.dumps({
                "status": "success",
                "skill_name": skill_name,
                "skill_path": str(result["skill_path"]),
                "method": result.get("method"),
                "new_knowledge_ids": new_ids,
                "total_source_ids": all_ids,
                "note": f"Skill '{skill_name}' has been updated in persistent memory."
            })

        def write_technical_document(
            request: str,
            filename: str = None,
            title: str = None,
            source_files: str = None,
            use_literature: bool = True,
            revise_path: str = None,
        ):
            """Author a grounded technical document and save it.

            Not a plan: no experiment schema, no campaign state, no plan
            report. A roadmap or estimate that went through the plan tool
            came back with a build sequence as its `hypothesis` and invented
            optimization ranges (live, cdoc facility roadmap).
            """
            try:
                planner = self.orch.planner
                # REVISION: read the document, rewrite it whole, put it back
                # where it was. Without this, "revise the paper you wrote"
                # authored into the CURRENT delegation folder while the
                # original sat untouched — and the agent then tried to reach
                # it with `../../`, which save_file's sandbox collapses to a
                # subfolder name, so it rebuilt the file by hand in chunks
                # and truncated it doing so (live).
                from .user_interface import format_path, record_deliverable
                current = None
                if revise_path:
                    rp = Path(revise_path)
                    if not rp.is_absolute():
                        rp = self._output_dir() / rp
                    rp = rp.resolve()
                    root = Path(self.orch.base_dir).resolve()
                    if root not in rp.parents:
                        return json.dumps({
                            "status": "error",
                            "message": (f"revise_path must be inside the "
                                        f"session directory ({root}).")})
                    if not rp.exists():
                        return json.dumps({
                            "status": "error",
                            "message": f"No such document: {rp}"})
                    current = rp.read_text(errors="replace")
                lit = None
                if use_literature:
                    _lit = self._load_campaign_literature()
                    if _lit is not None:
                        lit = _lit["text"]

                # Prior documents the agent names — this is how a revision or
                # a merge builds on what the session already wrote instead of
                # re-deriving it.
                sources, missing = [], []
                for raw in (source_files or "").split(","):
                    raw = raw.strip()
                    if not raw:
                        continue
                    fp = Path(raw)
                    if not fp.is_absolute():
                        fp = self._output_dir() / raw
                    if fp.exists():
                        sources.append(f"### {fp.name}\n{fp.read_text()}")
                    else:
                        # The agent names its own earlier file; that file
                        # lives in a SIBLING delegation directory, so search
                        # the session root by basename before giving up.
                        hits = sorted(Path(self.orch.base_dir).rglob(fp.name))
                        if hits:
                            sources.append(f"### {hits[0].name}\n"
                                           + hits[0].read_text())
                        else:
                            missing.append(raw)

                # A revision inherits the document's OWN name: falling back
                # to the request titled the deliverable with the instruction
                # ("Revise this brief with two targeted changes...") and
                # replaced its real name in the files list. Live.
                doc_title = title
                if not doc_title and current:
                    m = re.search(r"^#\s+(.+)$", current, re.M)
                    doc_title = (m.group(1).strip() if m
                                 else rp.stem.replace("_", " "))
                doc_title = doc_title or (request[:70].strip()
                                          or "Technical document")
                result = author_technical_document(
                    request=request,
                    kb_docs=planner.kb_docs,
                    model=planner.model,
                    generation_config=planner.generation_config,
                    external_context=lit,
                    source_documents=("\n\n".join(sources) if sources else None),
                    skill_context=planner._build_skill_context("planning"),
                    revise_document=current,
                    task_name=("Technical Document (revision)" if current
                               else "Technical Document"),
                )
                if result.get("error"):
                    return json.dumps({"status": "error",
                                       "message": result["error"]})
                sections = result.get("sections") or []
                if not sections:
                    return json.dumps({
                        "status": "error",
                        "message": "The author returned no sections."})

                text = document_to_markdown(doc_title, sections)
                if revise_path:
                    out = rp
                    # Revising in place crosses delegation isolation, which
                    # exists so a reused child cannot clobber earlier outputs
                    # by accident. An explicit revision is not that — but the
                    # audit trail still matters, so the delegation MAKING the
                    # change keeps the version it replaced. The canonical file
                    # stays canonical; the evidence lives where the edit
                    # happened.
                    bak = None
                    try:
                        d = self._output_dir()
                        d.mkdir(parents=True, exist_ok=True)
                        # Never clobber an earlier backup: revising the same
                        # document twice from one delegation would otherwise
                        # leave only the second-to-last version, and the
                        # original — the one the user actually approved —
                        # would be the copy that vanished.
                        n, bak = 1, d / f"{rp.stem}.before_revision{rp.suffix}"
                        while bak.exists():
                            n += 1
                            bak = d / f"{rp.stem}.before_revision{n}{rp.suffix}"
                        bak.write_text(current or "")
                        # Listed (not starred) so the replaced version shows
                        # up in the files block rather than only on disk.
                        # Under the meta this is the delegation slug; in a
                        # standalone session it is just the session dir, and
                        # "revised by session" says nothing.
                        who = ("" if d.resolve() == root
                               else f" (revised by {d.name})")
                        record_deliverable(
                            self.orch.base_dir, bak,
                            f"Pre-revision copy of {rp.name}{who}")
                    except Exception as e:  # noqa: BLE001 - never block the edit
                        logging.warning(f"Pre-revision copy failed: {e}")
                        bak = None
                    # A revision that came back shorter than the original is
                    # nearly always the model summarising instead of
                    # revising. Refuse rather than overwrite the good copy.
                    if len(text) < 0.5 * len(current or ""):
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"Revision aborted: the rewritten document is "
                                f"{len(text)} chars against the original's "
                                f"{len(current)}. A revision must return the "
                                "WHOLE document with untouched sections "
                                "verbatim. The original is unchanged — retry, "
                                "reproducing every section."),
                        })
                else:
                    name = filename or "technical_document.md"
                    if not name.endswith(".md"):
                        name += ".md"
                    out = self._output_dir() / name
                out.parent.mkdir(parents=True, exist_ok=True)
                # A roadmap or staging memo carries the same campaign flow a
                # white paper does. Skipped on a revision: the document
                # already has its figure, and a second pass would append a
                # duplicate section.
                if not revise_path:
                    text = self._maybe_embed_workflow_diagram(
                        text, out.parent, stem=f"{out.stem}_workflow")
                out.write_text(text)
                # A revised document's exported PDF twin (the white paper's
                # forwarded copy) must not keep serving the pre-revision
                # content; re-export is deterministic, so it does not touch
                # the content freeze a revision may be operating under.
                pdf_refreshed = self._refresh_pdf_twin(out) if revise_path \
                    else False

                record_deliverable(self.orch.base_dir, out, doc_title,
                                   deliverable=True)
                if revise_path:
                    # The transcript is the first place anyone looks, so the
                    # cross-delegation edit is named there, not just on disk.
                    print(f"    ✏️  Revised IN PLACE: {format_path(out)}")
                    if bak:
                        print(f"    ↩️  Previous version kept: "
                              f"{format_path(bak)}")
                else:
                    print(f"    📄 Document saved: {format_path(out)}")
                res = {"status": "success", "path": str(out),
                       "revised_in_place": bool(revise_path),
                       "previous_version": (str(bak) if revise_path and bak
                                            else None),
                       "revised_by": (self._output_dir().name if revise_path
                                      else None),
                       "pdf_refreshed": pdf_refreshed,
                       "title": doc_title,
                       "sections": [s.get("heading") for s in sections
                                    if isinstance(s, dict)],
                       "words": len(text.split()),
                       "literature_used": bool(lit),
                       "sources_used": len(sources)}
                if missing:
                    res["source_files_not_found"] = missing
                return json.dumps(res)
            except Exception as e:
                logging.error(f"Document authoring error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=write_technical_document,
            name="write_technical_document",
            description=(
                "Author a grounded technical DOCUMENT and save it as the "
                "deliverable: a roadmap, staging or build plan, cost or "
                "footprint estimate, consolidation memo, brief, summary or "
                "review. USE THIS — not generate_initial_plan — whenever the "
                "user says 'plan' in the everyday sense of a course of "
                "action ('plan how we build the facility', 'outline the "
                "stages', 'estimate the space we need'). "
                "generate_initial_plan is only for an EXPERIMENT: a testable "
                "hypothesis with measurements. Grounds the document in the "
                "campaign's literature and in prior session documents you "
                "name, so you get retrieval-backed authoring rather than "
                "writing it unaided into save_file. Use save_file for short "
                "notes and for content you have already composed. "
                "NOT for an EXPERIMENTAL PROTOCOL: a procedure with a "
                "hypothesis and measurements — including 'the runnable bench "
                "protocol for direction X' — is generate_initial_plan with "
                "selection_profile='lab', which gives it conformance "
                "checking, the critic, and refinement against results later. "
                "A document cannot be refined with results. NOT for editing "
                "a research portfolio either — 'drop the weakest direction', "
                "'harden this one', 'consolidate these' all change the "
                "portfolio itself and belong in refine_portfolio; a document "
                "about the revised portfolio leaves the portfolio unrevised."
            ),
            parameters={
                "request": {
                    "type": "string",
                    "description": (
                        "What the document must cover, in full — the user's "
                        "ask plus any structure, constraints, audience or "
                        "sections it must have. This is the authoring brief."
                    ),
                },
                "filename": {
                    "type": "string",
                    "description": (
                        "File to write, e.g. 'build_roadmap.md'. Defaults to "
                        "technical_document.md; give it a descriptive name."
                    ),
                },
                "title": {
                    "type": "string",
                    "description": (
                        "Document title, also the label shown beside the "
                        "file, e.g. 'cdoc staged build roadmap'."
                    ),
                },
                "source_files": {
                    "type": "string",
                    "description": (
                        "Comma-separated files this document should build on "
                        "— a roadmap you are revising, two documents you are "
                        "merging. Names of files in the session are resolved "
                        "for you. Omit for a fresh document."
                    ),
                },
                "use_literature": {
                    "type": "boolean",
                    "description": (
                        "Ground in the campaign's most recent literature "
                        "search (default true). Set false for a document "
                        "that is purely internal, e.g. merging two documents "
                        "you already wrote."
                    ),
                },
                "revise_path": {
                    "type": "string",
                    "description": (
                        "REVISE an existing document IN PLACE: pass its path "
                        "(any document in this session, including one written "
                        "by an earlier delegation). The file is read, the "
                        "whole document is rewritten with your change "
                        "applied, and it is written back over the SAME path — "
                        "so use this for 'add references to the paper you "
                        "wrote', 'tighten section 3', 'split this in two'. "
                        "Do NOT rebuild an existing document by hand with "
                        "save_file/append_file chunks: that writes into the "
                        "current delegation folder instead, and a save_file "
                        "call truncates what is already there. Omit for a new "
                        "document; `filename` is ignored when this is set."
                    ),
                },
            },
            required=["request"]
        )

        def generate_ideation_portfolio(
            specific_objective: str,
            knowledge_paths: str = None,
            additional_context: str = None,
            skill: str = None,
            literature_context: str = None,
            n_candidates: int = None,
            white_paper: bool = None,
            new_campaign: bool = None,
            selection_profile: str = None,   # accepted, ignored — see below
        ):
            """Author a PORTFOLIO of research directions (ideation).

            The ideation twin of generate_initial_plan: same grounding,
            best-of-N, judge, critic and campaign machinery, different
            contract — directions, not a bench protocol.
            """
            # `selection_profile` is NOT in this tool's schema, but its
            # description has to name the lab profile to say where a chosen
            # direction goes next — which puts the parameter in scope, and a
            # model duly passed it (live). The profile is meaningless here:
            # a portfolio is ideation by construction. Accept and ignore
            # rather than spend a round trip on a TypeError.
            if selection_profile:
                logging.info("generate_ideation_portfolio ignoring "
                             "selection_profile=%r (always ideation)",
                             selection_profile)
            # Same implementation, different contract — see `kind`.
            return generate_initial_plan(
                specific_objective=specific_objective,
                knowledge_paths=knowledge_paths,
                additional_context=additional_context,
                skill=skill,
                literature_context=literature_context,
                n_candidates=n_candidates,
                selection_profile="ideation",
                white_paper=white_paper,
                new_campaign=new_campaign,
                kind="portfolio",
            )

        self._register_tool(
            func=generate_ideation_portfolio,
            name="generate_ideation_portfolio",
            description=(
                "Generate a PORTFOLIO of research directions — brainstorming, "
                "ideation, 'what should we work on', a slate of use cases, a "
                "hedge against one direction failing, or consolidating earlier "
                "threads into a standalone class. Each direction carries its "
                "own hypothesis, rationale and novelty; the portfolio carries "
                "an organizing thesis. USE THIS INSTEAD OF "
                "generate_initial_plan whenever the ask is which directions "
                "are worth pursuing rather than how to run one on the bench — "
                "generate_initial_plan designs a lab experiment, and a "
                "portfolio forced into that schema comes back with its "
                "directions flattened into protocol steps. Produces a "
                "sponsor-facing white paper by default, and an all-candidates "
                "dossier when several candidate portfolios were generated. "
                "Once the user PICKS a direction and wants the runnable "
                "bench protocol for it, that is generate_initial_plan with "
                "selection_profile='lab' — an experimental plan, NOT a "
                "written document."
            ),
            parameters={
                "specific_objective": {
                    "type": "string",
                    "description": (
                        "What to ideate on, in full — the domain, what makes "
                        "a direction interesting here, any structure the user "
                        "asked for (how many directions, how to rank or "
                        "cluster them, elements each must carry)."
                    ),
                },
                "knowledge_paths": {
                    "type": "string",
                    "description": (
                        "Comma-separated paths to papers/reports/docs "
                        "folders — for document CORPORA too large to read "
                        "directly (triggers a full embedding KB build). "
                        "Documents already read this session belong in "
                        "additional_context, not here."),
                },
                "additional_context": {
                    "type": "string",
                    "description": "Constraints, prior findings, or user preferences to honour",
                },
                "skill": {
                    "type": "string",
                    "description": "Optional planning skill to load for domain guidance",
                },
                "literature_context": {
                    "type": "string",
                    "description": (
                        "Path to a literature search file to ground the "
                        "portfolio in (from search_literature)."
                    ),
                },
                "n_candidates": {
                    "type": "integer",
                    "description": (
                        "How many DISTINCT candidate portfolios to author "
                        "before the judge picks one. Defaults to best-of-3 on "
                        "a campaign's first portfolio and 1 for follow-ups. "
                        "Only a multi-candidate run produces the "
                        "all-candidates dossier, since that dossier IS the "
                        "report over the candidate set."
                    ),
                },
                "white_paper": {
                    "type": "boolean",
                    "description": (
                        "Force (true) or suppress (false) the sponsor-facing "
                        "white paper. Omit for the default, which produces one."
                    ),
                },
                "new_campaign": {
                    "type": "boolean",
                    "description": (
                        "TRUE when this starts a NEW line of work unrelated to "
                        "the current campaign, so its literature and history "
                        "do not leak in. Omit to continue the current campaign."
                    ),
                },
            },
            required=["specific_objective"]
        )

        def refine_portfolio(request: str, literature_context: str = None,
                             additional_context: str = None):
            """Revise a research PORTFOLIO: harden, drop, add, re-rank,
            consolidate.

            Same refinement engine as refine_plan_with_results — which is
            reachable but not FINDABLE for this: its name promises results,
            and dropping a direction has none. Live, "drop the weakest
            direction" went to write_technical_document instead, which wrote
            a document ABOUT the revised portfolio and left the portfolio
            itself untouched.
            """
            return refine_plan_with_results(
                result_data=request,
                use_literature_rag=False,
                literature_context=literature_context,
                additional_context=additional_context,
            )

        self._register_tool(
            func=refine_portfolio,
            name="refine_portfolio",
            description=(
                "Revise the current research PORTFOLIO in place. Use for any "
                "edit to the set of directions, none of which needs "
                "experimental results: HARDEN one (sharpen its hypothesis, "
                "name its failure mode), DROP one, ADD one, RE-RANK them, or "
                "CONSOLIDATE several into a single class. Directions the "
                "request does not touch are preserved verbatim. This is the "
                "tool for 'drop the weakest', 'harden class 2', 'consolidate "
                "these threads into one class' — do NOT write a document "
                "about the revised portfolio instead, and do not re-author "
                "the whole portfolio with generate_ideation_portfolio, which "
                "would lose the directions you are keeping. Use "
                "refine_plan_with_results instead when actual experimental "
                "results are what is driving the change."
            ),
            parameters={
                "request": {
                    "type": "string",
                    "description": (
                        "What to change, naming the directions by id where "
                        "the user did — e.g. 'harden DIR-1: sharpen its "
                        "hypothesis and name its dominant failure mode; "
                        "leave the others unchanged' or 'drop DIR-4 and "
                        "re-rank the rest by feasibility'."
                    ),
                },
                "literature_context": {
                    "type": "string",
                    "description": "Optional path to a literature file to ground the revision in.",
                },
                "additional_context": {
                    "type": "string",
                    "description": "Constraints or preferences to honour in the revision.",
                },
            },
            required=["request"]
        )

        self._register_tool(
            func=update_skill,
            name="update_skill",
            description=(
                "Update a graduated skill with new knowledge entries. "
                "Use when new knowledge has been synthesized and a linked skill "
                "should incorporate the new findings. The old version is preserved."
            ),
            parameters={
                "skill_name": {
                    "type": "string",
                    "description": "Name of the graduated skill to update"
                },
                "knowledge_ids": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Specific knowledge IDs to incorporate (omit to auto-detect from matching focus area)"
                },
                "domain": {
                    "type": "string",
                    "description": "Domain the skill was graduated under (default: 'planning')"
                }
            },
            required=["skill_name"]
        )
    
    def _register_tool(self, func: Callable, name: str, description: str, 
                      parameters: Dict[str, Any], required: list = None):
        """
        Register a tool in both OpenAI and Gemini formats.
        
        Args:
            func: The Python function to call
            name: Function name
            description: What the function does
            parameters: Dict of parameter definitions
            required: List of required parameter names
        """
        # Add to function map for execution
        self.functions_map[name] = func
        
        # Add to Gemini format (just the function object)
        self.gemini_functions.append(func)
        
        # Build OpenAI schema
        openai_schema = {
            "type": "function",
            "function": {
                "name": name,
                "description": description,
                "parameters": {
                    "type": "object",
                    "properties": parameters,
                    "required": required or []
                }
            }
        }
        self.openai_schemas.append(openai_schema)

    def _update_skill_description(self, custom_skills: dict = None) -> None:
        """Refresh the ``skill`` parameter description in ``generate_initial_plan``.

        Called after a skill is registered at runtime (e.g. ``graduate_to_skill``)
        so newly available skills become visible to the orchestrator LLM. The
        schema dict is mutated in place, so the change propagates to
        ``tools_for_model`` (which is the same ``openai_schemas`` list object).
        """
        new_desc = _build_planning_skill_description(custom_skills)
        for schema in self.openai_schemas:
            fn = schema.get("function", {})
            if fn.get("name") != "generate_initial_plan":
                continue
            skill_prop = fn.get("parameters", {}).get("properties", {}).get("skill")
            if skill_prop is not None:
                skill_prop["description"] = new_desc
            break

    def execute_tool(self, tool_name: str, **kwargs) -> str:
        """
        Execute a tool by name with given arguments.
        
        Args:
            tool_name: Name of the tool to execute
            **kwargs: Arguments to pass to the tool
            
        Returns:
            JSON string with result
        """
        if tool_name not in self.functions_map:
            return json.dumps({
                "status": "error",
                "message": f"Tool '{tool_name}' not found in registry"
            })

        # A tool call whose arguments hit the output-token cap mid-generation
        # can arrive as VALID but incomplete JSON (later keys dropped), which
        # slips past the malformed-JSON guard in the chat loop. Check the
        # schema's required list here so the model gets a clean, actionable
        # error instead of a TypeError traceback.
        missing = [p for p in self._required_params(tool_name) if p not in kwargs]
        if missing:
            return json.dumps({
                "status": "error",
                "tool": tool_name,
                "message": (
                    f"Missing required argument(s): {', '.join(missing)}. "
                    "The tool-call arguments were likely truncated by the "
                    "response length limit — resend the call with all required "
                    "arguments, splitting long text across multiple smaller "
                    "calls (e.g. save_file then append_file chunks)."
                ),
            })

        try:
            return self.functions_map[tool_name](**kwargs)
        except TypeError as e:
            # An unexpected keyword reaches here as a bare TypeError whose
            # message names the offending argument but not the valid ones,
            # so the model has to guess its way back. Name them.
            if "unexpected keyword argument" in str(e):
                import inspect as _inspect
                try:
                    accepted = [n for n in _inspect.signature(
                        self.functions_map[tool_name]).parameters]
                except (TypeError, ValueError):
                    accepted = []
                logging.warning(f"Tool {tool_name}: {e}")
                return json.dumps({
                    "status": "error",
                    "tool": tool_name,
                    "message": (f"{e}. Accepted arguments: "
                                f"{', '.join(accepted) or 'unknown'}. "
                                "Re-send the call using only those."),
                })
            logging.error(f"Tool execution error ({tool_name}): {e}", exc_info=True)
            return json.dumps({
                "status": "error",
                "message": str(e),
                "tool": tool_name
            })
        except Exception as e:
            logging.error(f"Tool execution error ({tool_name}): {e}", exc_info=True)
            return json.dumps({
                "status": "error",
                "message": str(e),
                "tool": tool_name
            })

    def _required_params(self, tool_name: str) -> list:
        """Return the schema-declared required parameter names for a tool."""
        for schema in self.openai_schemas:
            fn = schema.get("function", {})
            if fn.get("name") == tool_name:
                return fn.get("parameters", {}).get("required", []) or []
        return []


